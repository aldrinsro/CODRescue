from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.db.models import Count, Q, Sum, F, Avg
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.core.paginator import Paginator
import json
from parametre.models import Operateur, Ville
from commande.models import Commande, EtatCommande, EnumEtatCmd, Operation, Panier, Envoi
from django.urls import reverse
import barcode
from barcode.writer import ImageWriter
from io import BytesIO
import base64
import qrcode
import csv
from article.models import Article, MouvementStock, VarianteArticle
from commande.models import Envoi
from .forms import ArticleForm, AjusterStockForm
from .utils import creer_mouvement_stock
from .decorators import superviseur_preparation_required, superviseur_only_required, preparation_team_required
from article.models import Article, Promotion, VarianteArticle, Categorie, Genre, Couleur, Pointure
from django.views.decorators.http import require_POST
from article.forms import PromotionForm
from decimal import Decimal


@superviseur_preparation_required
def home_view(request):
    """Tableau de bord de supervision avec métriques globales et suivi en temps réel"""
    try:
        operateur_profile = request.user.profil_operateur
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('login')

    # Date de référence
    today = timezone.now().date()
    # Début de la semaine (lundi)
    start_of_week = today - timedelta(days=today.weekday())

    # Récupérer les états nécessaires
    try:
        etat_confirmee = EnumEtatCmd.objects.get(libelle__iexact='Confirmée')
        etat_en_preparation = EnumEtatCmd.objects.get(libelle__iexact='En préparation')
        etat_preparee = EnumEtatCmd.objects.get(libelle__iexact='Préparée')
    except EnumEtatCmd.DoesNotExist as e:
        messages.error(request, f"État manquant dans le système: {str(e)}")
        return redirect('login')


    # 3. Commandes livrées partiellement (supervision)
    commandes_livrees_partiellement = Commande.objects.filter(
        etats__enum_etat__libelle='Livrée Partiellement',
        etats__date_fin__isnull=True

    ).distinct().count()
    commandes_retournees = Commande.objects.filter(
        etats__enum_etat__libelle='Retournée',
        etats__date_fin__isnull=True
    ).distinct().count()

    commandes_en_cours = Commande.objects.filter(
        etats__enum_etat=etat_en_preparation,
        etats__date_fin__isnull=True
    ).distinct().count()

    ma_performance = EtatCommande.objects.filter(
        enum_etat=etat_preparee,
        date_debut__date=today,
        operateur=operateur_profile
    ).distinct().count()

    commandes_preparees_today = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        date_fin__date=today
    ).count()



    # Commandes préparées cette semaine (toutes)

    commandes_preparees_week = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        date_fin__date__gte=start_of_week,
        date_fin__date__lte=today
    ).count()



    # Commandes actuellement en préparation (toutes)

    commandes_en_preparation = Commande.objects.filter(
        etats__enum_etat=etat_en_preparation,
        etats__date_fin__isnull=True
    ).count()

    # Performance de l'opérateur aujourd'hui (commandes préparées par lui)
    ma_performance_today = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        date_fin__date=today,
        operateur=operateur_profile
    ).count()


    # Valeur totale (DH) des commandes préparées aujourd'hui
    commandes_ids_today = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        date_fin__date=today
    ).values_list('commande_id', flat=True)

    valeur_preparees_today = Commande.objects.filter(id__in=commandes_ids_today).aggregate(total=Sum('total_cmd'))['total'] or 0



    # Articles populaires (semaine en cours)
    commandes_ids_week = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        date_fin__date__gte=start_of_week,
        date_fin__date__lte=today
    ).values_list('commande_id', flat=True)

    articles_populaires = Panier.objects.filter(
        commande_id__in=commandes_ids_week
    ).values('article__nom', 'article__reference').annotate(
        total_quantite=Sum('quantite'),
        total_commandes=Count('commande', distinct=True)
    ).order_by('-total_quantite')[:5]



    # Activité récente (5 dernières préparations de l'opérateur)
    activite_recente = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        operateur=operateur_profile,
        date_fin__isnull=False
    ).select_related('commande', 'commande__client').order_by('-date_fin')[:5]



    # STATISTIQUES GLOBALES ADDITIONNELLES

    # Commandes confirmées aujourd'hui

    commandes_confirmees_today = Commande.objects.filter(
        etats__enum_etat__libelle='Confirmée',
        etats__date_debut__date=today
    ).distinct().count()

    

    # Total global des commandes actives
    total_commandes_actives = commandes_en_cours + commandes_livrees_partiellement + commandes_retournees
    

    # Performance globale (toutes les préparations terminées aujourd'hui)

    performance_globale_today = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        date_fin__date=today
    ).count()

    

    # Valeur totale traitée aujourd'hui

    valeur_totale_today = Commande.objects.filter(
        etats__enum_etat=etat_en_preparation,
        etats__date_fin__date=today
    ).aggregate(total=Sum('total_cmd'))['total'] or 0

    

    # Taux de réussite (commandes terminées vs commandes reçues)
    taux_reussite = 0
    if commandes_confirmees_today > 0:
        taux_reussite = round((performance_globale_today / commandes_confirmees_today) * 100, 1)

    # Alertes de supervision
    alertes = []
    if commandes_retournees > 0:
        alertes.append({
            'type': 'warning',
            'icon': 'fas fa-exclamation-triangle',
            'message': f'{commandes_retournees} commande(s) retournée(s) nécessite(nt) une attention',
            'url': 'Superpreparation:commandes_retournees'
        })
    if commandes_livrees_partiellement > 0:
        alertes.append({
            'type': 'info',
            'icon': 'fas fa-info-circle',
            'message': f'{commandes_livrees_partiellement} commande(s) livrée(s) partiellement',
            'url': 'Superpreparation:commandes_livrees_partiellement'
        })

        alertes.append({
            'type': 'urgent',
            'icon': 'fas fa-fire',

        })
    

    # Préparer les statistiques de supervision

    stats = {
        # Statistiques principales de supervision
        
        'commandes_en_cours': commandes_en_cours,
        'commandes_livrees_partiellement': commandes_livrees_partiellement,
        'commandes_retournees': commandes_retournees,
        'commandes_confirmees_today': commandes_confirmees_today,
        'total_commandes_actives': total_commandes_actives,
        # Performance et KPIs
        'performance_globale_today': performance_globale_today,
        'ma_performance_today': ma_performance_today,
        'taux_reussite': taux_reussite,
        'valeur_totale_today': valeur_totale_today,
        'valeur_preparees_today': valeur_preparees_today,

        # Données historiques
        'commandes_preparees_today': commandes_preparees_today,
        'commandes_preparees_week': commandes_preparees_week,
        'commandes_en_preparation': commandes_en_preparation,

        # Données détaillées
        'articles_populaires': articles_populaires,
        'activite_recente': activite_recente,
        'alertes': alertes

    }

    context = {
        'page_title': 'Centre de Supervision',
        'page_subtitle': f'Tableau de bord global - {total_commandes_actives} commandes actives',
        'profile': operateur_profile,
        'stats': stats,
        'alertes': alertes,
        'is_supervision': True
    }
    return render(request, 'composant_generale/Superpreparation/home.html', context)

def _get_commandes_en_preparation():
    """Récupère toutes les commandes en préparation sans doublons"""
    from django.db.models import Q, Max
    from django.db.models import Prefetch
    
    # États de préparation valides
    etats_preparation = ['En préparation', 'Collectée', 'Emballée', 'Préparée']
    
    # Préparer un Prefetch optimisé pour les états
    etats_qs = (
        EtatCommande.objects.only('id', 'commande_id', 'date_debut', 'date_fin', 'enum_etat_id', 'operateur_id')
            .select_related('enum_etat', 'operateur')
            .order_by('date_debut')
    )

    # Récupérer les commandes en une seule requête avec annotations pour trier côté DB
    # current_etat_date = date_debut max des états actifs de préparation
    qs = (
        Commande.objects.filter(
            Q(etats__enum_etat__libelle__in=etats_preparation),
            etats__date_fin__isnull=True
        )
        .select_related('client', 'ville', 'ville__region')
        .prefetch_related(
            Prefetch('etats', queryset=etats_qs),
            'paniers__article'
        )
        .annotate(
            current_etat_date=Max(
                'etats__date_debut',
                filter=Q(etats__enum_etat__libelle__in=etats_preparation, etats__date_fin__isnull=True)
            )
        )
        .order_by('-current_etat_date', '-id')
        .distinct()
    )

    # Limiter les colonnes ramenées par Commande (évite de gros transferts)
    qs = qs.only('id', 'id_yz', 'num_cmd', 'total_cmd', 'client', 'ville')
    return qs


def _get_etat_actuel_preparation(commande):
    """Trouve l'état actuel de préparation d'une commande"""
    etats_preparation = ['À imprimer', 'En préparation', 'Collectée', 'Emballée', 'Préparée']
    
    for etat in commande.etats.all().order_by('date_debut'):
        if etat.enum_etat.libelle in etats_preparation and not etat.date_fin:
            return etat
    return None


def _get_etat_precedent(commande, etat_actuel):
    """Trouve l'état précédent d'une commande"""
    etats_preparation = ['À imprimer', 'En préparation', 'Collectée', 'Emballée', 'Préparée']
    
    for etat in reversed(commande.etats.all().order_by('date_debut')):
        if (etat.date_fin and 
            etat.date_fin < etat_actuel.date_debut and 
            etat.enum_etat.libelle not in etats_preparation):
            return etat
    return None


def _enrichir_commande(commande):
    """Enrichit une commande avec ses états et informations supplémentaires"""
    etat_actuel = _get_etat_actuel_preparation(commande)
    if not etat_actuel:
        return commande
    
    # Ajouter l'état actuel
    commande.etat_actuel_preparation = etat_actuel
    
    # Ajouter l'état précédent
    commande.etat_precedent = _get_etat_precedent(commande, etat_actuel)
    
    # Ajouter l'état de confirmation
    etat_conf = commande.etats.filter(
        enum_etat__libelle='Confirmée',
        operateur__type_operateur='CONFIRMATION'
    ).order_by('-date_debut').first()
    
    if not etat_conf:
        etat_conf = commande.etats.filter(enum_etat__libelle='Confirmée').order_by('date_debut').first()
    
    commande.etat_confirmation = etat_conf
    
    # Générer le code-barres
    if commande.id_yz:
        code128 = barcode.get_barcode_class('code128')
        barcode_instance = code128(str(commande.id_yz), writer=ImageWriter())
        buffer = BytesIO()
        barcode_instance.write(buffer, options={'write_text': False, 'module_height': 10.0})
        commande.barcode_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    else:
        commande.barcode_base64 = None
    
    return commande


def _filtrer_commandes_par_type(commandes, filter_type):
    """Filtre les commandes selon le type de filtre"""
    from commande.models import Operation
    
    if filter_type == 'retournees':
        return redirect('Superpreparation:commandes_retournees')
    
    if filter_type == 'all':
        # Pour "Toutes les commandes", filtrer les commandes avec des états problématiques
        commandes_filtrees = []
        for commande in commandes:
            etat_actuel = _get_etat_actuel_preparation(commande)
            if not etat_actuel:
                continue
                
            # Vérifier s'il y a des états ultérieurs problématiques
            if etat_actuel.enum_etat.libelle != 'Préparée':
                etats_problematiques = ['Livrée', 'Livrée Partiellement', 'En cours de livraison']
                for etat in commande.etats.all():
                    if (etat.date_debut > etat_actuel.date_debut and 
                        etat.enum_etat.libelle in etats_problematiques):
                        break
                else:
                    commandes_filtrees.append(commande)
            else:
                commandes_filtrees.append(commande)
        return commandes_filtrees
    
    elif filter_type == 'livrees_partiellement':
        # Commandes avec historique de livraison partielle
        commandes_filtrees = []
        for commande in commandes:
            etat_actuel = _get_etat_actuel_preparation(commande)
            if not etat_actuel:
                continue
                
            # Vérifier l'historique de livraison partielle
            for etat in commande.etats.all().order_by('date_debut'):
                if (etat.enum_etat.libelle == 'Livrée Partiellement' and 
                    etat.date_fin and 
                    etat.date_fin < etat_actuel.date_debut):
                    commandes_filtrees.append(commande)
                    break
        return commandes_filtrees
    
    elif filter_type == 'renvoyees_logistique':
        # Commandes renvoyées par la logistique
        commandes_filtrees = []
        for commande in commandes:
            # Vérifier les opérations de renvoi
            operation_renvoi = Operation.objects.filter(
                commande=commande,
                type_operation='RENVOI_PREPARATION'
            ).first()
            
            if operation_renvoi:
                commandes_filtrees.append(commande)
                continue
            
            # Vérifier les commandes de renvoi
            if commande.num_cmd and commande.num_cmd.startswith('RENVOI-'):
                num_cmd_original = commande.num_cmd.replace('RENVOI-', '')
                commande_originale = Commande.objects.filter(
                    num_cmd=num_cmd_original,
                    etats__enum_etat__libelle='Livrée Partiellement'
                ).first()
                if commande_originale:
                    commandes_filtrees.append(commande)
                    continue
            
            # Vérifier l'historique des états
            etat_actuel = _get_etat_actuel_preparation(commande)
            if etat_actuel:
                for etat in reversed(commande.etats.all().order_by('date_debut')):
                    if (etat.date_fin and 
                        etat.date_fin < etat_actuel.date_debut and 
                        etat.enum_etat.libelle == 'En cours de livraison'):
                        commandes_filtrees.append(commande)
                        break
        return commandes_filtrees
    
    return list(commandes)


def _calculer_statistiques(operateur_profile):
    """Calcule les statistiques des commandes"""
    from commande.models import Operation
    
    stats = {
        'renvoyees_logistique': 0,
        'livrees_partiellement': 0,
        'affectees_moi': 0,
    }
    
    # Commandes en préparation
    commandes_preparation = _get_commandes_en_preparation()
    
    for commande in commandes_preparation:
        # Vérifier les opérations de renvoi
        if Operation.objects.filter(
            commande=commande,
            type_operation='RENVOI_PREPARATION'
        ).exists():
            stats['renvoyees_logistique'] += 1
            continue
        
        # Vérifier les commandes de renvoi
        if (commande.num_cmd and 
            commande.num_cmd.startswith('RENVOI-') and
            Commande.objects.filter(
                num_cmd=commande.num_cmd.replace('RENVOI-', ''),
                etats__enum_etat__libelle='Livrée Partiellement'
            ).exists()):
            stats['renvoyees_logistique'] += 1
            continue
        
        # Vérifier l'historique des états
        etat_actuel = _get_etat_actuel_preparation(commande)
        if etat_actuel:
            for etat in reversed(commande.etats.all().order_by('date_debut')):
                if (etat.date_fin and 
                    etat.date_fin < etat_actuel.date_debut):
                    if etat.enum_etat.libelle == 'En cours de livraison':
                        stats['renvoyees_logistique'] += 1
                    elif etat.enum_etat.libelle == 'Livrée Partiellement':
                        stats['livrees_partiellement'] += 1
                    break
    
    # Commandes affectées à l'opérateur
    try:
        superviseur_nom = operateur_profile.nom_complet
        if superviseur_nom:
            stats['affectees_moi'] = Commande.objects.filter(
                etats__enum_etat__libelle='En préparation',
                etats__commentaire__icontains=f"Affectée à la préparation par {superviseur_nom}"
            ).distinct().count()
    except Exception:
        pass
    
    return stats


@superviseur_preparation_required
def liste_prepa(request):
    """Liste des commandes à préparer pour les opérateurs de préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé. Réservé à l'équipe préparation.")
            return redirect('Superpreparation:home')
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('login')
    
    # Récupérer les paramètres
    filter_type = request.GET.get('filter', 'all')
    search_query = request.GET.get('search', '')
    items_per_page = int(request.GET.get('items_per_page', 10))
    etat_param = request.GET.get('etat', '').strip()
    
    # Récupérer et filtrer les commandes
    commandes = _get_commandes_en_preparation()
    commandes_filtrees = _filtrer_commandes_par_type(commandes, filter_type)
    
    # Appliquer la recherche
    if search_query:
        from django.db.models import Q
        commandes_filtrees = [cmd for cmd in commandes_filtrees if 
            search_query.lower() in str(cmd.id_yz).lower() or
            search_query.lower() in (cmd.num_cmd or '').lower() or
            search_query.lower() in cmd.client.nom.lower() or
            search_query.lower() in cmd.client.prenom.lower() or
            search_query.lower() in (cmd.client.numero_tel or '').lower()
        ]
    
    # Filtre serveur par État (actif)
    if etat_param:
        etat_lower = etat_param.lower()
        filtered_by_state = []
        for cmd in commandes_filtrees:
            etat_actif = cmd.etats.filter(date_fin__isnull=True).select_related('enum_etat').first()
            if etat_actif and etat_lower in etat_actif.enum_etat.libelle.lower():
                filtered_by_state.append(cmd)
        commandes_filtrees = filtered_by_state
    
    # Trier par date d'affectation sur la liste brute (avant enrichissement)
    commandes_filtrees.sort(
        key=lambda x: x.etats.filter(date_fin__isnull=True).first().date_debut 
        if x.etats.filter(date_fin__isnull=True).first() 
        else timezone.now(), 
        reverse=True
    )
    
    # Calculer les statistiques
    stats_par_type = _calculer_statistiques(operateur_profile)
    
    # Statistiques générales
    total_affectees = len(commandes_filtrees)
    valeur_totale = sum(getattr(cmd, 'total_cmd', 0) or 0 for cmd in commandes_filtrees)
    date_limite_urgence = timezone.now() - timedelta(days=1)
    commandes_urgentes = sum(1 for cmd in commandes_filtrees if 
        cmd.etats.filter(date_debut__lt=date_limite_urgence).exists()
    )
    
    # Pagination côté serveur
    from django.core.paginator import Paginator
    try:
        items_per_page = int(items_per_page)
        if items_per_page <= 0:
            items_per_page = 10
    except (ValueError, TypeError):
        items_per_page = 10
    paginator = Paginator(commandes_filtrees, items_per_page)
    page_number = request.GET.get('page', 1)
    commandes_page = paginator.get_page(page_number)
    
    # Enrichir uniquement la page courante (optimisation mémoire/CPU)
    try:
        current_list = list(commandes_page.object_list)
        enriched_current = [_enrichir_commande(cmd) for cmd in current_list]
        commandes_page.object_list = enriched_current
    except Exception:
        pass
    
    # Contexte
    context = {
        'page_title': 'Suivi Général - Préparation',
        'page_subtitle': f'Vue d\'ensemble des commandes en file de préparation ({total_affectees})',
        'commandes_affectees': commandes_page,
        'search_query': search_query,
        'filter_type': filter_type,
        'items_per_page': items_per_page,
        'etat': etat_param,
        'page_obj': commandes_page,
        'stats': {
            'total_affectees': total_affectees,
            'valeur_totale': valeur_totale,
            'commandes_urgentes': commandes_urgentes,
        },
        'stats_par_type': stats_par_type,
        'operateur_profile': operateur_profile,
        'api_produits_url_base': reverse('Superpreparation:api_commande_produits', args=[99999999]),
    }
    
    return render(request, 'Superpreparation/liste_prepa.html', context)

@superviseur_preparation_required
def commandes_preparees(request):
    """Liste des commandes préparées (état final) pour les superviseurs"""
    try:
        operateur_profile = request.user.profil_operateur
        if not (operateur_profile.is_preparation or operateur_profile.is_superviseur_preparation):
            messages.error(request, "Accès non autorisé. Réservé à l'équipe préparation.")
            return redirect('Superpreparation:home')
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('login')
    commandes_preparees = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée',
        etats__date_fin__isnull=True  # État actif
    ).select_related('client', 'ville', 'ville__region').prefetch_related('paniers__article', 'etats').distinct()
    search_query = request.GET.get('search', '')
    if search_query:
        commandes_preparees = commandes_preparees.filter(
            Q(id_yz__icontains=search_query) |
            Q(num_cmd__icontains=search_query) |
            Q(client__nom__icontains=search_query) |
            Q(client__prenom__icontains=search_query) |
            Q(client__numero_tel__icontains=search_query)
        ).distinct()

    # Statistiques
    total_preparees = commandes_preparees.count()
    valeur_totale = commandes_preparees.aggregate(total=Sum('total_cmd'))['total'] or 0
    
    # Commandes préparées aujourd'hui
    today = timezone.now().date()
    preparees_today = commandes_preparees.filter(
        etats__enum_etat__libelle='Préparée',
        etats__date_debut__date=today
    ).count()

    # Pagination
    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
        if items_per_page <= 0:
            items_per_page = 10
    except (ValueError, TypeError):
        items_per_page = 10

    paginator = Paginator(commandes_preparees, items_per_page)
    page_number = request.GET.get('page', 1)
    commandes_page = paginator.get_page(page_number)

    context = {
        'page_title': 'Commandes Préparées',
        'page_subtitle': f'Commandes finalisées et prêtes pour livraison ({total_preparees})',
        'commandes_preparees': commandes_page,
        'search_query': search_query,
        'items_per_page': items_per_page,
        'stats': {
            'total_preparees': total_preparees,
            'valeur_totale': valeur_totale,
            'preparees_today': preparees_today,
        },
        'operateur_profile': operateur_profile,
    }
    return render(request, 'Superpreparation/commandes_preparees.html', context)

@superviseur_preparation_required
def commandes_en_preparation(request):
    """Page de suivi (lecture seule) des commandes en préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        if not (operateur_profile.is_preparation or operateur_profile.is_superviseur_preparation):
            messages.error(request, "Accès non autorisé. Réservé à l'équipe préparation.")
            return redirect('Superpreparation:home')
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('Superpreparation:home')
    commandes_en_preparation = Commande.objects.filter(
        etats__enum_etat__libelle='En préparation',
        etats__date_fin__isnull=True  # État actif (en cours)
    ).select_related('client', 'ville', 'ville__region').prefetch_related('paniers__article', 'etats__operateur').distinct()
    search_query = request.GET.get('search', '')
    if search_query:
        commandes_en_preparation = commandes_en_preparation.filter(
            Q(id_yz__icontains=search_query) |
            Q(num_cmd__icontains=search_query) |
            Q(client__nom__icontains=search_query) |
            Q(client__prenom__icontains=search_query) |
            Q(client__numero_tel__icontains=search_query)
        ).distinct()
    stats = {
        'total_commandes': commandes_en_preparation.count(),
        'commandes_urgentes': commandes_en_preparation.filter(
            etats__date_debut__lt=timezone.now() - timedelta(days=1)
        ).count(),
        'valeur_totale': commandes_en_preparation.aggregate(total=Sum('total_cmd'))['total'] or 0
    }
    context = {
        'page_title': 'Suivi - Commandes en Préparation',
        'page_subtitle': f'Suivi en temps réel de {commandes_en_preparation.count()} commande(s) en cours de préparation',
        'profile': operateur_profile,
        'commandes': commandes_en_preparation,
        'search_query': search_query,
        'stats': stats,
        'active_tab': 'en_preparation',
        'is_readonly': True,
        'is_tracking_page': True
    }
    return render(request, 'Superpreparation/commandes_en_preparation.html', context)

@superviseur_preparation_required
def commandes_emballees(request):
    """Page de suivi des commandes emballées qui attendent la finalisation par le superviseur"""
    try:
        operateur_profile = request.user.profil_operateur
        
        # Autoriser superviseur ou équipe préparation
        if not (operateur_profile.is_preparation or operateur_profile.is_superviseur_preparation):
            messages.error(request, "Accès non autorisé. Réservé à l'équipe préparation.")
            return redirect('Superpreparation:home')
            
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('Superpreparation:home')

    # Récupérer TOUTES les commandes emballées qui attendent la finalisation
    commandes_emballees = Commande.objects.filter(
        etats__enum_etat__libelle='Emballée',
        etats__date_fin__isnull=True  # État actif (en cours)
    ).select_related('client', 'ville', 'ville__region').prefetch_related('paniers__article', 'etats__operateur').distinct()

    # Recherche
    search_query = request.GET.get('search', '')
    if search_query:
        commandes_emballees = commandes_emballees.filter(
            Q(id_yz__icontains=search_query) |
            Q(num_cmd__icontains=search_query) |
            Q(client__nom__icontains=search_query) |
            Q(client__prenom__icontains=search_query) |
            Q(client__numero_tel__icontains=search_query)
        ).distinct()

    # Statistiques
    stats = {
        'total_commandes': commandes_emballees.count(),
        'commandes_urgentes': commandes_emballees.filter(
            etats__date_debut__lt=timezone.now() - timedelta(hours=2)  # Urgent si emballée depuis plus de 2h
        ).count(),
        'valeur_totale': commandes_emballees.aggregate(total=Sum('total_cmd'))['total'] or 0
    }

    context = {
        'page_title': 'Suivi - Commandes Emballées',
        'page_subtitle': f'Suivi en temps réel de {commandes_emballees.count()} commande(s) emballées en attente de finalisation',
        'profile': operateur_profile,
        'commandes': commandes_emballees,
        'search_query': search_query,
        'stats': stats,
        'active_tab': 'emballees',

        'is_readonly': False,  # Les superviseurs peuvent finaliser

        'is_tracking_page': True

    }

    return render(request, 'Superpreparation/commandes_emballees.html', context)


@superviseur_preparation_required
def traiter_commande_retournee_api(request, commande_id):
    """API pour traiter une commande retournée et gérer la réincrémentation du stock"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            return JsonResponse({'success': False, 'message': "Accès réservé à l'équipe préparation"})
        commande = get_object_or_404(Commande, id=commande_id)
        
        # Vérifier que la commande est bien en état "Retournée"
        etat_retournee = commande.etats.filter(
            enum_etat__libelle='Retournée',
            date_fin__isnull=True
        ).first()
        
        if not etat_retournee:
            return JsonResponse({'success': False, 'message': 'Commande non retournée'})
        
        # Récupérer les données de la requête
        data = json.loads(request.body)
        type_traitement = data.get('type_traitement')
        etat_stock = data.get('etat_stock')
        commentaire = data.get('commentaire')
        if not all([type_traitement, etat_stock, commentaire]):
            return JsonResponse({'success': False, 'message': 'Données manquantes'})
        
        # Traitement selon le type
        with transaction.atomic(): 
            if type_traitement == 'repreparer':
                # Traiter tous les paniers de la commande
                for panier in commande.paniers.all():
                    quantite = panier.quantite
                    
                    if panier.variante:
                        # Cas avec variante : traiter la variante spécifique
                        variante = panier.variante
                        
                        # Vérifier que la variante est encore active et existe
                        if not variante.actif:
                            return JsonResponse({
                                'success': False, 
                                'message': f'Variante {variante} désactivée. Impossible de traiter le stock.'
                            })
                        
                        # Créer un mouvement de stock pour tracer (dans tous les cas)
                        from article.models import MouvementStock
                        
                        if etat_stock == 'bon':
                            # Réincrémenter le stock de la variante
                            variante.qte_disponible += quantite
                            variante.save()
                            
                            MouvementStock.objects.create(
                                article=panier.article,
                                variante=variante,
                                quantite=quantite,
                                type_mouvement='entree',
                                commentaire=f"Réincrémentation - Commande retournée {commande.id_yz} - {variante.couleur.nom if variante.couleur else 'N/A'} - {variante.pointure.pointure if variante.pointure else 'N/A'} - Produits en bon état - {commentaire}",
                                operateur=operateur_profile,
                                qte_apres_mouvement=variante.qte_disponible
                            )
                        else:
                            # Produits défectueux : juste tracer sans réincrémenter
                            MouvementStock.objects.create(
                                article=panier.article,
                                variante=variante,
                                quantite=0,  # Quantité 0 car pas de réincrémentation
                                type_mouvement='perte',
                                commentaire=f"Produits défectueux - Commande retournée {commande.id_yz} - {variante.couleur.nom if variante.couleur else 'N/A'} - {variante.pointure.pointure if variante.pointure else 'N/A'} - Produits défectueux (non réintégrés) - {commentaire}",
                                operateur=operateur_profile,
                                qte_apres_mouvement=variante.qte_disponible
                            )
                    else:
                        # Cas sans variante : gérer les articles simples (si ce cas existe dans votre logique métier)
                        # Note: Dans ce système, les articles devraient normalement avoir des variantes
                        return JsonResponse({
                            'success': False, 
                            'message': f'Article {panier.article.nom} sans variante définie. Impossible de gérer le stock.'
                        })
                
                # GESTION DES ÉTATS DE COMMANDE
                # Terminer l'état "Retournée" avec la date de fin (la commande reste à l'état "Retournée")
                etat_retournee.terminer_etat(operateur_profile)
                
                # Créer une opération pour tracer l'action
               
                
                # Compter les variantes traitées pour le message de confirmation
                variantes_traitees = commande.paniers.filter(variante__isnull=False).count()
                
                if etat_stock == 'bon':
                    message = f"Stock réincrémenté avec succès pour {variantes_traitees} variante(s). Commande reste en état 'Retournée' avec date de confirmation."
                else:
                    message = f"Traitement effectué pour {variantes_traitees} variante(s) - Stock NON réincrémenté (produits défectueux). Commande reste en état 'Retournée' avec date de confirmation."
                
                return JsonResponse({
                    'success': True,
                    'message': message,
                    'commande_id': commande.id,
                    'variantes_traitees': variantes_traitees,
                    'etat_actuel': 'Retournée',
                    'date_confirmation': timezone.now().strftime('%d/%m/%Y %H:%M')
                })

            return JsonResponse({
                'success': True,
                'message': message,
                'commande_id': commande.id
            })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})
    


@superviseur_preparation_required
def profile_view(request):

    try:

        operateur_profile = request.user.profil_operateur

    except Operateur.DoesNotExist:

        messages.error(request, "Votre profil opérateur n'existe pas.")

        return redirect('login') # Ou une page d'erreur appropriée



    context = {

        'page_title': 'Mon Profil',

        'page_subtitle': 'Gérez les informations de votre profil',

        'profile': operateur_profile,

    }

    return render(request, 'Superpreparation/profile.html', context)

@superviseur_preparation_required
def modifier_profile_view(request):

    try:

        operateur_profile = request.user.profil_operateur

    except Operateur.DoesNotExist:

        messages.error(request, "Votre profil opérateur n'existe pas.")

        return redirect('login')



    if request.method == 'POST':

        # Récupérer les données du formulaire

        nom = request.POST.get('nom')

        prenom = request.POST.get('prenom')

        mail = request.POST.get('mail')

        telephone = request.POST.get('telephone')

        adresse = request.POST.get('adresse')



        # Validation minimale (vous pouvez ajouter plus de validations ici)

        if not nom or not prenom or not mail:

            messages.error(request, "Nom, prénom et email sont requis.")

            context = {

                'page_title': 'Modifier Profil',

                'page_subtitle': 'Mettez à jour vos informations personnelles',

                'profile': operateur_profile,

                'form_data': request.POST, # Pour pré-remplir le formulaire

            }

            return render(request, 'Superpreparation/modifier_profile.html', context)

        

        # Mettre à jour l'utilisateur Django

        request.user.first_name = prenom

        request.user.last_name = nom

        request.user.email = mail

        request.user.save()



        # Mettre à jour le profil de l'opérateur

        operateur_profile.nom = nom

        operateur_profile.prenom = prenom

        operateur_profile.mail = mail

        operateur_profile.telephone = telephone

        operateur_profile.adresse = adresse



        # Gérer l'image si elle est fournie

        if 'photo' in request.FILES:

            operateur_profile.photo = request.FILES['photo']

        

        operateur_profile.save()



        messages.success(request, "Votre profil a été mis à jour avec succès.")

        return redirect('Superpreparation:profile') # Rediriger vers la page de profil après succès

    else:

        context = {

            'page_title': 'Modifier Profil',

            'page_subtitle': 'Mettez à jour vos informations personnelles',

            'profile': operateur_profile,

        }

        return render(request, 'Superpreparation/modifier_profile.html', context)

@superviseur_preparation_required
def changer_mot_de_passe_view(request):

    """Page de changement de mot de passe pour l'opérateur de préparation - Désactivée"""

    return redirect('Superpreparation:profile')

@superviseur_preparation_required
def detail_prepa(request, pk):
    """Vue détaillée pour la préparation d'une commande spécifique"""
    try:
        operateur_profile = request.user.profil_operateur
        # Autoriser superviseur ou équipe préparation
        if not (operateur_profile.is_preparation or operateur_profile.is_superviseur_preparation):
            messages.error(request, "Accès non autorisé. Réservé à l'équipe préparation.")
            return redirect('Superpreparation:home')
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('login')

    # Récupérer la commande spécifique

    try:
        commande = Commande.objects.select_related(
            'client', 'ville', 'ville__region'
        ).prefetch_related(
            'paniers__article', 'paniers__variante', 'etats__enum_etat', 'etats__operateur'
        ).get(id=pk)

    except Commande.DoesNotExist:
        messages.error(request, "La commande demandée n'existe pas.")
        return redirect('Superpreparation:liste_prepa')

    # Récupérer l'état de préparation actuel (peut être affecté à n'importe quel opérateur)

    etat_preparation = commande.etats.filter(
        Q(enum_etat__libelle='À imprimer') | Q(enum_etat__libelle='En préparation'),
        date_fin__isnull=True
    ).first()

    # Récupérer les paniers (articles) de la commande
    paniers = commande.paniers.all().select_related('article', 'variante', 'variante__couleur', 'variante__pointure')
    
    # Initialiser les variables pour les cas de livraison partielle/renvoi
    articles_livres = []
    articles_renvoyes = []
    is_commande_livree_partiellement = False
    commande_renvoi = None # Initialiser à None
    commande_originale = None # Initialiser à None
    etat_articles_renvoyes = {} # Initialiser à un dictionnaire vide

    # Ajouter le prix unitaire, le total de chaque ligne, et l'URL d'image si disponible
    articles_image_urls = {}
    for panier in paniers:
        panier.prix_unitaire = panier.sous_total / panier.quantite if panier.quantite > 0 else 0
        panier.total_ligne = panier.sous_total
        # Préparer affichage variante (couleur/pointure)
        try:
            panier.couleur_display = (
                panier.variante.couleur.nom if getattr(panier, 'variante', None) and getattr(panier, 'variante', 'couleur', None) else getattr(panier.article, 'couleur', '')
            )
        except Exception:
            panier.couleur_display = getattr(panier.article, 'couleur', '')
        try:
            panier.pointure_display = (
                panier.variante.pointure.pointure if getattr(panier, 'variante', None) and getattr(panier.variante, 'pointure', None) else getattr(panier.article, 'pointure', '')
            )
        except Exception:
            panier.pointure_display = getattr(panier.article, 'pointure', '')
        image_url = None
        try:
            # Protéger l'accès à .url si aucun fichier n'est associé
            if getattr(panier.article, 'image', None):
                # Certains backends lèvent une exception si l'image n'existe pas
                if hasattr(panier.article.image, 'url'):
                    image_url = panier.article.image.url
        except Exception:
            image_url = None
        # Rendre accessible directement dans le template via panier.article.image_url
        setattr(panier.article, 'image_url', image_url)
        # Egalement disponible dans le context si nécessaire
        if getattr(panier.article, 'id', None) is not None:
            articles_image_urls[panier.article.id] = image_url
    
    # Récupérer tous les états de la commande pour afficher l'historique

    etats_commande = commande.etats.all().select_related('enum_etat', 'operateur').order_by('date_debut')

    # Déterminer l'état actuel

    etat_actuel = etats_commande.filter(date_fin__isnull=True).first()

    # Récupérer l'état précédent pour comprendre d'où vient la commande
    etat_precedent = None
    if etat_actuel:
        # Trouver l'état précédent (le dernier état terminé avant l'état actuel)
        for etat in reversed(etats_commande):
            if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                if etat.enum_etat.libelle not in ['À imprimer', 'En préparation']:
                    etat_precedent = etat
                    break
    # Analyser les articles pour les commandes livrées partiellement

    articles_livres = []
    articles_renvoyes = []
    is_commande_livree_partiellement = False

    

    # Import pour JSON
    import json
    # Récupérer l'état des articles renvoyés depuis l'opération de livraison partielle (si elle existe)
    etat_articles_renvoyes = {}
    operation_livraison_partielle = None

    # Cas 1: La commande actuelle est la commande originale livrée partiellement

    if etat_actuel and etat_actuel.enum_etat.libelle == 'Livrée Partiellement':
        is_commande_livree_partiellement = True
        # Les articles dans cette commande sont ceux qui ont été livrés partiellement

        for panier in paniers:
            articles_livres.append({
                'article': panier.article,
                'quantite_livree': panier.quantite,
                'prix': panier.article.prix_unitaire,
                'sous_total': panier.sous_total
            })

        # Chercher la commande de renvoi associée
        commande_renvoi = Commande.objects.filter(
            num_cmd__startswith=f"RENVOI-{commande.num_cmd}",
            client=commande.client
        ).first()
        # La commande source pour les articles renvoyés est la commande actuelle

        operation_livraison_partielle = commande.operations.filter(
            type_operation='LIVRAISON_PARTIELLE'
        ).order_by('-date_operation').first()

    # Cas 2: La commande actuelle est une commande de renvoi suite à une livraison partielle
    elif etat_precedent and etat_precedent.enum_etat.libelle == 'Livrée Partiellement':
        is_commande_livree_partiellement = True
        # Chercher la commande originale qui a été livrée partiellement
        commande_originale = Commande.objects.filter(
            num_cmd=commande.num_cmd.replace('RENVOI-', ''),
            client=commande.client
        ).first()
        # La commande source pour les articles renvoyés est la commande originale
        if commande_originale:
            operation_livraison_partielle = commande_originale.operations.filter(
                type_operation='LIVRAISON_PARTIELLE'
            ).order_by('-date_operation').first()
    # Si une opération de livraison partielle est trouvée, extraire les états des articles renvoyés
    if operation_livraison_partielle:
        try:
            details = json.loads(operation_livraison_partielle.conclusion)
            if 'recap_articles_renvoyes' in details:
                for item in details['recap_articles_renvoyes']:
                    etat_articles_renvoyes[item['article_id']] = item['etat']
        except Exception:
            pass

    # Populer articles_renvoyes si c'est une commande de renvoi ou si elle a une commande de renvoi associée

    if is_commande_livree_partiellement:
        # Si la commande actuelle est une commande de renvoi (Cas 2)
        if commande.num_cmd and commande.num_cmd.startswith('RENVOI-'):
            for panier_renvoi in paniers:
                etat = etat_articles_renvoyes.get(panier_renvoi.article.id, 'bon')
                articles_renvoyes.append({
                    'article': panier_renvoi.article,
                    'quantite': panier_renvoi.quantite,
                    'prix': panier_renvoi.article.prix_unitaire,
                    'sous_total': panier_renvoi.sous_total,
                    'etat': etat
                })
        # Si la commande actuelle est la commande originale livrée partiellement et qu'une commande de renvoi existe (Cas 1)
        elif commande_renvoi:
            for panier_renvoi in commande_renvoi.paniers.all():
                etat = etat_articles_renvoyes.get(panier_renvoi.article.id, 'bon')
                articles_renvoyes.append({
                    'article': panier_renvoi.article,
                    'quantite': panier_renvoi.quantite,
                    'prix': panier_renvoi.article.prix_unitaire,
                    'sous_total': panier_renvoi.sous_total,
                    'etat': etat
                })
    # Pour les articles livrés, on lit l'opération de livraison partielle sur la commande originale
    if is_commande_livree_partiellement and commande_originale:
        operation_livraison_partielle_for_livres = commande_originale.operations.filter(
            type_operation='LIVRAISON_PARTIELLE'
        ).order_by('-date_operation').first()
        if operation_livraison_partielle_for_livres:
            try:
                details = json.loads(operation_livraison_partielle_for_livres.conclusion)
                if 'articles_livres' in details:
                    for article_livre in details['articles_livres']:
                        article_id = article_livre.get('article_id')
                        if article_id:
                            article = Article.objects.filter(id=article_id).first()
                        if article:
                            articles_livres.append({
                                'article': article,
                                'quantite_livree': article_livre.get('quantite', 0),
                                'prix': article.prix_unitaire,
                                'sous_total': article.prix_unitaire * article_livre.get('quantite', 0)
                            })
            except Exception:
                pass
    # Calculer le total des articles
    total_articles = sum(panier.total_ligne for panier in paniers)
    # Récupérer les opérations associées à la commande
    operations = commande.operations.select_related('operateur').order_by('-date_operation')
    # Générer le code-barres pour la commande
    code128 = barcode.get_barcode_class('code128')
    barcode_instance = code128(str(commande.id_yz), writer=ImageWriter())
    buffer = BytesIO()
    barcode_instance.write(buffer, options={'write_text': False, 'module_height': 10.0})
    barcode_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    commande_barcode = f"data:image/png;base64,{barcode_base64}"
    # Gestion des actions POST (marquer comme préparée, etc.)
    if request.method == 'POST':
        action = request.POST.get('action')
        # Action 'commencer_preparation' supprimée car les commandes passent maintenant 
        # directement en "En préparation" lors de l'affectation
        if action == 'marquer_preparee':
            with transaction.atomic():
                # Marquer l'état 'En préparation' comme terminé
                etat_en_preparation, created = EnumEtatCmd.objects.get_or_create(libelle='En préparation')
                etat_actuel = EtatCommande.objects.filter(
                    commande=commande,
                    enum_etat=etat_en_preparation,
                    date_fin__isnull=True
                ).first()
                if etat_actuel:
                    etat_actuel.date_fin = timezone.now()
                    etat_actuel.operateur = operateur_profile
                    etat_actuel.save()
                # Créer le nouvel état 'Préparée'
                etat_preparee, created = EnumEtatCmd.objects.get_or_create(libelle='Préparée')
                EtatCommande.objects.create(
                    commande=commande,
                    enum_etat=etat_preparee,
                    operateur=operateur_profile
                )
                # Log de l'opération
               
            messages.success(request, f"La commande {commande.id_yz} a bien été marquée comme préparée.")
            return redirect('Superpreparation:detail_prepa', pk=commande.pk)
        elif action == 'signaler_probleme':
            with transaction.atomic():
                # 1. Terminer l'état "En préparation" actuel
                etat_en_preparation_enum = get_object_or_404(EnumEtatCmd, libelle='En préparation')
                etat_actuel = EtatCommande.objects.filter(
                    commande=commande,
                    enum_etat=etat_en_preparation_enum,
                    date_fin__isnull=True
                ).first()
                if etat_actuel:
                    etat_actuel.date_fin = timezone.now()
                    etat_actuel.commentaire = "Problème signalé par le préparateur."
                    etat_actuel.save()
                # 2. Trouver l'opérateur de confirmation d'origine
                operateur_confirmation_origine = None
                etats_precedents = commande.etats.select_related('operateur').order_by('-date_debut')
                for etat in etats_precedents:
                    if etat.operateur and etat.operateur.is_confirmation:
                        operateur_confirmation_origine = etat.operateur
                        break
                # 3. Créer l'état "Retour Confirmation" et l'affecter
                etat_retour_enum, _ = EnumEtatCmd.objects.get_or_create(
                    libelle='Retour Confirmation',
                    defaults={'ordre': 25, 'couleur': '#D97706'}
                )
                EtatCommande.objects.create(
                    commande=commande,
                    enum_etat=etat_retour_enum,
                    operateur=operateur_confirmation_origine, # Affectation directe
                    date_debut=timezone.now(),
                    commentaire="Retourné par la préparation pour vérification."
                )
                # 4. Log et message de succès
                if operateur_confirmation_origine:
                    log_conclusion = f"Problème signalé par {operateur_profile.nom_complet}. Commande retournée et affectée à l'opérateur {operateur_confirmation_origine.nom_complet}."
                    messages.success(request, f"La commande {commande.id_yz} a été retournée à {operateur_confirmation_origine.nom_complet} pour vérification.")
                else:
                    log_conclusion = f"Problème signalé par {operateur_profile.nom_complet}. Opérateur d'origine non trouvé, commande renvoyée au pool de confirmation."
                    messages.warning(request, f"La commande {commande.id_yz} a été renvoyée au pool de confirmation (opérateur d'origine non trouvé).")
            return redirect('Superpreparation:liste_prepa')
    # Avant le return render dans detail_prepa
    commande_renvoi_id = None
    if commande_renvoi:
        commande_renvoi_id = commande_renvoi.id
    context = {
        'page_title': f'Préparation Commande {commande.id_yz}',
        'page_subtitle': f'Détails de la commande et étapes de préparation',
        'commande': commande,
        'paniers': paniers,
        'etats_commande': etats_commande,
        'etat_actuel': etat_actuel,
        'etat_precedent': etat_precedent,
        'etat_preparation': etat_preparation,
        'total_articles': total_articles,
        'operations': operations,
        'commande_barcode': commande_barcode,
        'is_commande_livree_partiellement': is_commande_livree_partiellement,
        'articles_livres': articles_livres,
        'articles_renvoyes': articles_renvoyes,
        'articles_image_urls': articles_image_urls,
        # Variables de debug/informations supplémentaires
        'commande_originale': commande_originale,
        'commande_renvoi': commande_renvoi,
        'etat_articles_renvoyes': etat_articles_renvoyes,
        'commande_renvoi_id': commande_renvoi_id,
    }
    return render(request, 'Superpreparation/detail_prepa.html', context)




@superviseur_preparation_required
def api_commandes_confirmees(request):
    """API pour récupérer toutes les commandes confirmées"""
    try:
        commandes_confirmees = Commande.objects.filter(
            etats__enum_etat__libelle='Confirmée'
        ).select_related('client').distinct()
        commandes_data = []
        for commande in commandes_confirmees:
            commandes_data.append({
                'id': commande.id_yz,
                'client_nom': f"{commande.client.prenom} {commande.client.nom}",
                'date_creation': commande.date_creation.strftime('%Y-%m-%d %H:%M:%S'),
                'total': float(commande.total_cmd)
            })
        return JsonResponse({
            'success': True,
            'commandes': commandes_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    

@superviseur_preparation_required
def api_articles_commande(request, commande_id):
    """API pour récupérer les articles d'une commande avec leurs codes-barres"""
    try:
        commande = get_object_or_404(
            Commande.objects.select_related('client').prefetch_related(
                'paniers__article', 'paniers__variante'
            ),
            id_yz=commande_id
        )
        articles_data = []
        for panier in commande.paniers.all():
            reference = panier.article.reference
            if panier.variante and panier.variante.reference_variante:
                reference = panier.variante.reference_variante
            try:
                from .templatetags.barcode_filters import barcode_image_url
                barcode_url = barcode_image_url(reference)
            except Exception as e:
                barcode_url = ""
            article_data = {
                'nom': panier.article.nom,
                'reference': panier.article.reference,
                'quantite': panier.quantite,
                'sous_total': float(panier.sous_total),
                'barcode_url': barcode_url,
            }
            if panier.variante:
                variante_info = {
                    'reference_variante': panier.variante.reference_variante,
                }
                if panier.variante.couleur:
                    variante_info['couleur'] = panier.variante.couleur.nom
                if panier.variante.pointure:
                    variante_info['pointure'] = panier.variante.pointure.pointure
                article_data['variante'] = variante_info
            articles_data.append(article_data)
        return JsonResponse({
            'success': True,
            'articles': articles_data,
            'commande_id': commande.id_yz,
            'client': f"{commande.client.prenom} {commande.client.nom}"
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@superviseur_preparation_required
def api_commande_produits(request, commande_id):
    """API pour récupérer les produits d'une commande pour les étiquettes"""
    try:
        # Récupérer la commande. La sécurité est déjà gérée par la page
        # qui appelle cette API, qui ne liste que les commandes autorisées.
        commande = Commande.objects.get(id=commande_id)
        # Récupérer tous les produits de la commande (avec variantes)
        paniers = (
            commande.paniers.all()
            .select_related('article', 'variante', 'variante__couleur', 'variante__pointure')
        )

        # Construire un rendu HTML des produits pour une meilleure lisibilité dans le ticket
        elements_html = []
        total_qte = 0
        for panier in paniers:
            if panier.article:
                nom = panier.article.nom or ''
                reference = panier.article.reference or ''
                couleur = getattr(getattr(panier.variante, 'couleur', None), 'nom', '') if getattr(panier, 'variante', None) else ''
                pointure = getattr(getattr(panier.variante, 'pointure', None), 'pointure', '') if getattr(panier, 'variante', None) else ''
                ref_var = getattr(panier.variante, 'reference_variante', None) if getattr(panier, 'variante', None) else None
                reference_affichee = ref_var or reference

                details = " ".join([v for v in [reference_affichee, couleur, f"P{pointure}" if pointure else ''] if v])
                elements_html.append(f"<li><strong>{nom}</strong> {details} ×{panier.quantite}</li>")
                total_qte += panier.quantite

        

        # Joindre tous les produits en une liste HTML
        produits_text = (
            "<ul class=\"ticket-products\">" + "".join(elements_html) + "</ul>"
            if elements_html else "PRODUITS NON SPÉCIFIÉS"
        )
        
        return JsonResponse({
            'success': True,
            'produits': produits_text,
            'nombre_articles': total_qte,
        })
    except Commande.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Commande non trouvée'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})


# ==================== FONCTIONS HELPERS UPSELL ====================

def determiner_type_prix_gele(article, compteur):
    """
    Détermine le type de prix gelé à enregistrer dans le panier.

    PRIORITÉ 1: Les phases spéciales (promotion, liquidation, test) sont TOUJOURS gelées,
                même pour les articles upsell.

    PRIORITÉ 2: Les articles upsell en phase normale → enregistrer le niveau upsell actuel
                basé sur le compteur au moment de la création du panier.

    PRIORITÉ 3: Les articles normaux en phase normale ont le type 'normal'.
    """
    # PRIORITÉ 1: Phases spéciales et promotions (même pour les articles upsell)
    # Ces types doivent être gelés car ils représentent des prix spéciaux
    if hasattr(article, 'has_promo_active') and article.has_promo_active:
        return 'promotion'
    elif article.phase == 'LIQUIDATION':
        return 'liquidation'
    elif article.phase == 'EN_TEST':
        return 'test'

    # PRIORITÉ 2: Articles upsell en phase normale → enregistrer le niveau selon le compteur
    # Le niveau upsell est gelé au moment de l'ajout du panier
    if article.isUpsell:
        if compteur == 0:
            return 'normal'  # Pas encore de niveau upsell
        elif compteur == 1:
            return 'upsell_niveau_1'
        elif compteur == 2:
            return 'upsell_niveau_2'
        elif compteur == 3:
            return 'upsell_niveau_3'
        elif compteur >= 4:
            return 'upsell_niveau_4'
        else:
            return 'normal'

    # PRIORITÉ 3: Articles normaux en phase normale
    return 'normal'


def mettre_a_jour_types_prix_gele_upsell(commande):
    """
    Met à jour dynamiquement les type_prix_gele de tous les paniers upsell
    en fonction du compteur actuel de la commande.

    Cette fonction doit être appelée après chaque modification du panier qui peut
    impacter le compteur (ajout, suppression, modification de quantité).

    IMPORTANT: Seuls les paniers upsell en phase normale sont mis à jour.
    Les paniers en promotion, liquidation ou test conservent leur type_prix_gele fixe.
    """
    from commande.models import Panier

    # Récupérer tous les paniers upsell de la commande
    paniers_upsell = commande.paniers.filter(article__isUpsell=True)

    for panier in paniers_upsell:
        article = panier.article

        # Recalculer le type_prix_gele basé sur le compteur actuel
        nouveau_type = determiner_type_prix_gele(article, commande.compteur)

        # Mettre à jour uniquement si le type a changé et que ce n'est pas une phase spéciale
        # (les phases spéciales restent figées)
        if nouveau_type != panier.type_prix_gele and nouveau_type not in ['promotion', 'liquidation', 'test']:
            ancien_type = panier.type_prix_gele
            panier.type_prix_gele = nouveau_type
            panier.save(update_fields=['type_prix_gele'])
            print(f"🔄 Panier {panier.id} mis à jour: {ancien_type} → {nouveau_type} (compteur={commande.compteur})")


def _recalculer_compteur_upsell(commande):
    """
    Recalcule le compteur upsell de la commande et met à jour tous les paniers concernés.

    Cette fonction centralise la logique de recalcul du compteur upsell:
    - Compte les articles upsell dans la commande
    - Applique la règle: compteur = max(0, total_upsell - 1) si total >= 2, sinon 0
    - Met à jour les type_prix_gele de tous les paniers upsell
    - Recalcule tous les totaux

    Args:
        commande: L'instance de la commande à recalculer
    """
    from django.db.models import Sum

    # Compter la quantité totale d'articles upsell
    total_quantite_upsell = commande.paniers.filter(
        article__isUpsell=True
    ).aggregate(total=Sum('quantite'))['total'] or 0

    # Règle métier: Le compteur s'incrémente à partir de 2 unités d'articles upsell
    # 0-1 unités → compteur = 0
    # 2+ unités → compteur = total - 1
    if total_quantite_upsell >= 2:
        commande.compteur = total_quantite_upsell - 1
    else:
        commande.compteur = 0

    commande.save()

    # Mettre à jour les type_prix_gele de tous les paniers upsell
    mettre_a_jour_types_prix_gele_upsell(commande)

    # Recalculer tous les totaux
    commande.recalculer_totaux_upsell()

    print(f"🔄 Compteur upsell recalculé: {commande.compteur} (total articles upsell: {total_quantite_upsell})")


@superviseur_preparation_required
def modifier_commande_superviseur(request, commande_id):
    """Page de modification complète d'une commande pour les superviseurs de préparation"""
    import json
    from commande.models import Commande, Operation
    from parametre.models import Ville
    try:
        # Accepter PREPARATION et SUPERVISEUR_PREPARATION
        operateur = Operateur.objects.get(user=request.user, actif=True)
        if operateur.type_operateur not in ['SUPERVISEUR_PREPARATION']:
            messages.error(request, "Accès non autorisé. Réservé à l'équipe préparation.")
            return redirect('Superpreparation:home')
    except Operateur.DoesNotExist:
        # Pas de profil: continuer (le décorateur a déjà validé l'accès via groupes)
        operateur = None
    # Récupérer la commande avec ses relations
    commande = get_object_or_404(Commande.objects.select_related('client', 'ville', 'ville__region'), id=commande_id)
    
    # Pour les superviseurs, on ne vérifie pas l'affectation spécifique
    # Ils peuvent modifier toutes les commandes en préparation
    if request.method == 'POST':
        try:
            # ================ GESTION DES ACTIONS AJAX SPÉCIFIQUES ================
            action = request.POST.get('action')
            
            if action == 'add_article':
                # Ajouter un nouvel article immédiatement
                from article.models import Article, VarianteArticle
                from commande.models import Panier

                article_id = request.POST.get('article_id')
                variante_id = request.POST.get('variante_id')
                quantite = int(request.POST.get('quantite', 1))

                print(f"📥 Ajout article: article_id={article_id}, variante_id={variante_id}, quantite={quantite}")

                try:
                    article = Article.objects.get(id=article_id)
                    variante = None

                    # Récupérer la variante si spécifiée
                    if variante_id:
                        try:
                            variante = VarianteArticle.objects.get(id=variante_id, article=article)
                            print(f"✅ Variante trouvée: ID={variante.id}, couleur={variante.couleur}, pointure={variante.pointure}")
                        except VarianteArticle.DoesNotExist:
                            print(f"⚠️ Variante {variante_id} non trouvée pour l'article {article_id}")

                    # Vérifier si l'article+variante existe déjà dans la commande
                    panier_existant = Panier.objects.filter(
                        commande=commande,
                        article=article,
                        variante=variante
                    ).first()

                    # ========== CRÉATION OU MISE À JOUR DU PANIER ==========
                    if panier_existant:
                        # Article existe déjà → Incrémenter la quantité
                        panier_existant.quantite += quantite
                        panier_existant.sous_total = float(panier_existant.prix_panier * panier_existant.quantite)
                        panier_existant.save()
                        panier = panier_existant
                        print(f"🔄 Article existant mis à jour: ID={article.id}, variante={variante.id if variante else 'None'}, nouvelle quantité={panier.quantite}")
                    else:
                        # Nouvel article → Créer un panier avec le bon prix et type_prix_gele
                        from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
                        prix_panier_initial = get_prix_upsell_avec_compteur(article, commande.compteur)
                        sous_total_initial = float(prix_panier_initial * quantite)
                        type_prix = determiner_type_prix_gele(article, commande.compteur)

                        panier = Panier.objects.create(
                            commande=commande,
                            article=article,
                            quantite=quantite,
                            prix_panier=float(prix_panier_initial),
                            sous_total=sous_total_initial,
                            variante=variante,
                            type_prix_gele=type_prix
                        )
                        print(f"➕ Nouvel article ajouté: ID={article.id}, variante={variante.id if variante else 'None'}, quantité={quantite}, type_prix_gele={type_prix}")

                    # ========== RECALCUL DU COMPTEUR UPSELL ==========
                    if article.isUpsell and hasattr(article, 'prix_upsell_1') and article.prix_upsell_1 is not None:
                        _recalculer_compteur_upsell(commande)

                    # ========== RECALCUL DU TOTAL AVEC FRAIS ==========
                    commande.recalculer_total_avec_frais()

                    # ========== RÉPONSE JSON ==========
                    # Calculer les statistiques pour la réponse
                    total_articles = commande.paniers.aggregate(total=Sum('sous_total'))['total'] or 0
                    frais_livraison = commande.ville.frais_livraison if commande.ville else 0
                    total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
                        total=Sum('quantite')
                    )['total'] or 0

                    message = 'Article ajouté avec succès' if not panier_existant else f'Quantité mise à jour ({panier.quantite})'

                    return JsonResponse({
                        'success': True,
                        'message': message,
                        'article_id': article.id,
                        'article_nom': article.nom,
                        'quantite': panier.quantite,
                        'sous_total': float(panier.sous_total),
                        'sous_total_articles': float(total_articles),
                        'compteur_upsell': commande.compteur,
                        'total_quantite_upsell': total_quantite_upsell,
                        'total_commande': float(commande.total_cmd),
                        'frais_livraison': float(frais_livraison),
                        'frais_livraison_actif': commande.frais_livraison,
                        'total_final': float(commande.total_cmd),
                        'articles_count': commande.paniers.count(),
                        'panier_id': panier.id
                    })
                    
                except Article.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Article non trouvé'})
                except Exception as e:
                    print(f"❌ Erreur lors de l'ajout d'article: {e}")
                    return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})
            
            elif action == 'toggle_frais_livraison':
                # Toggle des frais de livraison
                frais_livraison_actif = request.POST.get('frais_livraison_actif', 'false').lower() == 'true'

                # Mettre à jour le statut des frais de livraison
                commande.frais_livraison = frais_livraison_actif

                # Recalculer le total avec ou sans frais
                total_articles = commande.paniers.aggregate(total=Sum('sous_total'))['total'] or 0
                frais_livraison_montant = commande.ville.frais_livraison if commande.ville else 0

                if frais_livraison_actif:
                    commande.total_cmd = float(total_articles) + float(frais_livraison_montant)
                else:
                    commande.total_cmd = float(total_articles)

                commande.save()

                return JsonResponse({
                    'success': True,
                    'message': f"Frais de livraison {'activés' if frais_livraison_actif else 'désactivés'} avec succès",
                    'nouveau_statut': frais_livraison_actif,
                    'frais_livraison': float(frais_livraison_montant),
                    'total_articles': float(total_articles),
                    'total_commande': float(commande.total_cmd)
                })

            elif action in ['modify_quantity', 'modifier_quantite_directe', 'update_article']:
                # Modifier la quantité d'un article existant
                from commande.models import Panier

                panier_id = request.POST.get('panier_id')
                # Les différentes actions peuvent envoyer des noms différents
                quantite_str = request.POST.get('nouvelle_quantite') or request.POST.get('quantite') or request.POST.get('new_quantity')
                try:
                    nouvelle_quantite = int(quantite_str) if quantite_str is not None else 1
                except ValueError:
                    nouvelle_quantite = 1

                try:
                    panier = Panier.objects.get(id=panier_id, commande=commande)
                    article = panier.article

                    # ========== SAUVEGARDE DES INFORMATIONS ==========
                    ancienne_quantite = panier.quantite
                    etait_upsell = article.isUpsell

                    print(f"🔢 Modification quantité panier {panier_id}: {ancienne_quantite} → {nouvelle_quantite}")

                    # ========== MODIFICATION DE LA QUANTITÉ ==========
                    if nouvelle_quantite <= 0:
                        # Supprimer l'article si quantité <= 0
                        panier.delete()
                        message = f'Article {article.nom} supprimé de la commande'
                    else:
                        # IMPORTANT: Le prix_panier reste INCHANGÉ (prix historique gelé)
                        panier.quantite = nouvelle_quantite
                        panier.sous_total = float(panier.prix_panier * nouvelle_quantite)
                        panier.save()
                        message = f'Quantité de {article.nom} mise à jour: {nouvelle_quantite}'

                    # ========== RECALCUL DU COMPTEUR UPSELL SI NÉCESSAIRE ==========
                    if etait_upsell:
                        _recalculer_compteur_upsell(commande)

                    # ========== RECALCUL DU TOTAL AVEC FRAIS ==========
                    commande.recalculer_total_avec_frais()

                    # ========== RÉPONSE JSON ==========
                    total_articles = commande.paniers.aggregate(total=Sum('sous_total'))['total'] or 0
                    frais_livraison = commande.ville.frais_livraison if commande.ville else 0

                    return JsonResponse({
                        'success': True,
                        'message': message,
                        'sous_total': float(panier.sous_total) if nouvelle_quantite > 0 else 0,
                        'sous_total_articles': float(total_articles),
                        'compteur_upsell': commande.compteur,
                        'total_commande': float(commande.total_cmd),
                        'frais_livraison': float(frais_livraison),
                        'frais_livraison_actif': commande.frais_livraison,
                        'total_final': float(commande.total_cmd),
                        'articles_count': commande.paniers.count()
                    })
                    
                except Panier.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Article non trouvé dans la commande'})
                except Exception as e:
                    print(f"❌ Erreur lors de la modification de quantité: {e}")
                    return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})
            
            
            elif action == 'delete_article':
                # Supprimer un article de la commande
                from commande.models import Panier

                panier_id = request.POST.get('panier_id')

                try:
                    panier = Panier.objects.get(id=panier_id, commande=commande)

                    # ========== SAUVEGARDE DES INFORMATIONS AVANT SUPPRESSION ==========
                    article_nom = panier.article.nom
                    etait_upsell = panier.article.isUpsell

                    print(f"🗑️ Suppression panier {panier_id}: {article_nom} (upsell: {etait_upsell})")

                    # ========== SUPPRESSION DU PANIER ==========
                    panier.delete()

                    # ========== RECALCUL DU COMPTEUR UPSELL SI NÉCESSAIRE ==========
                    if etait_upsell:
                        _recalculer_compteur_upsell(commande)

                    # ========== RECALCUL DU TOTAL AVEC FRAIS ==========
                    commande.recalculer_total_avec_frais()

                    # ========== RÉPONSE JSON ==========
                    total_articles = commande.paniers.aggregate(total=Sum('sous_total'))['total'] or 0
                    frais_livraison = commande.ville.frais_livraison if commande.ville else 0

                    return JsonResponse({
                        'success': True,
                        'message': f'Article "{article_nom}" supprimé avec succès',
                        'compteur_upsell': commande.compteur,
                        'articles_count': commande.paniers.count(),
                        'sous_total_articles': float(total_articles),
                        'total_commande': float(commande.total_cmd),
                        'frais_livraison': float(frais_livraison),
                        'frais_livraison_actif': commande.frais_livraison,
                        'total_final': float(commande.total_cmd)
                    })
                    
                except Panier.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Article non trouvé dans la commande'})
                except Exception as e:
                    print(f"❌ Erreur lors de la suppression d'article: {e}")
                    return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})
            
            else:
                # Soumission du formulaire principal (pas d'action AJAX)
                try:
                    nouvelle_adresse = request.POST.get('adresse', '').strip()
                    nouvelle_ville_id = request.POST.get('ville_id')
                    if nouvelle_adresse:
                        commande.adresse = nouvelle_adresse
                    if nouvelle_ville_id:
                        try:
                            nouvelle_ville = Ville.objects.get(id=nouvelle_ville_id)
                            commande.ville = nouvelle_ville
                        except Ville.DoesNotExist:
                            messages.error(request, "Ville sélectionnée non trouvée.")
                            return redirect('Superpreparation:modifier_commande_superviseur', commande_id=commande.id)
                    commande.save()
                    messages.success(request, f"Les modifications de la commande {commande.id_yz} ont été enregistrées avec succès.")
                    return redirect('Superpreparation:detail_prepa', pk=commande.id)
                except Exception as e:
                    messages.error(request, f"Erreur lors de l'enregistrement: {str(e)}")
                    return redirect('Superpreparation:modifier_commande_superviseur', commande_id=commande.id)
        except Exception as e:
            print(f"❌ Erreur générale dans modifier_commande_superviseur: {e}")
            return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})
    
    # ================ AFFICHAGE DE LA PAGE (GET) ================
    
    # Récupérer les articles de la commande
    paniers = commande.paniers.select_related('article').all()
    # Calculer le total des articles
    total_articles = sum(panier.sous_total for panier in paniers)
    
    # Récupérer les informations de livraison
    frais_livraison = commande.ville.frais_livraison if commande.ville else 0
    
    # Récupérer toutes les villes pour le formulaire
    villes = Ville.objects.all().order_by('nom')
    
    # Vérifier si c'est une commande renvoyée
    is_commande_renvoyee = False
    commande_originale_obj = None
    commande_renvoi_obj = None
    etat_articles_renvoyes = None
    articles_livres = []
    articles_renvoyes = []
    
    # Vérifier si cette commande est un renvoi
    if hasattr(commande, 'commande_renvoi') and commande.commande_renvoi:
        is_commande_renvoyee = True
        commande_originale_obj = commande.commande_renvoi
        commande_renvoi_obj = commande
        
        # Récupérer l'état des articles renvoyés
        try:
            etat_articles_renvoyes = Operation.objects.filter(
                commande=commande_originale_obj,
                enum_etat__libelle='Livrée partiellement'
            ).first()
            
            if etat_articles_renvoyes:
                # Récupérer les articles livrés et renvoyés
                articles_livres = etat_articles_renvoyes.articles_livres.all() if hasattr(etat_articles_renvoyes, 'articles_livres') else []
                articles_renvoyes = etat_articles_renvoyes.articles_renvoyes.all() if hasattr(etat_articles_renvoyes, 'articles_renvoyes') else []
        except Exception as e:
            print(f"⚠️ Erreur lors de la récupération des articles renvoyés: {e}")
    # Préparer le contexte
    context = {
        'commande': commande,
        'paniers': paniers,
        'total_articles': total_articles,
        'operateur': operateur,
        'is_commande_renvoyee': is_commande_renvoyee,
        'articles_livres': articles_livres,
        'articles_renvoyes': articles_renvoyes,
        # Variables de debug/informations supplémentaires
        'commande_originale': commande_originale_obj,
        'commande_renvoi': commande_renvoi_obj,
        'etat_articles_renvoyes': etat_articles_renvoyes,
        # Informations de livraison
        'frais_livraison': float(frais_livraison),
        'villes': villes,
    }
    return render(request, 'Superpreparation/modifier_commande_superviseur.html', context)



def api_articles_disponibles_prepa(request):
    """API pour récupérer les articles disponibles pour les opérateurs de préparation"""
    from article.models import Article
    try:
        # Accepter PREPARATION, SUPERVISEUR_PREPARATION et ADMIN
        operateur = Operateur.objects.get(user=request.user, actif=True)
        if operateur.type_operateur not in ['PREPARATION', 'SUPERVISEUR_PREPARATION', 'ADMIN']:
            return JsonResponse({'success': False, 'message': 'Accès non autorisé'}, status=403)
    except Operateur.DoesNotExist:
        # Pas de profil: continuer (le décorateur a déjà validé l'accès via groupes)
        operateur = None
    try:
        search_query = request.GET.get('search', '')
        filter_type = request.GET.get('filter', 'tous')
        # Récupérer les articles actifs

        articles = Article.objects.filter(actif=True)


        # Appliquer les filtres selon le type

        if filter_type == 'disponible':

            # Filtrer les articles qui ont au moins une variante avec stock > 0

            articles = articles.filter(variantes__qte_disponible__gt=0, variantes__actif=True).distinct()

        elif filter_type == 'upsell':

            articles = articles.filter(isUpsell=True)

        elif filter_type == 'liquidation':

            articles = articles.filter(phase='LIQUIDATION')

        elif filter_type == 'test':

            articles = articles.filter(phase='EN_TEST')

        

        # Recherche textuelle

        if search_query:

            articles = articles.filter(

                Q(nom__icontains=search_query) |

                Q(reference__icontains=search_query) |

                Q(couleur__icontains=search_query) |

                Q(pointure__icontains=search_query) |

                Q(description__icontains=search_query)

            )

        

        # Filtrer les articles en promotion si nécessaire (approche plus sûre)

        if filter_type == 'promo':

            # Filtrer les articles qui ont un prix actuel inférieur au prix unitaire

            articles = articles.filter(

                prix_actuel__isnull=False,

                prix_actuel__lt=F('prix_unitaire')

            )

        

        # Limiter les résultats

        articles = articles[:50]

        

        articles_data = []

        for article in articles:

            # Mettre à jour le prix actuel si nécessaire

            if article.prix_actuel is None:

                article.prix_actuel = article.prix_unitaire

                article.save(update_fields=['prix_actuel'])

            

            # Récupérer toutes les variantes actives (inclet celles en rupture)

            variantes_actives = article.variantes.filter(actif=True)

            

            # Si pas de variantes, créer une entrée avec les propriétés de compatibilité

            if not variantes_actives.exists():

                # Propriétés de compatibilité du modèle Article

                stock = article.qte_disponible

                couleur = article.couleur

                pointure = article.pointure

                

                if stock > 0:

                    articles_data.append({

                        'id': article.id,

                        'nom': article.nom,

                        'reference': article.reference or '',

                        'couleur': couleur or '',

                        'pointure': pointure or '',

                        'description': article.description or '',

                        'prix_unitaire': float(article.prix_unitaire),

                        'prix_actuel': float(article.prix_actuel or article.prix_unitaire),

                        'prix_upsell_1': float(article.prix_upsell_1) if article.prix_upsell_1 else 0.0,

                        'prix_upsell_2': float(article.prix_upsell_2) if article.prix_upsell_2 else 0.0,

                        'prix_upsell_3': float(article.prix_upsell_3) if article.prix_upsell_3 else 0.0,

                        'prix_upsell_4': float(article.prix_upsell_4) if article.prix_upsell_4 else 0.0,

                        'qte_disponible': stock,

                        'isUpsell': bool(article.isUpsell),

                        'phase': article.phase or 'NORMAL',

                        'has_promo_active': article.has_promo_active,

                        'image_url': article.image.url if article.image else article.image_url,

                        'categorie': str(article.categorie) if article.categorie else '',

                        'genre': str(article.genre) if article.genre else '',

                        'modele': article.modele_complet(),

                        'variantes_count': 0,

                        'is_variante': False,

                        'variante_id': None,

                        # Propriétés pour compatibilité avec le template

                        'prix': float(article.prix_actuel or article.prix_unitaire),

                        'prix_original': float(article.prix_unitaire),

                        'has_reduction': article.has_promo_active,

                        'reduction_pourcentage': round(((float(article.prix_unitaire) - float(article.prix_actuel or article.prix_unitaire)) / float(article.prix_unitaire)) * 100, 0) if article.has_promo_active else 0,

                        'article_type': 'normal',

                        'type_icon': 'fas fa-box',

                        'type_color': 'text-gray-600',

                        'display_text': f"{article.nom} - {couleur or ''} - {pointure or ''} ({float(article.prix_actuel or article.prix_unitaire):.2f} DH)"

                    })

            else:

                # Préparer la liste complète des variantes pour affichage (y compris rupture)

                variantes_list = []

                for v in variantes_actives:

                    variantes_list.append({

                        'id': v.id,

                        'couleur': v.couleur.nom if v.couleur else '',

                        'pointure': v.pointure.pointure if v.pointure else '',

                        'qte_disponible': v.qte_disponible,

                        'reference_variante': v.reference_variante,

                    })



                # Créer une entrée pour chaque variante (même celles en rupture)

                for variante in variantes_actives:

                    # Déterminer le type d'article pour l'affichage

                    article_type = 'normal'

                    type_icon = 'fas fa-box'

                    type_color = 'text-gray-600'

                    

                    if article.isUpsell:

                        article_type = 'upsell'

                        type_icon = 'fas fa-arrow-up'

                        type_color = 'text-purple-600'

                    elif article.phase == 'LIQUIDATION':

                        article_type = 'liquidation'

                        type_icon = 'fas fa-money-bill-wave'

                        type_color = 'text-red-600'

                    elif article.phase == 'EN_TEST':

                        article_type = 'test'

                        type_icon = 'fas fa-flask'

                        type_color = 'text-yellow-600'

                    

                    # Vérifier si l'article est en promotion

                    if article.has_promo_active:

                        article_type = 'promo'

                        type_icon = 'fas fa-fire'

                        type_color = 'text-orange-600'

                    

                    articles_data.append({

                        'id': article.id,

                        'nom': article.nom,

                        'reference': article.reference or '',

                        'couleur': variante.couleur.nom if variante.couleur else '',

                        'pointure': variante.pointure.pointure if variante.pointure else '',

                        'description': article.description or '',

                        'prix_unitaire': float(article.prix_unitaire),

                        'prix_actuel': float(article.prix_actuel or article.prix_unitaire),

                        'prix_upsell_1': float(article.prix_upsell_1) if article.prix_upsell_1 else 0.0,

                        'prix_upsell_2': float(article.prix_upsell_2) if article.prix_upsell_2 else 0.0,

                        'prix_upsell_3': float(article.prix_upsell_3) if article.prix_upsell_3 else 0.0,

                        'prix_upsell_4': float(article.prix_upsell_4) if article.prix_upsell_4 else 0.0,

                        'qte_disponible': variante.qte_disponible,

                        'isUpsell': bool(article.isUpsell),

                        'phase': article.phase or 'NORMAL',

                        'has_promo_active': article.has_promo_active,

                        'image_url': article.image.url if article.image else article.image_url,

                        'categorie': str(article.categorie) if article.categorie else '',

                        'genre': str(article.genre) if article.genre else '',

                        'modele': article.modele_complet(),

                        'variantes_count': variantes_actives.count(),

                        'is_variante': True,

                        'variante_id': variante.id,

                        'reference_variante': variante.reference_variante,

                        'variantes_all': variantes_list,

                        # Propriétés pour compatibilité avec le template

                        'prix': float(article.prix_actuel or article.prix_unitaire),

                        'prix_original': float(article.prix_unitaire),

                        'has_reduction': article.has_promo_active,

                        'reduction_pourcentage': round(((float(article.prix_unitaire) - float(article.prix_actuel or article.prix_unitaire)) / float(article.prix_unitaire)) * 100, 0) if article.has_promo_active else 0,

                        'article_type': article_type,

                        'type_icon': type_icon,

                        'type_color': type_color,

                        'display_text': f"{article.nom} - {variante.couleur.nom if variante.couleur else ''} - {variante.pointure.pointure if variante.pointure else ''} ({float(article.prix_actuel or article.prix_unitaire):.2f} DH)"

                    })

        

        # Calculer les statistiques pour la réponse

        stats = {

            "total": len(articles_data),

            "disponible": len([a for a in articles_data if a.get('qte_disponible', 0) > 0]),

            "upsell": len([a for a in articles_data if a.get('isUpsell', False)]),

            "liquidation": len([a for a in articles_data if a.get('phase') == 'LIQUIDATION']),

            "test": len([a for a in articles_data if a.get('phase') == 'EN_TEST']),

            "promo": len([a for a in articles_data if a.get('has_promo_active', False)])

        }

        

        # Retourner le format attendu par le frontend

        return JsonResponse({

            'success': True,

            'articles': articles_data,

            'stats': stats

        })

        

    except Exception as e:

        import traceback

        traceback.print_exc()  # This will print to the server console

        return JsonResponse({

            'success': False,

            'message': 'Erreur interne du serveur lors du chargement des articles.',

            'details': str(e)

        }, status=500)


@superviseur_preparation_required
def api_panier_commande_prepa(request, commande_id):
    """API pour récupérer le panier d'une commande pour les opérateurs de préparation"""
    try:
        # Accepter PREPARATION et SUPERVISEUR_PREPARATION
        operateur = Operateur.objects.get(user=request.user, actif=True)
        if operateur.type_operateur not in ['PREPARATION', 'SUPERVISEUR_PREPARATION']:
            return JsonResponse({'success': False, 'message': 'Accès non autorisé'})
    except Operateur.DoesNotExist:
        # Pas de profil: continuer (le décorateur a déjà validé l'accès via groupes)
        operateur = None
    # Récupérer la commande
    try:
        commande = Commande.objects.get(id=commande_id)
    except Commande.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Commande non trouvée'})
    # Pour les superviseurs, on ne vérifie pas l'affectation spécifique
    # Ils peuvent accéder à toutes les commandes en préparation
    if operateur and operateur.type_operateur == 'SUPERVISEUR_PREPARATION':
        # Vérifier seulement que la commande est en préparation (incluant les états finaux)
        etat_preparation = commande.etats.filter(
            Q(enum_etat__libelle='En préparation') | Q(enum_etat__libelle='Collectée') | Q(enum_etat__libelle='Emballée') | Q(enum_etat__libelle='Préparée'),
            date_fin__isnull=True
        ).first()
    else:
        # Pour les opérateurs normaux, vérifier l'affectation spécifique
        etat_preparation = commande.etats.filter(
            Q(enum_etat__libelle='En préparation') | Q(enum_etat__libelle='Collectée') | Q(enum_etat__libelle='Emballée') | Q(enum_etat__libelle='Préparée'),
            operateur=operateur,
            date_fin__isnull=True
        ).first()
    if not etat_preparation:
        return JsonResponse({'success': False, 'message': 'Commande non affectée'})
    # Récupérer les paniers
    paniers = commande.paniers.all().select_related('article')
    # Construire le format "articles" attendu par le front (compatibilité)
    articles = []
    total_articles_montant = 0.0
    for panier in paniers:
        prix_unitaire = float(panier.article.prix_unitaire) if panier.article and panier.article.prix_unitaire is not None else 0.0
        sous_total = float(panier.sous_total) if panier.sous_total is not None else round(prix_unitaire * (panier.quantite or 0), 2)
        total_articles_montant += sous_total
        articles.append({
            'id': panier.id,
            'article_id': getattr(panier.article, 'id', None),
            'nom': getattr(panier.article, 'nom', ''),
            'reference': getattr(panier.article, 'reference', '') or '',
            'couleur': getattr(panier.article, 'couleur', ''),
            'pointure': getattr(panier.article, 'pointure', ''),
            'quantite': panier.quantite or 0,
            'prix_unitaire': prix_unitaire,
            'sous_total': sous_total,
        })
    # Objet commande attendu par le front
    client_nom = ''
    try:
        if commande.client:
            prenom = getattr(commande.client, 'prenom', '') or ''
            nom = getattr(commande.client, 'nom', '') or ''
            client_nom = (prenom + ' ' + nom).strip()
    except Exception:
        client_nom = ''
    commande_payload = {
        'id': commande.id,
        'id_yz': getattr(commande, 'id_yz', ''),
        'client_nom': client_nom,
        'total_articles': len(articles),
        'total_montant': round(total_articles_montant, 2),
        'total_final': float(commande.total_cmd) if getattr(commande, 'total_cmd', None) is not None else round(total_articles_montant, 2),
    }
    # Réponse incluant les deux formats (nouveau et legacy) pour compatibilité
    return JsonResponse({
        'success': True,
        # Nouveau format utilisé par le front de Suivi Général
        'commande': commande_payload,
        'articles': articles,
        # Legacy (au cas où)
        'paniers': articles,
        'total_commande': commande_payload['total_final'],
        'nb_articles': len(articles),
    })

@superviseur_preparation_required
def api_finaliser_commande(request, commande_id):
    """API pour finaliser une commande depuis la liste (pour les superviseurs)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)
    try:
        operateur_profile = request.user.profil_operateur
        if not (operateur_profile.is_preparation or operateur_profile.is_superviseur_preparation):
            return JsonResponse({'success': False, 'message': 'Accès non autorisé'}, status=403)
    except Operateur.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Profil opérateur non trouvé'}, status=403)
    try:
        commande = Commande.objects.get(id=commande_id)
    except Commande.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Commande non trouvée'}, status=404)
    # Vérifier que la commande est emballée (prête pour finalisation)
    etat_actuel = commande.etat_actuel
    if not etat_actuel or etat_actuel.enum_etat.libelle != 'Emballée':
        return JsonResponse({
            'success': False, 
            'message': f'La commande n\'est pas emballée (état actuel: {etat_actuel.enum_etat.libelle if etat_actuel else "Aucun"})'
        }, status=400)
    try:
        with transaction.atomic():
            # Marquer l'état 'Emballée' comme terminé
            etat_actuel.date_fin = timezone.now()
            etat_actuel.operateur = operateur_profile
            etat_actuel.save()
            # Créer le nouvel état 'Préparée' (final)
            etat_preparee, created = EnumEtatCmd.objects.get_or_create(libelle='Préparée')
            EtatCommande.objects.create(
                commande=commande,
                enum_etat=etat_preparee,
                operateur=operateur_profile
            )
            
        return JsonResponse({
            'success': True,
            'message': f'La commande {commande.id_yz} a été préparée avec succès.'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de la finalisation: {str(e)}'
        }, status=500)


@superviseur_preparation_required
def api_panier_commande(request, commande_id):
    """API pour récupérer le contenu du panier d'une commande (pour le modal)"""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)
    try:
        operateur_profile = request.user.profil_operateur
        if not (operateur_profile.is_preparation or operateur_profile.is_superviseur_preparation):
            return JsonResponse({'success': False, 'message': 'Accès non autorisé'}, status=403)
    except Operateur.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Profil opérateur non trouvé'}, status=403)
    try:
        commande = Commande.objects.get(id=commande_id)
    except Commande.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Commande non trouvée'}, status=404)
    try:
        # Récupérer les articles du panier
        paniers = commande.paniers.all().select_related('article')
        if not paniers.exists():
            html_content = '''
                <div class="text-center py-8">
                    <i class="fas fa-shopping-cart text-4xl text-gray-300 mb-4"></i>
                    <p class="text-gray-500">Aucun article dans le panier</p>
                </div>
            '''
            return JsonResponse({
                'success': True,
                'html': html_content
            })

        # Construire le HTML du panier
        html_content = '''
            <div class="space-y-4">
                <div class="bg-gray-50 p-4 rounded-lg">
                    <h4 class="font-medium text-gray-900 mb-2">Commande #''' + str(commande.id_yz) + '''</h4>
                    <p class="text-sm text-gray-600">Client: ''' + (commande.client.prenom + ' ' + commande.client.nom if commande.client else 'Non défini') + '''</p>
                </div>
                <div class="space-y-3">
        '''
        total_articles = 0
        total_montant = 0
        for panier in paniers:
            article = panier.article
            quantite = panier.quantite or 0
            prix = getattr(article, 'prix', 0) or 0
            sous_total = quantite * prix
            total_articles += quantite
            total_montant += sous_total

            # Informations de l'article
            nom_article = getattr(article, 'nom', 'Article sans nom') or 'Article sans nom'
            reference = getattr(article, 'reference', '') or ''
            couleur = getattr(article, 'couleur', '') or ''
            pointure = getattr(article, 'pointure', '') or ''
            html_content += '''
                <div class="border border-gray-200 rounded-lg p-4">
                    <div class="flex justify-between items-start mb-2">
                        <div class="flex-1">
                            <h5 class="font-medium text-gray-900">''' + nom_article + '''</h5>
                            <p class="text-sm text-gray-600">Réf: ''' + reference + '''</p>
                        </div>
                        <div class="text-right">
                            <span class="text-lg font-bold text-green-600">''' + str(prix) + ''' DH</span>
                        </div>
                    </div>
                    <div class="grid grid-cols-2 gap-4 text-sm">
                        <div>
                            <span class="text-gray-500">Quantité:</span>
                            <span class="font-medium text-gray-900 ml-1">''' + str(quantite) + '''</span>
                        </div>
                        <div>
                            <span class="text-gray-500">Sous-total:</span>
                            <span class="font-medium text-green-600 ml-1">''' + str(sous_total) + ''' DH</span>
                        </div>
            '''
       # Ajouter couleur et pointure si disponibles
            if couleur or pointure:
                html_content += '''
                        <div class="col-span-2 flex space-x-4">
                '''
                if couleur:
                    html_content += '''
                            <div>
                                <span class="text-gray-500">Couleur:</span>
                                <span class="font-medium text-gray-900 ml-1">''' + couleur + '''</span>
                            </div>
                    '''
                if pointure:
                    html_content += '''
                            <div>
                                <span class="text-gray-500">Pointure:</span>
                                <span class="font-medium text-gray-900 ml-1">''' + pointure + '''</span>
                            </div>
                    '''
                html_content += '''
                        </div>
                '''
            html_content += '''
                    </div>
                </div>
            '''
        # Ajouter le total
        html_content += '''
                </div>
                <div class="border-t border-gray-200 pt-4">
                    <div class="flex justify-between items-center">
                        <div class="text-lg font-medium text-gray-900">
                            Total (''' + str(total_articles) + ''' article''' + ('s' if total_articles > 1 else '') + '''):
                        </div>
                        <div class="text-2xl font-bold text-green-600">
                            ''' + str(round(total_montant, 2)) + ''' DH
                        </div>
                    </div>
                </div>
            </div>
        '''
        return JsonResponse({
            'success': True,
            'html': html_content
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors du chargement du panier: {str(e)}'
        }, status=500)


@superviseur_preparation_required
def imprimer_tickets_preparation(request):
    """
    Vue pour imprimer les tickets de préparation SANS changer l'état des commandes.
    Permet d'imprimer ou de réimprimer des tickets pour les commandes en préparation.
    """
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            return HttpResponse("Accès non autorisé.", status=403)
    except Operateur.DoesNotExist:
        return HttpResponse("Profil opérateur non trouvé.", status=403)
    commande_ids_str = request.GET.get('ids')
    if not commande_ids_str:
        return HttpResponse("Aucun ID de commande fourni.", status=400)
    try:
        commande_ids = [int(id) for id in commande_ids_str.split(',') if id.isdigit()]
    except ValueError:
        return HttpResponse("IDs de commande invalides.", status=400)
    commandes = Commande.objects.filter(
        id__in=commande_ids,
        etats__operateur=operateur_profile,
        etats__enum_etat__libelle='En préparation',
        etats__date_fin__isnull=True
    ).distinct()

    if not commandes.exists():
        messages.info(request, "L'impression des tickets est désactivée. Utilisez les outils de gestion.")
        return redirect('Superpreparation:liste_prepa')
    code128 = barcode.get_barcode_class('code128')
    messages.info(request, "L'impression des tickets a été retirée de l'interface superviseur.")
    return redirect('Superpreparation:liste_prepa')


def get_operateur_display_name(operateur):
    """Fonction helper pour obtenir le nom d'affichage d'un opérateur"""
    if not operateur:
        return "Opérateur inconnu"
    
    if hasattr(operateur, 'nom_complet') and operateur.nom_complet:
        return operateur.nom_complet
    elif operateur.nom and operateur.prenom:
        return f"{operateur.prenom} {operateur.nom}"
    elif operateur.nom:
        return operateur.nom
    elif hasattr(operateur, 'user') and operateur.user:
        return operateur.user.username
    else:
        return "Opérateur inconnu"

@superviseur_preparation_required
def export_envois(request):
    """Export des envois journaliers - Service de préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé.")
            return redirect('login')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil opérateur non trouvé.")
        return redirect('login')
    
    from parametre.models import Region, Operateur
    from django.utils import timezone
    import datetime
    
    # Date par défaut : aujourd'hui
    today = timezone.now().date()
    date_envoi = request.GET.get('date_envoi', today)
    region_id = request.GET.get('region')
    livreur_id = request.GET.get('livreur')
    
    # Obtenir tous les livreurs (opérateurs de livraison)
    livreurs = Operateur.objects.filter(is_livraison=True, actif=True)
    regions = Region.objects.all()
    
    # Simuler des envois (à remplacer par votre modèle Envoi)
    envois = []
    
    # Commandes PRÉPARÉES à être envoyées
    commandes_pretes = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée',
        etats__date_fin__isnull=True
    ).select_related('ville__region')
    
    if region_id and region_id.strip():
        commandes_pretes = commandes_pretes.filter(ville__region_id=region_id)
    
    # Statistiques
    stats = {
        'total_envois': len(envois),
        'total_commandes': 0,
        'commandes_pretes': commandes_pretes.count(),
        'livreurs_actifs': livreurs.filter(actif=True).count(),
    }
    
    context = {
        'envois': envois,
        'commandes_pretes': commandes_pretes,
        'livreurs': livreurs,
        'regions': regions,
        'stats': stats,
        'today': today,
    }
    
    return render(request, 'Superpreparation/export_envois.html', context)


def details_region_view(request):
    """Vue détaillée pour afficher les commandes par région - Service de préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé.")
            return redirect('login')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil opérateur non trouvé.")
        return redirect('login')
    
    from parametre.models import Region, Ville
    
    # Récupérer les paramètres de filtrage
    region_name = request.GET.get('region')
    ville_name = request.GET.get('ville')
    
    # Base queryset pour toutes les commandes en traitement
    commandes_reparties = Commande.objects.filter(
        etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée', 'En cours de livraison'],
        etats__date_fin__isnull=True,
        ville__isnull=False,  # Exclure les commandes sans ville
        ville__region__isnull=False  # Exclure les commandes sans région
    ).select_related('client', 'ville', 'ville__region').prefetch_related(
        'etats__operateur', 'etats__enum_etat', 'paniers__article'
    ).distinct()
    
    # Appliquer les filtres
    if region_name:
        commandes_reparties = commandes_reparties.filter(ville__region__nom_region=region_name)
    if ville_name:
        commandes_reparties = commandes_reparties.filter(ville__nom=ville_name)
    
    # Statistiques par ville dans la région/ville filtrée
    stats_par_ville = commandes_reparties.values(
        'ville__id', 'ville__nom', 'ville__region__nom_region'
    ).annotate(
        nb_commandes=Count('id'),
        total_montant=Sum('total_cmd')
    ).order_by('ville__region__nom_region', 'ville__nom')
    
    # Statistiques des commandes PRÉPARÉES par ville dans la région/ville filtrée
    commandes_preparees = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée',
        etats__date_fin__isnull=True,
        ville__isnull=False,
        ville__region__isnull=False
    ).select_related('ville', 'ville__region')
    
    # Appliquer les mêmes filtres que pour les commandes en traitement
    if region_name:
        commandes_preparees = commandes_preparees.filter(ville__region__nom_region=region_name)
    if ville_name:
        commandes_preparees = commandes_preparees.filter(ville__nom=ville_name)
    
    stats_preparees_par_ville = commandes_preparees.values(
        'ville__nom', 'ville__region__nom_region'
    ).annotate(
        nb_commandes_preparees=Count('id')
    ).order_by('ville__region__nom_region', 'ville__nom')
    
    # Créer un dictionnaire pour un accès rapide
    preparees_par_ville = {(stat['ville__nom'], stat['ville__region__nom_region']): stat['nb_commandes_preparees'] for stat in stats_preparees_par_ville}
    
    # Calculer les totaux
    total_commandes = commandes_reparties.count()
    total_montant = commandes_reparties.aggregate(total=Sum('total_cmd'))['total'] or 0
    
    # Définir le titre selon le filtre appliqué
    if region_name:
        page_title = f"Détails - {region_name}"
        page_subtitle = f"Commandes en traitement dans la région {region_name}"
    elif ville_name:
        page_title = f"Détails - {ville_name}"
        page_subtitle = f"Commandes en traitement à {ville_name}"
    else:
        page_title = "Détails par Région"
        page_subtitle = "Répartition détaillée des commandes en traitement"
    
    context = {
        'operateur': operateur_profile,
        'commandes_reparties': commandes_reparties,
        'stats_par_ville': stats_par_ville,
        'preparees_par_ville': preparees_par_ville,
        'total_commandes': total_commandes,
        'total_montant': total_montant,
        'region_name': region_name,
        'ville_name': ville_name,
        'page_title': page_title,
        'page_subtitle': page_subtitle,
    }
    
    return render(request, 'Superpreparation/details_region.html', context)

@superviseur_preparation_required
def exporter_envoi(request, envoi_id, format):
    """
    Exporte un envoi dans un format spécifique (CSV/PDF).
    """
    envoi = get_object_or_404(Envoi, id=envoi_id)
    if format == 'csv':
        # Logique d'export CSV
        return HttpResponse(f"Export CSV de l'envoi {envoi_id}", content_type="text/csv")
    elif format == 'pdf':
        # Logique d'export PDF
        return HttpResponse(f"Export PDF de l'envoi {envoi_id}", content_type="application/pdf")
    return HttpResponse("Format non supporté", status=400)


@superviseur_preparation_required
def envois_view(request):
    """Page de gestion des envois (création/en cours)."""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé.")
            return redirect('login')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil opérateur non trouvé.")
        return redirect('login')

    # Données placeholder: commandes prêtes (même logique que export_envois)
    from parametre.models import Region
    from django.db.models import Count, Q

    regions = Region.objects.all()
    commandes_pretes = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée',
        etats__date_fin__isnull=True
    ).select_related('ville__region')

    # Récupérer les envois actifs (non clôturés)
    envois_actifs = Envoi.objects.exclude(
        status=False
    ).select_related('region').order_by('-date_creation')

    # Mettre à jour le compteur de commandes pour chaque envoi
    for envoi in envois_actifs:
        nb_commandes = Commande.objects.filter(
            ville__region=envoi.region,
            etats__enum_etat__libelle='Préparée',
            etats__date_fin__isnull=True
        ).count()

        # Mettre à jour seulement si le nombre a changé
        if envoi.nb_commandes != nb_commandes:
            envoi.nb_commandes = nb_commandes
            envoi.save(update_fields=['nb_commandes'])

    # Récupérer les IDs des régions qui ont déjà un envoi actif
    regions_avec_envoi_actif = set(envois_actifs.values_list('region_id', flat=True))

    # Annoter chaque région avec le nombre de commandes prêtes
    for region in regions:
        region.nb_commandes_pretes = Commande.objects.filter(
            ville__region=region,
            etats__enum_etat__libelle='Préparée',
            etats__date_fin__isnull=True
        ).count()

    context = {
        'regions': regions,
        'commandes_pretes': commandes_pretes,
        'envois_actifs': envois_actifs,
        'regions_avec_envoi_actif': regions_avec_envoi_actif,
        'page_title': 'Gestion des envois',
        'page_subtitle': 'Créer et suivre les envois en cours',
    }
    return render(request, 'Superpreparation/envois.html', context)

@superviseur_preparation_required
def commandes_envoi(request, envoi_id):
    """Récupérer les commandes d'un envoi spécifique"""
    try:
        envoi = Envoi.objects.get(id=envoi_id)
        
        # Récupérer seulement les commandes préparées de cet envoi spécifique
        commandes = Commande.objects.filter(
            etats__enum_etat__libelle='Préparée',
            etats__date_fin__isnull=True  # État actuel (pas terminé)
        ).select_related('client', 'ville', 'ville__region').prefetch_related(
            'etats', 
            'paniers__article', 
            'paniers__variante__couleur', 
            'paniers__variante__pointure'
        ).distinct()

        # Préparer les données pour le JSON
        commandes_data = []
        for commande in commandes:
            try:
                # Récupérer l'état actuel de manière sécurisée
                etat_actuel = commande.etat_actuel
                libelle_etat = None
                if etat_actuel and etat_actuel.enum_etat:
                    libelle_etat = etat_actuel.enum_etat.libelle
                else:
                    print(f"DEBUG: Aucun état actuel trouvé pour la commande {commande.id}")
                
                # Récupérer les paniers avec leurs variantes
                paniers_data = []
                for panier in commande.paniers.all():
                    panier_info = {
                        'id': panier.id,
                        'article': {
                            'id': panier.article.id,
                            'nom': panier.article.nom,
                            'reference': panier.article.reference,
                        },
                        'quantite': panier.quantite,
                        'sous_total': float(panier.sous_total) if panier.sous_total else 0,
                        'variante': None
                    }
                    
                    # Ajouter les informations de variante si elle existe
                    if panier.variante:
                        panier_info['variante'] = {
                            'id': panier.variante.id,
                            'reference_variante': panier.variante.reference_variante,
                            'couleur': panier.variante.couleur.nom if panier.variante.couleur else None,
                            'pointure': panier.variante.pointure.pointure if panier.variante.pointure else None,
                            'qte_disponible': panier.variante.qte_disponible,
                        }
                    else:
                        print(f"DEBUG: Aucune variante pour panier {panier.id}")
                    
                    paniers_data.append(panier_info)
                
                commandes_data.append({
                    'id': commande.id,
                    'id_yz': commande.id_yz,
                    'num_cmd': commande.num_cmd,
                    'total_cmd': float(commande.total_cmd) if commande.total_cmd else 0,
                    'client': {
                        'nom': commande.client.nom if commande.client else None,
                        'prenom': commande.client.prenom if commande.client else None,
                        'email': commande.client.email if commande.client else None,
                        'numero_tel': commande.client.numero_tel if commande.client else None,
                    },
                    'ville': {
                        'nom_ville': commande.ville.nom if commande.ville else None,
                    },
                    'etat_actuel': {
                        'enum_etat': {
                            'libelle': libelle_etat,
                        }
                    } if etat_actuel else None,
                    'paniers': paniers_data,
                })
            except Exception as e:
                # Si une commande spécifique pose problème, on la saute et on continue
                print(f"Erreur lors du traitement de la commande {commande.id}: {str(e)}")
                continue
        

        return JsonResponse({
            'success': True,
            'commandes': commandes_data,
            'envoi': {
                'id': envoi.id,
                'numero_envoi': envoi.numero_envoi,
                'region': envoi.region.nom_region,
                'nb_commandes': envoi.nb_commandes,
            }
        })
        
    except Envoi.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Envoi non trouvé'
        }, status=404)
    except Exception as e:
        import traceback
        print(f"Erreur dans commandes_envoi: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la récupération des commandes: {str(e)}'
        }, status=500)


@superviseur_preparation_required
def commandes_envoi_historique(request, envoi_id):
    """Récupérer les commandes d'un envoi clôturé pour l'historique"""
    try:
        envoi = Envoi.objects.get(id=envoi_id, status=False)  # Seulement les envois clôturés
        
        commandes = Commande.objects.filter(
            envoi =envoi
        ).select_related('client', 'ville', 'ville__region').prefetch_related(
            'etats', 
            'paniers__article', 
            'paniers__variante__couleur', 
            'paniers__variante__pointure'
        ).distinct()
        
        commandes_data = []
        for commande in commandes:
            try:
                
                # Récupérer l'état actuel de manière sécurisée
                etat_actuel = commande.etat_actuel
                libelle_etat = None
                if etat_actuel and etat_actuel.enum_etat:
                    libelle_etat = etat_actuel.enum_etat.libelle
                else:
                    print(f"DEBUG: Aucun état actuel trouvé pour la commande {commande.id}")
                
                # Récupérer les paniers avec leurs variantes
                paniers_data = []
                for panier in commande.paniers.all():
                    panier_info = {
                        'id': panier.id,
                        'article': {
                            'id': panier.article.id,
                            'nom': panier.article.nom,
                            'reference': panier.article.reference,
                        },
                        'quantite': panier.quantite,
                        'sous_total': float(panier.sous_total) if panier.sous_total else 0,
                        'variante': None
                    }
                    
                    # Ajouter les informations de variante si elle existe
                    if panier.variante:
                        panier_info['variante'] = {
                            'id': panier.variante.id,
                            'reference_variante': panier.variante.reference_variante,
                            'couleur': panier.variante.couleur.nom if panier.variante.couleur else None,
                            'pointure': panier.variante.pointure.pointure if panier.variante.pointure else None,
                            'qte_disponible': panier.variante.qte_disponible,
                        }
                    else:
                        print(f"DEBUG: Aucune variante pour panier {panier.id}")
                    
                    paniers_data.append(panier_info)
                
                commandes_data.append({
                    'id': commande.id,
                    'id_yz': commande.id_yz,
                    'num_cmd': commande.num_cmd,
                    'total_cmd': float(commande.total_cmd) if commande.total_cmd else 0,
                    'client': {
                        'nom': commande.client.nom if commande.client else None,
                        'prenom': commande.client.prenom if commande.client else None,
                        'email': commande.client.email if commande.client else None,
                        'numero_tel': commande.client.numero_tel if commande.client else None,
                    },
                    'ville': {
                        'nom_ville': commande.ville.nom if commande.ville else None,
                    },
                    'etat_actuel': {
                        'enum_etat': {
                            'libelle': libelle_etat,
                        }
                    } if etat_actuel else None,
                    'paniers_data': paniers_data,
                })
            except Exception as e:
                # Si une commande spécifique pose problème, on la saute et on continue
                print(f"Erreur lors du traitement de la commande {commande.id}: {str(e)}")
                continue
        
        
        return JsonResponse({
            'success': True,
            'commandes': commandes_data,
            'envoi': {
                'id': envoi.id,
                'numero_envoi': envoi.numero_envoi,
                'region': envoi.region.nom_region,
                'nb_commandes': envoi.nb_commandes,
                'date_creation': envoi.date_creation.strftime('%d/%m/%Y'),
                'date_cloture': envoi.date_livraison_effective.strftime('%d/%m/%Y') if envoi.date_livraison_effective else None,
            }
        })
        
    except Envoi.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Envoi clôturé non trouvé'
        }, status=404)
    except Exception as e:
        import traceback
        print(f"Erreur dans commandes_envoi_historique: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la récupération des commandes: {str(e)}'
        }, status=500)


@csrf_exempt
@login_required
def historique_envois_view(request):
    """Page d'historique des envois clôturés."""
    from django.core.paginator import Paginator
    from datetime import datetime, timedelta
    
    # Récupérer les envois clôturés (status=False)
    envois_clotures = Envoi.objects.filter(
        status=False  # False = clôturé
    ).select_related('region', 'operateur_creation').order_by('-date_creation')
    
    # Mettre à jour le compteur de commandes pour chaque envoi clôturé
    for envoi in envois_clotures:
        # Pour les envois clôturés, on compte TOUTES les commandes de la région
        # peu importe leur état actuel
        nb_commandes = Commande.objects.filter(
            envoi=envoi
        ).count()
        
        # Mettre à jour seulement si le nombre a changé
        if envoi.nb_commandes != nb_commandes:
            envoi.nb_commandes = nb_commandes
            envoi.save(update_fields=['nb_commandes'])
    
    # Filtrage par période si demandé
    periode = request.GET.get('periode')
    if periode:
        today = datetime.now().date()
        if periode == '7j':
            date_debut = today - timedelta(days=7)
            envois_clotures = envois_clotures.filter(date_creation__gte=date_debut)
        elif periode == '30j':
            date_debut = today - timedelta(days=30)
            envois_clotures = envois_clotures.filter(date_creation__gte=date_debut)
        elif periode == '90j':
            date_debut = today - timedelta(days=90)
            envois_clotures = envois_clotures.filter(date_creation__gte=date_debut)
    
    # Filtrage par région si demandé
    region_filter = request.GET.get('region')
    if region_filter:
        envois_clotures = envois_clotures.filter(region__id=region_filter)
    
    # Pagination
    paginator = Paginator(envois_clotures, 12)  # 12 envois par page
    page_number = request.GET.get('page')
    envois = paginator.get_page(page_number)
    
    # Récupérer toutes les régions pour le filtre
    from parametre.models import Region
    regions = Region.objects.all().order_by('nom_region')
    
    context = {
        'envois': envois,
        'regions': regions,
        'periode_selected': periode,
        'region_selected': region_filter,
        'page_title': 'Historique des envois',
        'page_subtitle': 'Consulter les envois clôturés et exporter leurs données',
    }
    return render(request, 'Superpreparation/historique_envois.html', context)

@csrf_exempt
@login_required
def creer_envoi_region(request):

    try:
        from parametre.models import Operateur
        operateur = Operateur.objects.get(user=request.user, actif=True)
    except Operateur.DoesNotExist:
        print(f"🔍 DEBUG - Aucun profil opérateur trouvé")
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)

    try:
        from parametre.models import Region
        region_id = request.POST.get('region_id')

        if not region_id:
            return JsonResponse({'success': False, 'message': 'Région requise'}, status=400)

        region = Region.objects.get(id=region_id)

        # Vérifier qu'il n'y a pas déjà un envoi actif pour cette région
        envoi_actif_existant = Envoi.objects.filter(region=region, status=True).exists()
        if envoi_actif_existant:
            return JsonResponse({
                'success': False, 
                'message': f'Il existe déjà un envoi actif pour la région {region.nom_region}. Veuillez clôturer l\'envoi existant avant d\'en créer un nouveau.'
            }, status=400)

        # Générer un numéro d'envoi via le modèle Envoi
        from django.utils import timezone
        
        # Récupérer le profil opérateur
        try:
            from parametre.models import Operateur
            operateur_creation = Operateur.objects.get(user=request.user, actif=True)
        except Operateur.DoesNotExist:
            operateur_creation = None
            
        envoi = Envoi.objects.create(
            region=region,
            operateur_creation=operateur_creation,
            date_envoi=timezone.now().date(),
            date_livraison_prevue=timezone.now().date(),
        )

        return JsonResponse({'success': True, 'envoi_id': envoi.id, 'numero': envoi.numero_envoi})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@superviseur_preparation_required
def creer_envois_multiples(request):
    """Créer plusieurs envois en une seule opération pour les régions sélectionnées"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)
    
    try:
        import json
        from parametre.models import Region, Operateur
        from django.utils import timezone
        
        # Récupérer les IDs des régions sélectionnées
        data = json.loads(request.body)
        regions_ids = data.get('regions_ids', [])
        
        if not regions_ids:
            return JsonResponse({'success': False, 'message': 'Aucune région sélectionnée'}, status=400)
        
        # Récupérer le profil opérateur
        try:
            operateur_creation = Operateur.objects.get(user=request.user, actif=True)
        except Operateur.DoesNotExist:
            operateur_creation = None
        
        envois_crees = []
        erreurs = []
        
        for region_id in regions_ids:
            try:
                region = Region.objects.get(id=region_id)
                
                # Vérifier qu'il n'y a pas déjà un envoi actif pour cette région
                envoi_actif_existant = Envoi.objects.filter(region=region, status=True).exists()
                if envoi_actif_existant:
                    erreurs.append(f'Région {region.nom_region}: Un envoi actif existe déjà')
                    continue
                
                # Créer l'envoi
                envoi = Envoi.objects.create(
                    region=region,
                    operateur_creation=operateur_creation,
                    date_envoi=timezone.now().date(),
                    date_livraison_prevue=timezone.now().date(),
                )
                
                envois_crees.append({
                    'id': envoi.id,
                    'numero': envoi.numero_envoi,
                    'region': region.nom_region
                })
                
            except Region.DoesNotExist:
                erreurs.append(f'Région ID {region_id}: Région non trouvée')
            except Exception as e:
                erreurs.append(f'Région ID {region_id}: {str(e)}')
        
        # Préparer la réponse
        if envois_crees and not erreurs:
            return JsonResponse({
                'success': True,
                'message': f'{len(envois_crees)} envoi(s) créé(s) avec succès',
                'envois_crees': envois_crees
            })
        elif envois_crees and erreurs:
            return JsonResponse({
                'success': True,
                'message': f'{len(envois_crees)} envoi(s) créé(s), {len(erreurs)} erreur(s)',
                'envois_crees': envois_crees,
                'erreurs': erreurs
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Aucun envoi créé',
                'erreurs': erreurs
            }, status=400)
            
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Données JSON invalides'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@superviseur_preparation_required
def cloturer_envois_multiples(request):
    """Clôturer plusieurs envois en une seule opération"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)
    
    try:
        import json
        from django.utils import timezone
        from parametre.models import Operateur
        
        # Récupérer les IDs des envois sélectionnés
        data = json.loads(request.body)
        envois_ids = data.get('envois_ids', [])
        
        if not envois_ids:
            return JsonResponse({'success': False, 'message': 'Aucun envoi sélectionné'}, status=400)
        
        # Récupérer le profil opérateur
        try:
            operateur_cloture = Operateur.objects.get(user=request.user, actif=True)
        except Operateur.DoesNotExist:
            operateur_cloture = None
        
        envois_clotures = []
        erreurs = []
        
        for envoi_id in envois_ids:
            try:
                envoi = Envoi.objects.get(id=envoi_id)
                
                # Vérifier que l'envoi est encore actif
                if not envoi.status:
                    erreurs.append(f'Envoi {envoi.numero_envoi}: Déjà clôturé')
                    continue
                
                # Clôturer l'envoi
                envoi.status = False
                envoi.operateur_cloture = operateur_cloture
                envoi.date_cloture = timezone.now()
                envoi.save()
                
                envois_clotures.append({
                    'id': envoi.id,
                    'numero': envoi.numero_envoi,
                    'region': envoi.region.nom_region
                })
                
            except Envoi.DoesNotExist:
                erreurs.append(f'Envoi ID {envoi_id}: Envoi non trouvé')
            except Exception as e:
                erreurs.append(f'Envoi ID {envoi_id}: {str(e)}')
        
        # Préparer la réponse
        if envois_clotures and not erreurs:
            return JsonResponse({
                'success': True,
                'message': f'{len(envois_clotures)} envoi(s) clôturé(s) avec succès',
                'envois_clotures': envois_clotures
            })
        elif envois_clotures and erreurs:
            return JsonResponse({
                'success': True,
                'message': f'{len(envois_clotures)} envoi(s) clôturé(s), {len(erreurs)} erreur(s)',
                'envois_clotures': envois_clotures,
                'erreurs': erreurs
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Aucun envoi clôturé',
                'erreurs': erreurs
            }, status=400)
            
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Données JSON invalides'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required
def cloturer_envoi(request):
    """Clôturer un envoi (POST)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)

    try:
        envoi_id = request.POST.get('envoi_id')
        if not envoi_id:
            return JsonResponse({'success': False, 'message': 'ID d\'envoi requis'}, status=400)

        envoi = Envoi.objects.get(id=envoi_id)
        
        # Vérifier que l'envoi n'est pas déjà clôturé
        if not envoi.status:  # False = déjà clôturé
            return JsonResponse({'success': False, 'message': 'Cet envoi est déjà clôturé'}, status=400)

        from django.db import transaction
        from django.utils import timezone
        from commande.models import Commande, EtatCommande, EnumEtatCmd

        with transaction.atomic():
            # Passer toutes les commandes liées à l'envoi en "En livraison"
            # On cible explicitement les commandes dont l'état courant est "Préparée"
            commandes = Commande.objects.filter(
                envoi=envoi,
                etats__enum_etat__libelle='Préparée',
                etats__date_fin__isnull=True,
            ).distinct()
            if not commandes.exists():
                # Fallback: prendre les commandes 'Préparée' de la région de l'envoi
                commandes = (
                    Commande.objects.filter(
                        ville__region=envoi.region,
                        etats__enum_etat__libelle='Préparée',
                        etats__date_fin__isnull=True,
                    )
                    .select_related('client', 'ville')
                    .distinct()
                )
            etat_enum, _ = EnumEtatCmd.objects.get_or_create(
                libelle='Mise en distribution',
                defaults={'ordre': 17, 'couleur': '#8B5CF6'}
            )

            for commande in commandes:
                try:
                    # Lier la commande à cet envoi si ce n'est pas déjà fait
                    if commande.envoi_id != envoi.id:
                        commande.envoi = envoi
                        commande.save(update_fields=['envoi'])
                    
                    # Fermer l'état courant "Préparée" s'il existe
                    etat_actuel = commande.etats.filter(date_fin__isnull=True).order_by('-date_debut').first()
                    if etat_actuel and etat_actuel.enum_etat.libelle == 'Préparée':
                        etat_actuel.date_fin = timezone.now()
                        etat_actuel.save(update_fields=['date_fin'])

                    # Créer le nouvel état
                    EtatCommande.objects.create(
                        commande=commande,
                        enum_etat=etat_enum,
                        operateur=getattr(request.user, 'profil_operateur', None),
                        date_debut=timezone.now(),
                        commentaire=f"Envoi clôturé: {envoi.numero_envoi}"
                    )
                except Exception as state_err:
                    print(f"⚠️ Erreur transition état commande {commande.id}: {state_err}")
                    continue

            # Marquer l'envoi comme livré (clôturé)
            envoi.marquer_comme_livre(getattr(request.user, 'profil_operateur', None))

        return JsonResponse({
            'success': True, 
            'message': f'Envoi {envoi.numero_envoi} clôturé avec succès'
        })
    except Envoi.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Envoi non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


def rafraichir_articles_commande_prepa(request, commande_id):
    """Rafraîchir la section des articles de la commande en préparation"""
    print(f"🔄 Rafraîchissement des articles pour la commande {commande_id}")
    
    try:
        operateur = Operateur.objects.get(user=request.user, actif=True)
        
        # Vérifier le type d'opérateur
        if operateur.type_operateur not in ['SUPERVISEUR_PREPARATION', 'PREPARATION', 'ADMIN']:
            return JsonResponse({'error': 'Accès non autorisé.'}, status=403)
            
    except Operateur.DoesNotExist:
        # Fallback si pas de profil: autoriser selon groupes Django
        if not request.user.groups.filter(name__in=['superviseur', 'operateur_preparation']).exists():
            return JsonResponse({'error': 'Profil d\'opérateur de préparation non trouvé.'}, status=403)
        operateur = None
    
    try:
        commande = Commande.objects.get(id=commande_id)
        
        # Pour les superviseurs, on ne vérifie pas l'affectation spécifique
        # Ils peuvent accéder à toutes les commandes en préparation
        if operateur and operateur.type_operateur == 'SUPERVISEUR_PREPARATION':
            # Vérifier seulement que la commande est en préparation
            etat_preparation = commande.etats.filter(
                enum_etat__libelle__in=['En préparation', 'À imprimer'],
                date_fin__isnull=True
            ).first()
        else:
            # Pour les opérateurs normaux, vérifier l'affectation spécifique
            etat_preparation = commande.etats.filter(
                operateur=operateur,
                enum_etat__libelle__in=['En préparation', 'À imprimer'],
                date_fin__isnull=True
            ).first()
        
        if not etat_preparation:
            return JsonResponse({'error': 'Cette commande ne vous est pas affectée.'}, status=403)
        
        # Générer le HTML directement pour éviter les erreurs de template
        paniers = commande.paniers.select_related('article').all()
        html_rows = []
        
        for panier in paniers:
            # Utiliser le filtre Django pour calculer le prix selon la logique upsell
            from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
            
            prix_calcule = get_prix_upsell_avec_compteur(panier.article, commande.compteur)
            
            # Déterminer l'affichage selon le type d'article
            if panier.article.isUpsell:
                prix_class = "text-orange-600"
                upsell_badge = f'<span class="inline-flex items-center px-2 py-1 rounded-full text-xs font-bold bg-orange-100 text-orange-700 ml-2"><i class="fas fa-arrow-up mr-1"></i>Upsell</span>'
                
                # Déterminer le niveau d'upsell affiché
                if commande.compteur >= 4 and panier.article.prix_upsell_4 is not None:
                    niveau_upsell = 4
                elif commande.compteur >= 3 and panier.article.prix_upsell_3 is not None:
                    niveau_upsell = 3
                elif commande.compteur >= 2 and panier.article.prix_upsell_2 is not None:
                    niveau_upsell = 2
                elif commande.compteur >= 1 and panier.article.prix_upsell_1 is not None:
                    niveau_upsell = 1
                else:
                    niveau_upsell = 0
                
                if niveau_upsell > 0:
                    upsell_info = f'<div class="text-xs text-orange-600 flex items-center justify-center gap-1"><i class="fas fa-arrow-up"></i>Upsell Niveau {niveau_upsell}</div>'
                else:
                    upsell_info = f'<div class="text-xs text-orange-600 flex items-center justify-center gap-1"><i class="fas fa-arrow-up"></i>Upsell (Prix normal)</div>'
            else:
                # Prix normal
                prix_class = "text-gray-700"
                upsell_badge = ""
                upsell_info = ""
            
            sous_total = prix_calcule * panier.quantite
            
            html_row = f"""
            <tr data-panier-id="{panier.id}" data-article-id="{panier.article.id}">
                <td class="px-4 py-3">
                    <div class="flex items-center">
                        <div class="flex-1">
                            <div class="font-medium text-gray-900 flex items-center">
                                {panier.article.nom}
                                {upsell_badge}
                            </div>
                            <div class="text-sm text-gray-500">
                                Réf: {panier.article.reference or 'N/A'}
                                {f" · Réf variante: {getattr(panier.variante, 'reference', '')}" if hasattr(panier, 'variante') and getattr(panier, 'variante', None) and getattr(panier.variante, 'reference', None) else ''}
                                {f" · Couleur: {getattr(getattr(panier.variante,'couleur', None),'nom','')}" if hasattr(panier, 'variante') and getattr(panier, 'variante', None) and getattr(panier.variante, 'couleur', None) else ''}
                                {f" · Pointure: {getattr(getattr(panier.variante,'pointure', None),'pointure','')}" if hasattr(panier, 'variante') and getattr(panier, 'variante', None) and getattr(panier.variante, 'pointure', None) else ''}
                            </div>
                        </div>
                    </div>
                </td>
                <td class="px-4 py-3 text-center">
                    <div class="flex items-center justify-center space-x-2">
                        <button type="button" onclick="modifierQuantite({panier.id}, -1)" 
                                class="w-8 h-8 bg-gray-200 hover:bg-gray-300 rounded text-sm transition-colors">-</button>
                        <input type="number" id="quantite-{panier.id}" value="{panier.quantite}" min="1" 
                               class="w-16 p-2 border rounded-lg text-center" 
                               onchange="modifierQuantiteDirecte({panier.id}, this.value)">
                        <button type="button" onclick="modifierQuantite({panier.id}, 1)" 
                                class="w-8 h-8 bg-gray-200 hover:bg-gray-300 rounded text-sm transition-colors">+</button>
                    </div>
                </td>
                <td class="px-4 py-3 text-center">
                    <div class="font-medium {prix_class}" id="prix-unitaire-{panier.id}">
                        {prix_calcule:.2f} DH
                    </div>
                    {upsell_info}
                </td>
                <td class="px-4 py-3 text-center">
                    <div class="font-bold text-gray-900">
                        {sous_total:.2f} DH
                    </div>
                </td>
                <td class="px-4 py-3 text-center">
                    <button type="button" onclick="supprimerArticle({panier.id})" 
                            class="px-3 py-2 bg-red-500 hover:bg-red-600 text-white text-sm rounded-lg transition-colors">
                        <i class="text-white text-sm rounded-lg transition-colors">
                        <i class="fas fa-trash mr-1"></i>Supprimer
                    </button>
                </td>
            </tr>
            """
            html_rows.append(html_row)
        
        html = "".join(html_rows)
        
        # Calculer les statistiques upsell
        articles_upsell = commande.paniers.filter(article__isUpsell=True)
        total_quantite_upsell = articles_upsell.aggregate(
            total=Sum('quantite')
        )['total'] or 0
        
        response_data = {
            'success': True,
            'html': html,
            'articles_count': commande.paniers.count(),
            'total_commande': float(commande.total_cmd),
            'sous_total_articles': float(commande.sous_total_articles),
            'compteur': commande.compteur,
            'articles_upsell': articles_upsell.count(),
            'quantite_totale_upsell': total_quantite_upsell
        }
        
        print(f"✅ Rafraîchissement terminé - Articles: {response_data['articles_count']}, Compteur: {response_data['compteur']}")
        print(f"🔍 Détails du compteur: type={type(commande.compteur)}, valeur={commande.compteur}")
        if commande.compteur is None:
            print("⚠️ ATTENTION: Le compteur est None dans la base de données")
        elif commande.compteur == 0:
            print("ℹ️ Le compteur est 0 (normal si pas d'articles upsell)")
        else:
            print(f"✅ Le compteur a une valeur: {commande.compteur}")
        return JsonResponse(response_data)
        
    except Commande.DoesNotExist:
        print(f"❌ Commande {commande_id} non trouvée")
        return JsonResponse({'error': 'Commande non trouvée'}, status=404)
    except Exception as e:
        print(f"❌ Erreur lors du rafraîchissement: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Erreur interne: {str(e)}'}, status=500)

@superviseur_preparation_required
def ajouter_article_commande_prepa(request, commande_id):
    """Ajouter un article à la commande en préparation"""
    if request.method != 'POST':
        print("❌ Méthode non autorisée")
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        # Autoriser PREPARATION et SUPERVISEUR_PREPARATION
        operateur = Operateur.objects.get(user=request.user, actif=True)
        print(f"✅ Opérateur trouvé: {operateur.id} - Type: {operateur.type_operateur}")
        if operateur.type_operateur not in ['PREPARATION', 'SUPERVISEUR_PREPARATION']:
            print(f"❌ Type d'opérateur non autorisé: {operateur.type_operateur}")
            return JsonResponse({'error': "Accès réservé à l'équipe préparation"}, status=403)
    except Operateur.DoesNotExist:
        print("❌ Profil d'opérateur non trouvé")
        return JsonResponse({'error': 'Profil d\'opérateur non trouvé.'}, status=403)
    
    try:
        with transaction.atomic():
            print("🔧 Début de la transaction atomique")
            commande = Commande.objects.select_for_update().get(id=commande_id)
            print(f"✅ Commande trouvée: {commande.id} - ID YZ: {commande.id_yz}")
            
         
            article_id = request.POST.get('article_id')
            quantite = int(request.POST.get('quantite', 1))
            variante_id = request.POST.get('variante_id')

            print("[AJOUT VARIANTE] entrée:", {
                'commande_id': commande_id,
                'operateur': getattr(operateur, 'id', None),
                'article_id': article_id,
                'quantite': quantite,
                'variante_id': variante_id,
            })
            
            if not article_id or quantite <= 0:
                print(f"❌ Données invalides: article_id={article_id}, quantite={quantite}")
                return JsonResponse({'error': 'Données invalides'}, status=400)

            print(f"✅ Données reçues: article_id={article_id}, quantite={quantite}, variante_id={variante_id}")
            article = Article.objects.get(id=article_id)
            print(f"✅ Article trouvé: {article.id} - {article.nom}")
            variante_obj = None
            if variante_id:
                try:
                    from article.models import VarianteArticle
                    variante_obj = VarianteArticle.objects.get(id=int(variante_id), article=article)
                    print("[AJOUT VARIANTE] variante trouvée:", {
                        'id': variante_obj.id,
                        'couleur': getattr(getattr(variante_obj, 'couleur', None), 'nom', None),
                        'pointure': getattr(getattr(variante_obj, 'pointure', None), 'pointure', None),
                        'qte_disponible_avant': variante_obj.qte_disponible,
                    })
                except Exception:
                    variante_obj = None
                    print("[AJOUT VARIANTE] variante introuvable ou invalide", variante_id)
            
            # Décrémenter le stock et créer un mouvement
            print("[AJOUT VARIANTE] création mouvement stock", {
                'article': article.id,
                'quantite': quantite,
                'type': 'sortie',
                'variante': getattr(variante_obj, 'id', None),
            })
            creer_mouvement_stock(
                article=article, quantite=quantite, type_mouvement='sortie',
                commande=commande, operateur=operateur,
                commentaire=f'Ajout article pendant préparation cmd {commande.id_yz}',
                variante=variante_obj
            )
            
            # Vérifier si l'article existe déjà dans la commande
            filtre_panier = {'commande': commande, 'article': article}
            if hasattr(Panier, 'variante'):
                filtre_panier['variante'] = variante_obj
            panier_existant = Panier.objects.filter(**filtre_panier).first()
            print("[AJOUT VARIANTE] filtre panier:", {
                'commande': commande.id,
                'article': article.id,
                'variante': getattr(variante_obj, 'id', None),
                'existant': getattr(panier_existant, 'id', None)
            })
            
            if panier_existant:
                # Si l'article existe déjà, mettre à jour la quantité
                panier_existant.quantite += quantite
                panier_existant.save()
                panier = panier_existant
                print(f"[AJOUT VARIANTE] 🔄 panier existant {panier.id} mis à jour, nouvelle_quantite={panier.quantite}")
            else:
                # Si l'article n'existe pas, créer un nouveau panier
                create_kwargs = {
                    'commande': commande,
                    'article': article,
                    'quantite': quantite,
                    'sous_total': 0,
                }
                if hasattr(Panier, 'variante'):
                    create_kwargs['variante'] = variante_obj
                panier = Panier.objects.create(**create_kwargs)
                print(f"[AJOUT VARIANTE] ➕ nouveau panier créé id={panier.id}, article={article.id}, variante={getattr(variante_obj,'id',None)}, quantite={quantite}")
            
            # Recalculer le compteur après ajout (logique de confirmation)
            if article.isUpsell and hasattr(article, 'prix_upsell_1') and article.prix_upsell_1 is not None:
                # Compter la quantité totale d'articles upsell (après ajout)
                total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
                    total=Sum('quantite')
                )['total'] or 0
                
                # Le compteur ne s'incrémente qu'à partir de 2 unités d'articles upsell
                # 0-1 unités upsell → compteur = 0
                # 2+ unités upsell → compteur = total_quantite_upsell - 1
                if total_quantite_upsell >= 2:
                    commande.compteur = total_quantite_upsell - 1
                else:
                    commande.compteur = 0
                
                commande.save()
                
                # Recalculer TOUS les articles de la commande avec le nouveau compteur
                commande.recalculer_totaux_upsell()
            else:
                # Pour les articles normaux, juste calculer le sous-total
                from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
                prix_unitaire = get_prix_upsell_avec_compteur(article, commande.compteur)
                sous_total = prix_unitaire * panier.quantite
                panier.sous_total = float(sous_total)
                panier.save()
            
            # Recalculer le total
            commande.total_cmd = sum(p.sous_total for p in commande.paniers.all())
            commande.save()
            print("[AJOUT VARIANTE] totaux mis à jour:", {
                'commande_id': commande.id,
                'total_cmd': commande.total_cmd,
                'articles_count': commande.paniers.count(),
            })
            
            # Calculer les statistiques upsell
            articles_upsell = commande.paniers.filter(article__isUpsell=True)
            total_quantite_upsell = articles_upsell.aggregate(
                total=Sum('quantite')
            )['total'] or 0
            
            # Déterminer si c'était un ajout ou une mise à jour
            message = 'Article ajouté' if not panier_existant else f'Quantité mise à jour ({panier.quantite})'
            
            # Préparer les données de l'article pour le frontend
            article_data = {
                'panier_id': panier.id,
                'article_id': article.id,
                'nom': article.nom,
                'reference': article.reference,
                'couleur_fr': (variante_obj.couleur.nom if variante_obj and variante_obj.couleur else (article.couleur or '')),
                'couleur_ar': (variante_obj.couleur.nom if variante_obj and variante_obj.couleur else (article.couleur or '')),
                'pointure': (variante_obj.pointure.pointure if variante_obj and variante_obj.pointure else (article.pointure or '')),
                'quantite': panier.quantite,
                'prix': panier.sous_total / panier.quantite,  # Prix unitaire
                'sous_total': panier.sous_total,
                'is_upsell': article.isUpsell,
                'compteur': commande.compteur,
                'phase': article.phase,
                'has_promo_active': article.has_promo_active,
                'qte_disponible': variante_obj.qte_disponible if variante_obj else article.qte_disponible,
                'description': article.description or ''
            }
            
            response_data = {
                'success': True, 
                'message': message,
                'compteur': commande.compteur,
                'total_commande': float(commande.total_cmd),
                'was_update': panier_existant is not None,
                'new_quantity': panier.quantite,
                'article_data': article_data,
                'articles_count': commande.paniers.count(),
                'sous_total_articles': float(sum(p.sous_total for p in commande.paniers.all())),
                'articles_upsell': articles_upsell.count(),
                'quantite_totale_upsell': total_quantite_upsell
            }
            
            print("✅ ===== SUCCÈS ajouter_article_commande_prepa =====")
            print(f"📦 Réponse: {response_data}")
            
            return JsonResponse(response_data)
            
    except Article.DoesNotExist:
        print("❌ Article non trouvé")
        return JsonResponse({'error': 'Article non trouvé'}, status=404)
    except Exception as e:
        print(f"❌ ===== ERREUR ajouter_article_commande_prepa =====")
        print(f"❌ Exception: {str(e)}")
        import traceback
        print(f"❌ Traceback: {traceback.format_exc()}")
        return JsonResponse({'error': f'Erreur interne: {str(e)}'}, status=500)

@superviseur_preparation_required
def modifier_quantite_article_prepa(request, commande_id):
    """Modifier la quantité d'un article dans la commande en préparation"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='PREPARATION')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Profil d\'opérateur non trouvé.'}, status=403)

    try:
        with transaction.atomic():
            commande = Commande.objects.select_for_update().get(id=commande_id)
            
            # Vérifier l'affectation
            if not commande.etats.filter(operateur=operateur, enum_etat__libelle__in=['En préparation', 'À imprimer'], date_fin__isnull=True).exists():
                return JsonResponse({'error': 'Commande non affectée.'}, status=403)
            
            panier_id = request.POST.get('panier_id')
            nouvelle_quantite = int(request.POST.get('quantite', 1))

            panier = Panier.objects.get(id=panier_id, commande=commande)
            ancienne_quantite = panier.quantite
            article = panier.article
            difference = nouvelle_quantite - ancienne_quantite

            if difference > 0:
                creer_mouvement_stock(article, difference, 'sortie', commande, operateur, f'Ajustement qté cmd {commande.id_yz}')
            elif difference < 0:
                creer_mouvement_stock(article, abs(difference), 'entree', commande, operateur, f'Ajustement qté cmd {commande.id_yz}')

            panier.quantite = nouvelle_quantite
            
            # Recalculer le compteur si c'est un article upsell
            if article.isUpsell:
                # Compter la quantité totale d'articles upsell après modification
                total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
                    total=Sum('quantite')
                )['total'] or 0
                
                # Appliquer la logique : compteur = max(0, total_quantite_upsell - 1)
                if total_quantite_upsell >= 2:
                    commande.compteur = total_quantite_upsell - 1
                else:
                    commande.compteur = 0
                
                commande.save()
                
                # Recalculer TOUS les articles de la commande avec le nouveau compteur
                commande.recalculer_totaux_upsell()
            else:
                # Pour les articles normaux, juste calculer le sous-total
                from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
                prix_unitaire = get_prix_upsell_avec_compteur(article, commande.compteur)
                panier.sous_total = prix_unitaire * nouvelle_quantite
            panier.save()

            commande.total_cmd = sum(p.sous_total for p in commande.paniers.all())
            commande.save()

            # Calculer les statistiques upsell
            articles_upsell = commande.paniers.filter(article__isUpsell=True)
            total_quantite_upsell = articles_upsell.aggregate(
                total=Sum('quantite')
            )['total'] or 0

            return JsonResponse({
                'success': True, 
                'message': 'Quantité modifiée',
                'compteur': commande.compteur,
                'articles_upsell': articles_upsell.count(),
                'quantite_totale_upsell': total_quantite_upsell,
                'total_commande': float(commande.total_cmd),
                'sous_total_articles': float(commande.sous_total_articles)
            })

    except Panier.DoesNotExist:
        return JsonResponse({'error': 'Panier non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Erreur interne: {str(e)}'}, status=500)

@superviseur_preparation_required
def supprimer_article_commande_prepa(request, commande_id):
    """Supprimer un article de la commande en préparation"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='PREPARATION')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Profil d\'opérateur non trouvé.'}, status=403)

    try:
        with transaction.atomic():
            commande = Commande.objects.select_for_update().get(id=commande_id)
            
            # Vérifier l'affectation
            if not commande.etats.filter(operateur=operateur, enum_etat__libelle__in=['En préparation', 'À imprimer'], date_fin__isnull=True).exists():
                return JsonResponse({'error': 'Commande non affectée.'}, status=403)

            panier_id = request.POST.get('panier_id')
            panier = Panier.objects.get(id=panier_id, commande=commande)
            quantite_supprimee = panier.quantite
            article = panier.article
            
            creer_mouvement_stock(article, quantite_supprimee, 'entree', commande, operateur, f'Suppression article cmd {commande.id_yz}')
            
            # Sauvegarder l'info avant suppression
            etait_upsell = panier.article.isUpsell
            
            # Supprimer l'article
            panier.delete()

            # Recalculer le compteur après suppression (logique de confirmation)
            if etait_upsell:
                # Compter la quantité totale d'articles upsell restants (après suppression)
                total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
                    total=Sum('quantite')
                )['total'] or 0
                
                # Appliquer la logique : compteur = max(0, total_quantite_upsell - 1)
                if total_quantite_upsell >= 2:
                    commande.compteur = total_quantite_upsell - 1
                else:
                    commande.compteur = 0
                
                commande.save()
                
                # Recalculer TOUS les articles de la commande avec le nouveau compteur
                commande.recalculer_totaux_upsell()

            commande.total_cmd = sum(p.sous_total for p in commande.paniers.all())
            commande.save()

            # Calculer les statistiques upsell
            articles_upsell = commande.paniers.filter(article__isUpsell=True)
            total_quantite_upsell = articles_upsell.aggregate(
                total=Sum('quantite')
            )['total'] or 0

            return JsonResponse({
                'success': True, 
                'message': 'Article supprimé',
                'compteur': commande.compteur,
                'articles_upsell': articles_upsell.count(),
                'quantite_totale_upsell': total_quantite_upsell,
                'total_commande': float(commande.total_cmd),
                'sous_total_articles': float(commande.sous_total_articles)
            })

    except Panier.DoesNotExist:
        return JsonResponse({'error': 'Panier non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Erreur interne: {str(e)}'}, status=500)

@superviseur_preparation_required
def api_panier_commande_livraison(request, commande_id):
    """API pour récupérer le panier d'une commande pour les opérateurs de livraison"""
    try:
        # Vérifier que l'utilisateur est un opérateur de livraison
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Accès non autorisé'})
    
    # Récupérer la commande
    try:
        commande = Commande.objects.get(id=commande_id)
    except Commande.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Commande non trouvée'})
    
    # Récupérer les paniers de la commande
    paniers = commande.paniers.all().select_related('article')
    
    paniers_data = []
    for panier in paniers:
        paniers_data.append({
            'id': panier.id,
            'article_id': panier.article.id,
            'nom': panier.article.nom,
            'reference': panier.article.reference,
            'couleur': panier.article.couleur,
            'pointure': panier.article.pointure,
            'prix_unitaire': float(panier.article.prix_unitaire),
            'quantite': panier.quantite,
            'sous_total': float(panier.sous_total),
            'qte_disponible': panier.article.qte_disponible,
        })
    
    return JsonResponse({
        'success': True,
        'paniers': paniers_data,
        'total_commande': float(commande.total_cmd)
    })

@superviseur_preparation_required
def api_articles_commande_livree_partiellement(request, commande_id):
    """API pour récupérer les détails des articles d'une commande livrée partiellement"""
    import json
    from article.models import Article
    from commande.models import Commande, EtatCommande, EnumEtatCmd, Operation
    from parametre.models import Operateur

    try:
        # Vérifier que l'utilisateur est un opérateur de préparation
        operateur = Operateur.objects.get(user=request.user, type_operateur='PREPARATION')
    except Operateur.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Accès non autorisé'})
    
    # Récupérer la commande
    try:
        commande = Commande.objects.get(id=commande_id)
    except Commande.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Commande non trouvée'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur lors de la récupération de la commande: {str(e)}'})
    
    # Analyser les articles pour les commandes livrées partiellement
    articles_livres = []
    articles_renvoyes = []
    
    # Récupérer tous les états de la commande
    etats_commande = commande.etats.all().select_related('enum_etat', 'operateur').order_by('date_debut')
    
    # Déterminer l'état actuel
    etat_actuel = etats_commande.filter(date_fin__isnull=True).first()
    
    # Récupérer l'état précédent pour comprendre d'où vient la commande
    etat_precedent = None
    if etat_actuel:
        # Trouver l'état précédent (le dernier état terminé avant l'état actuel)
        for etat in reversed(etats_commande):
            if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                if etat.enum_etat.libelle not in ['À imprimer', 'En préparation']:
                    etat_precedent = etat
                    break
    
    # Vérifier si c'est une commande livrée partiellement
    if etat_actuel and etat_actuel.enum_etat.libelle == 'Livrée Partiellement':
        # Les articles dans cette commande sont ceux qui ont été livrés partiellement
        for panier in commande.paniers.all():
            articles_livres.append({
                'article_id': panier.article.id,
                'nom': panier.article.nom,
                'reference': panier.article.reference,
                'couleur': panier.article.couleur,
                'pointure': panier.article.pointure,
                'quantite_livree': panier.quantite,
                'prix': float(panier.article.prix_unitaire),
                'sous_total': float(panier.sous_total)
            })
        
        # Chercher la commande de renvoi associée
        commande_renvoi = Commande.objects.filter(
            num_cmd__startswith=f"RENVOI-{commande.num_cmd}",
            client=commande.client
        ).first()
        
        if commande_renvoi:
            # Récupérer l'état des articles renvoyés depuis l'opération de livraison partielle
            etat_articles_renvoyes = {}
            operation_livraison_partielle = commande.operations.filter(
                type_operation='LIVRAISON_PARTIELLE'
            ).order_by('-date_operation').first()
            if operation_livraison_partielle:
                try:
                    details = json.loads(operation_livraison_partielle.conclusion)
                    if 'recap_articles_renvoyes' in details:
                        for item in details['recap_articles_renvoyes']:
                            etat_articles_renvoyes[item['article_id']] = item['etat']
                except Exception:
                    pass
            if commande_renvoi:
                for panier_renvoi in commande_renvoi.paniers.all():
                    etat = etat_articles_renvoyes.get(panier_renvoi.article.id, 'bon')
                    articles_renvoyes.append({
                        'article_id': panier_renvoi.article.id,
                        'nom': panier_renvoi.article.nom,
                        'reference': panier_renvoi.article.reference,
                        'couleur': panier_renvoi.article.couleur,
                        'pointure': panier_renvoi.article.pointure,
                        'quantite': panier_renvoi.quantite,
                        'prix': float(panier_renvoi.article.prix_unitaire),
                        'sous_total': float(panier_renvoi.sous_total),
                        'etat': etat
                    })
    # Vérifier si c'est une commande renvoyée après livraison partielle
    elif etat_precedent and etat_precedent.enum_etat.libelle == 'Livrée Partiellement':
        # Chercher la commande originale qui a été livrée partiellement
        commande_originale = Commande.objects.filter(
            num_cmd=commande.num_cmd.replace('RENVOI-', ''),
            client=commande.client
        ).first()
        # Récupérer l'état des articles renvoyés depuis l'opération de livraison partielle
        etat_articles_renvoyes = {}
        if commande_originale:
            operation_livraison_partielle = commande_originale.operations.filter(
                type_operation='LIVRAISON_PARTIELLE'
            ).order_by('-date_operation').first()
            if operation_livraison_partielle:
                try:
                    details = json.loads(operation_livraison_partielle.conclusion)
                    if 'recap_articles_renvoyes' in details:
                        for item in details['recap_articles_renvoyes']:
                            etat_articles_renvoyes[item['article_id']] = item['etat']
                except Exception:
                    pass
        if commande_originale:
            # Les articles dans cette commande de renvoi sont ceux qui ont été renvoyés
            for panier in commande_originale.paniers.all():
                etat = etat_articles_renvoyes.get(panier.article.id, 'bon')
                articles_renvoyes.append({
                    'article': panier.article,
                    'quantite': panier.quantite,
                    'prix': panier.article.prix_unitaire,
                    'sous_total': panier.sous_total,
                    'etat': etat
                })
    
    # Récupérer les détails de la livraison partielle
    date_livraison_partielle = None
    commentaire_livraison_partielle = None
    operateur_livraison = None
    
    if etat_actuel and etat_actuel.enum_etat.libelle == 'Livrée Partiellement':
        date_livraison_partielle = etat_actuel.date_debut
        commentaire_livraison_partielle = etat_actuel.commentaire
        operateur_livraison = etat_actuel.operateur
    elif etat_precedent and etat_precedent.enum_etat.libelle == 'Livrée Partiellement':
        date_livraison_partielle = etat_precedent.date_debut
        commentaire_livraison_partielle = etat_precedent.commentaire
        operateur_livraison = etat_precedent.operateur
    
    try:
        return JsonResponse({
            'success': True,
            'commande': {
                'id': commande.id,
                'id_yz': commande.id_yz,
                'num_cmd': commande.num_cmd,
                'total_cmd': float(commande.total_cmd),
                'date_livraison_partielle': date_livraison_partielle.isoformat() if date_livraison_partielle else None,
                'commentaire_livraison_partielle': commentaire_livraison_partielle,
                'operateur_livraison': {
                    'nom': operateur_livraison.nom_complet if operateur_livraison else None,
                    'email': operateur_livraison.mail if operateur_livraison else None
                } if operateur_livraison else None
            },
            'articles_livres': articles_livres,
            'articles_renvoyes': articles_renvoyes,
            'total_articles_livres': len(articles_livres),
            'total_articles_renvoyes': len(articles_renvoyes)
        })
    except Exception as e:
        print(f"Erreur lors de la génération de la réponse JSON: {e}")
        return JsonResponse({
            'success': False, 
            'message': f'Erreur lors de la génération de la réponse: {str(e)}'
        })
    
def api_prix_upsell_articles(request, commande_id):
    """API pour récupérer les prix upsell mis à jour des articles"""
    print(f"🔄 Récupération des prix upsell pour la commande {commande_id}")

    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='PREPARATION')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Profil d\'opérateur de préparation non trouvé.'}, status=403)

    try:
        commande = Commande.objects.get(id=commande_id)
        
        # Vérifier que la commande est affectée à cet opérateur
        etat_preparation = commande.etats.filter(
            operateur=operateur,
            enum_etat__libelle__in=['En préparation', 'À imprimer'],
            date_fin__isnull=True
        ).first()
        
        if not etat_preparation:
            return JsonResponse({'error': 'Cette commande ne vous est pas affectée.'}, status=403)
        
        # Récupérer tous les articles du panier avec leurs prix mis à jour
        paniers = commande.paniers.select_related('article').all()
        prix_articles = {}
        
        for panier in paniers:
            # Utiliser le filtre Django pour calculer le prix selon la logique upsell
            from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
            
            prix_calcule = get_prix_upsell_avec_compteur(panier.article, commande.compteur)
            print(f"   - prix_calcule: {prix_calcule}")
            
            # Convertir en float pour éviter les problèmes de sérialisation JSON
            prix_calcule = float(prix_calcule) if prix_calcule is not None else 0.0
            
            # Déterminer le type de prix et les informations
            if panier.article.isUpsell:
                prix_type = "upsell"
                
                # Déterminer le niveau d'upsell affiché
                if commande.compteur >= 4 and panier.article.prix_upsell_4 is not None:
                    niveau_upsell = 4
                elif commande.compteur >= 3 and panier.article.prix_upsell_3 is not None:
                    niveau_upsell = 3
                elif commande.compteur >= 2 and panier.article.prix_upsell_2 is not None:
                    niveau_upsell = 2
                elif commande.compteur >= 1 and panier.article.prix_upsell_1 is not None:
                    niveau_upsell = 1
                else:
                    niveau_upsell = 0
                
                if niveau_upsell > 0:
                    libelle = f"Upsell Niveau {niveau_upsell}"
                else:
                    libelle = "Upsell (Prix normal)"
                    
                icone = "fas fa-arrow-up"
            else:
                prix_type = "normal"
                libelle = "Prix normal"
                icone = "fas fa-tag"
                niveau_upsell = 0
            
            prix_articles[panier.id] = {
                'prix': prix_calcule,
                'type': prix_type,
                'libelle': libelle,
                'icone': icone,
                'niveau_upsell': niveau_upsell,
                'is_upsell': panier.article.isUpsell,
                'sous_total': float(prix_calcule * panier.quantite)
            }
        
        return JsonResponse({
            'success': True,
            'prix_articles': prix_articles,
            'compteur': commande.compteur,
            'message': f'Prix mis à jour pour {len(prix_articles)} articles'
        })
        
    except Commande.DoesNotExist:
        return JsonResponse({'error': 'Commande non trouvée.'}, status=404)
    except Exception as e:
        print(f"❌ Erreur lors de la récupération des prix upsell: {e}")
        return JsonResponse({'error': f'Erreur serveur: {str(e)}'}, status=500)

@superviseur_preparation_required
def commandes_confirmees(request):
    """Page des commandes confirmées"""
    from django.utils import timezone
    from datetime import datetime, timedelta

    # Récupérer toutes les commandes confirmées (état actif sans date_fin)
    commandes_confirmees = Commande.objects.filter(
        etats__enum_etat__libelle='Confirmée',
        etats__date_fin__isnull=True  # État actif (en cours)
    ).select_related('client', 'ville', 'ville__region').prefetch_related('etats', 'operations').distinct()

    # Recherche
    search_query = request.GET.get('search', '')
    if search_query:
        commandes_confirmees = commandes_confirmees.filter(
            Q(id_yz__icontains=search_query) |
            Q(num_cmd__icontains=search_query) |
            Q(client__nom__icontains=search_query) |
            Q(client__prenom__icontains=search_query) |
            Q(client__numero_tel__icontains=search_query) |
            Q(etats__operateur__nom__icontains=search_query) |
            Q(etats__operateur__prenom__icontains=search_query)
        ).distinct()

    # Tri par date de confirmation (plus récentes en premier)
    commandes_confirmees = commandes_confirmees.order_by('-etats__date_debut')

    # Créer une copie des données non paginées pour les statistiques AVANT la pagination
    commandes_non_paginees = commandes_confirmees

    # Paramètres de pagination flexible
    items_per_page = request.GET.get('items_per_page', 25)
    start_range = request.GET.get('start_range')
    end_range = request.GET.get('end_range')

    # Gestion de la pagination flexible
    if start_range and end_range:
        try:
            start_range = int(start_range)
            end_range = int(end_range)
            if start_range > 0 and end_range >= start_range:
                # Pagination par plage personnalisée
                commandes_confirmees = commandes_confirmees[start_range-1:end_range]
                paginator = Paginator(commandes_confirmees, end_range - start_range + 1)
                page_obj = paginator.get_page(1)
            else:
                # Plage invalide, utiliser la pagination normale
                items_per_page = 25
                paginator = Paginator(commandes_confirmees, items_per_page)
                page_number = request.GET.get('page', 1)
                page_obj = paginator.get_page(page_number)
        except (ValueError, TypeError):
            # Erreur de conversion, utiliser la pagination normale
            items_per_page = 25
            paginator = Paginator(commandes_confirmees, items_per_page)
            page_number = request.GET.get('page', 1)
            page_obj = paginator.get_page(page_number)
    else:
        # Pagination normale
        if items_per_page == 'all':
            # Afficher toutes les commandes
            paginator = Paginator(commandes_confirmees, commandes_confirmees.count())
            page_obj = paginator.get_page(1)
        else:
            try:
                items_per_page = int(items_per_page)
                if items_per_page <= 0:
                    items_per_page = 25
            except (ValueError, TypeError):
                items_per_page = 25
            
            paginator = Paginator(commandes_confirmees, items_per_page)
            page_number = request.GET.get('page', 1)
            page_obj = paginator.get_page(page_number)

    # Statistiques
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    # Confirmées aujourd'hui
    confirmees_aujourd_hui = Commande.objects.filter(
        etats__enum_etat__libelle='Confirmée',
        etats__date_fin__isnull=True,  # État actif
        etats__date_debut__date=today
    ).distinct().count()

    # Confirmées cette semaine
    confirmees_semaine = Commande.objects.filter(
        etats__enum_etat__libelle='Confirmée',
        etats__date_fin__isnull=True,  # État actif
        etats__date_debut__date__gte=week_start
    ).distinct().count()

    # Confirmées ce mois
    confirmees_mois = Commande.objects.filter(
        etats__enum_etat__libelle='Confirmée',
        etats__date_fin__isnull=True,  # État actif
        etats__date_debut__date__gte=month_start
    ).distinct().count()

    # Total des commandes confirmées (utiliser les données non paginées)
    total_confirmees = commandes_non_paginees.count()

    # Montant total des commandes confirmées (utiliser les données non paginées)
    montant_total = commandes_non_paginees.aggregate(total=Sum('total_cmd'))['total'] or 0

    # Ajouter l'état de confirmation à chaque commande
    for commande in page_obj:
        etats_commande = list(commande.etats.all().order_by('date_debut'))
        
        # Trouver l'état de confirmation (le premier état "Confirmée")
        etat_confirmation = None
        for etat in etats_commande:
            if etat.enum_etat.libelle == 'Confirmée':
                etat_confirmation = etat
                break
        
        commande.etat_confirmation = etat_confirmation

    # Récupérer les opérateurs de préparation actifs
    operateurs_preparation = Operateur.objects.filter(
        type_operateur='PREPARATION',
        actif=True
    ).order_by('nom', 'prenom')

    # Vérifier si c'est une requête AJAX
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Rendre les templates partiels pour AJAX
        html_table_body = render_to_string('Superpreparation/partials/_confirmees_table_body.html', {
            'page_obj': page_obj
        }, request=request)
        
        html_pagination = render_to_string('Superpreparation/partials/_confirmees_pagination.html', {
            'page_obj': page_obj,
            'search_query': search_query,
            'items_per_page': items_per_page,
            'start_range': start_range,
            'end_range': end_range
        }, request=request)
        
        html_pagination_info = render_to_string('Superpreparation/partials/_confirmees_pagination_info.html', {
            'page_obj': page_obj
        }, request=request)
        
        return JsonResponse({
            'success': True,
            'html_table_body': html_table_body,
            'html_pagination': html_pagination,
            'html_pagination_info': html_pagination_info,
            'total_count': commandes_confirmees.count()
        })

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'total_confirmees': total_confirmees,
        'confirmees_aujourd_hui': confirmees_aujourd_hui,
        'confirmees_semaine': confirmees_semaine,
        'confirmees_mois': confirmees_mois,
        'montant_total': montant_total,
        'operateurs_preparation': operateurs_preparation,
        'page_title': 'Suivi des Commandes Confirmées',
        'page_subtitle': 'Suivi et gestion des commandes validées prêtes pour la préparation',
        'items_per_page': items_per_page,
        'start_range': start_range,
        'end_range': end_range,
    }

    return render(request, 'Superpreparation/commande_confirmees.html', context)

@login_required
def get_article_variants(request, article_id):
    """Récupère toutes les variantes disponibles d'un article donné"""
    try:
        from article.models import Article, VarianteArticle
        
        print(f"🔍 Recherche des variantes pour l'article ID: {article_id} (type: {type(article_id)})")
        
        # D'abord, vérifier si l'article existe (avec ou sans le filtre actif=True)
        try:
            article_test = Article.objects.get(id=article_id)
            print(f"📋 Article trouvé mais actif={article_test.actif}")
        except Article.DoesNotExist:
            print(f"❌ Aucun article trouvé avec l'ID {article_id}")
            # Lister quelques articles pour debug
            articles_existants = Article.objects.all()[:5]
            print("📚 Articles existants (premiers 5):")
            for art in articles_existants:
                print(f"   - ID: {art.id}, Nom: {art.nom}, Actif: {art.actif}")
        
        # Vérifier que l'article existe ET est actif
        article = get_object_or_404(Article, id=article_id, actif=True)
        print(f"✅ Article trouvé: {article.nom}")
        
        # Récupérer toutes les variantes actives de cet article
        variantes = VarianteArticle.objects.filter(
            article=article,
            actif=True
        ).select_related('couleur', 'pointure').order_by('couleur__nom', 'pointure__ordre')
        
        print(f"📊 Nombre de variantes trouvées: {variantes.count()}")
        
        # Construire la liste des variantes avec leurs informations
        variants_data = []
        for variante in variantes:
           
            variant_info = {
                'id': variante.id,
                'couleur': variante.couleur.nom if variante.couleur else None,
                'pointure': variante.pointure.pointure if variante.pointure else None,
                'taille': None,  # À adapter selon votre modèle si vous avez des tailles
                'stock': variante.qte_disponible,
                'qte_disponible': variante.qte_disponible,
                'prix_unitaire': float(variante.prix_unitaire),
                'prix_actuel': float(variante.prix_actuel),
                'reference_variante': variante.reference_variante,
                'est_disponible': variante.est_disponible
            }
            variants_data.append(variant_info)
        
        response_data = {
            'success': True,
            'variants': variants_data,
            'article': {
                'id': article.id,
                'nom': article.nom,
                'reference': article.reference
            }
        }
        
        print(f"📤 Réponse envoyée: {len(variants_data)} variantes")
        return JsonResponse(response_data)
        
    except Article.DoesNotExist:
        print(f"❌ Article {article_id} non trouvé")
        return JsonResponse({
            'success': False,
            'error': 'Article non trouvé'
        }, status=404)
        
    except Exception as e:
        import traceback
        print(f"❌ Erreur dans get_article_variants: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': 'Erreur lors de la récupération des variantes'
        }, status=500)


def generate_barcode_for_commande(commande_id_yz):
    """Fonction utilitaire pour générer le code-barres d'une commande"""
    try:
        barcode_data = str(commande_id_yz)
        print(f"📊 Génération du code-barres: {barcode_data}")
        
        # Créer un code-barres avec les mêmes options que Prepacommande
        code128 = barcode.get_barcode_class("code128")
        barcode_instance = code128(barcode_data, writer=ImageWriter())
        buffer = BytesIO()
        barcode_instance.write(
            buffer,
            options={
                "write_text": False,
                "module_height": 4.0,  # Hauteur augmentée pour meilleure lisibilité
                "module_width": 0.15,  # Largeur augmentée pour impression claire
                "quiet_zone": 2.0,     # Zone de silence autour du code-barres
            },
        )
        barcode_base64 = base64.b64encode(buffer.getvalue()).decode()
        print(f"✅ Code-barres généré avec succès (dimensions optimisées)")
        return barcode_base64
    except Exception as barcode_error:
            print(f"❌ Erreur lors de la génération du code-barres: {str(barcode_error)}")
    return ""

def generate_qr_code_for_commande(commande_id_yz):
    """Fonction utilitaire pour générer le code QR d'une commande"""
    try:
        import qrcode
        qr_data = str(commande_id_yz)
        print(f"📊 Génération du code QR: {qr_data}")
        
        # Créer un code QR avec des paramètres optimisés pour l'impression
        qr = qrcode.QRCode(
            version=1,  # Version 1 (21x21 modules)
            error_correction=qrcode.constants.ERROR_CORRECT_L,  # Niveau de correction d'erreur bas
            box_size=10,  # Taille de chaque module (pixel)
            border=4,  # Bordure autour du QR code
        )
        
        # Ajouter les données
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        # Créer l'image
        qr_image = qr.make_image(fill_color="black", back_color="white")
        
        # Créer un buffer pour l'image
        buffer = BytesIO()
        qr_image.save(buffer, format='PNG')
        
        # Convertir en base64
        qr_base64 = base64.b64encode(buffer.getvalue()).decode()
        print(f"✅ Code QR généré avec succès")
        return qr_base64
        
    except ImportError:
        print("❌ Erreur: La bibliothèque qrcode n'est pas installée")
        return ""
    except Exception as qr_error:
        print(f"❌ Erreur lors de la génération du code QR: {str(qr_error)}")
        return ""

@superviseur_preparation_required
def api_ticket_commande_new(request):
    """API pour récupérer le contenu HTML du nouveau ticket de commande (format 10x10)"""
    try:
        commande_id = request.GET.get('id')
        if not commande_id:
            return JsonResponse({'error': 'ID de la commande requis'}, status=400)
        
        print(f"🔍 Génération du ticket pour la commande: {commande_id}")
        
        # Récupérer la commande
        try:
            commande = Commande.objects.filter(
                id_yz=commande_id,
                etats__enum_etat__libelle='Confirmée'
            ).distinct().first()
            
            if not commande:
                print(f"❌ Commande {commande_id} non trouvée ou non confirmée")
                return JsonResponse({'error': f'Commande {commande_id} non trouvée ou non confirmée'}, status=404)
            
            print(f"✅ Commande trouvée: {commande.id_yz}")
        except Exception as e:
            print(f"❌ Erreur lors de la recherche de la commande {commande_id}: {str(e)}")
            return JsonResponse({'error': f'Erreur lors de la recherche de la commande {commande_id}'}, status=500)
        
        # Générer le code QR
        qr_base64 = generate_qr_code_for_commande(commande.id_yz)
        
        # Préparer les articles
        articles_data = []
        paniers = commande.paniers.all().select_related('article', 'variante__couleur', 'variante__pointure')
        
        for panier in paniers:
            article = panier.article
            variante = panier.variante
            
            # Calculer le prix unitaire
            prix_unitaire = panier.sous_total / panier.quantite if panier.quantite > 0 else 0
            
            article_data = {
                'nom': article.nom or "Article sans nom",
                'couleur': variante.couleur.nom if variante and variante.couleur else "Standard",
                'pointure': variante.pointure.pointure if variante and variante.pointure else "Standard",
                'quantite': panier.quantite,
                'prix': f"{prix_unitaire:.2f}"
            }
            articles_data.append(article_data)
        
        # Préparer les données de la commande
        commande_data = {
            'id_yz': commande.id_yz,
            'client_nom': f"{commande.client.nom} {commande.client.prenom}" if commande.client else "Client non défini",
            'client_telephone': commande.client.numero_tel if commande.client else "",
            'client_adresse': commande.adresse or "",
            'ville_client': commande.ville.nom if commande.ville else commande.ville_init or "",
            'total': f"{commande.total_cmd:.2f}",
            'frais_livraison': commande.frais_livraison,
            'date_commande': commande.date_creation,
            'qr_code': qr_base64,
            'articles': articles_data
        }
        
        # Rendre le template HTML
        ticket_html = render_to_string(
            'Superpreparation/partials/_ticket_commande_new.html',
            {'commande': commande_data}
        )
        
        return JsonResponse({
            'success': True,
            'html': ticket_html,
            'commande': commande_data
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Erreur dans api_ticket_commande_new: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la génération du ticket: {str(e)}'
        }, status=500)

def api_qr_codes_articles(request):
    """API pour récupérer les QR codes des articles d'une commande (format 10x10)"""
    try:
        commande_id = request.GET.get('id')
        if not commande_id:
            return JsonResponse({'error': 'ID de la commande requis'}, status=400)
        
        print(f"🔍 Génération des QR codes pour la commande: {commande_id}")
        
        # Récupérer la commande
        try:
            commande = Commande.objects.filter(
                id_yz=commande_id,
                etats__enum_etat__libelle='Confirmée'
            ).distinct().first()
            
            if not commande:
                print(f"❌ Commande {commande_id} non trouvée ou non confirmée")
                return JsonResponse({'error': f'Commande {commande_id} non trouvée ou non confirmée'}, status=404)
            
            print(f"✅ Commande trouvée: {commande.id_yz}")
        except Exception as e:
            print(f"❌ Erreur lors de la recherche de la commande {commande_id}: {str(e)}")
            return JsonResponse({'error': f'Erreur lors de la recherche de la commande {commande_id}'}, status=500)
        
        # Récupérer les articles du panier
        paniers = commande.paniers.all().select_related('article', 'variante__couleur', 'variante__pointure')
        
        if not paniers.exists():
            return JsonResponse({'error': 'Aucun article trouvé dans cette commande'}, status=404)
        
        # Générer les QR codes pour chaque article
        qr_codes_html = []
        
        for panier in paniers:
            article = panier.article
            variante = panier.variante
            
            # Créer le texte du QR code : référence_article|CMD-numéro_commande
            qr_text = f"{article.reference}|CMD-{commande.id_yz}"
            
            # Générer le QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(qr_text)
            qr.make(fit=True)
            
            # Créer l'image du QR code
            qr_image = qr.make_image(fill_color="black", back_color="white")
                
                # Convertir en base64
            buffer = BytesIO()
            qr_image.save(buffer, format='PNG')
            qr_base64 = base64.b64encode(buffer.getvalue()).decode()
            
            # Créer le HTML pour ce QR code
            qr_html = f'''
            <div class="qr-code-container">
                <img src="data:image/png;base64,{qr_base64}" alt="QR Code" class="qr-code">
                <div class="article-info">
                    <div class="article-ref">{article.reference}</div>
                    <div class="commande-ref">CMD-{commande.id_yz}</div>
                </div>
            </div>
            '''
            qr_codes_html.append(qr_html)
        
        # Combiner tous les QR codes
        all_qr_codes_html = ''.join(qr_codes_html)
        
        return JsonResponse({
            'success': True,
            'html': all_qr_codes_html,
            'count': len(qr_codes_html)
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Erreur dans api_qr_codes_articles: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la génération des QR codes: {str(e)}'
        }, status=500)


@superviseur_preparation_required
@transaction.atomic
def api_finaliser_preparation(request, commande_id):
    """API pour finaliser la préparation d'une commande (changer l'état vers "Préparée")
    Seul le superviseur peut finaliser la préparation"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)

    try:
        # Récupérer la commande
        commande = get_object_or_404(Commande, id=commande_id)
        
        # Vérifier que l'opérateur est bien un superviseur
        operateur = request.user.profil_operateur
        if not operateur or operateur.type_operateur != 'SUPERVISEUR_PREPARATION':
            return JsonResponse({
                'success': False, 
                'error': 'Seuls les superviseurs de préparation peuvent finaliser la préparation'
            }, status=403)
        
        # Vérifier que la commande est dans l'état "Emballée"
        etat_actuel = commande.etat_actuel
        if not etat_actuel or etat_actuel.enum_etat.libelle != 'Emballée':
            return JsonResponse({
                'success': False, 
                'error': 'La commande doit être dans l\'état "Emballée" pour être finalisée'
            }, status=400)
        
        # Récupérer l'état "Préparée"
        try:
            etat_preparee = EnumEtatCmd.objects.get(libelle='Préparée')
        except EnumEtatCmd.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'error': 'L\'état "Préparée" n\'existe pas dans le système'
            }, status=500)
        
        # Fermer l'état actuel (Emballée)
        etat_actuel.date_fin = timezone.now()
        etat_actuel.save()
        
        # Créer le nouvel état (Préparée)
        nouvel_etat = EtatCommande.objects.create(
            commande=commande,
            enum_etat=etat_preparee,
            operateur=operateur,
            date_debut=timezone.now(),
            date_fin=None  # État actif
        )
        
        # Créer une opération pour tracer l'action
       
        
        print(f"✅ Préparation finalisée pour la commande {commande.id_yz} par {operateur.nom}")
        
        return JsonResponse({
            'success': True,
            'message': f'La préparation de la commande {commande.id_yz} a été finalisée avec succès',
            'nouvel_etat': 'Préparée',
            'operateur': f'{operateur.nom} {operateur.prenom}',
            'date_finalisation': timezone.now().strftime('%d/%m/%Y %H:%M')
        })
        
    except Commande.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'Commande non trouvée'
        }, status=404)
    except Exception as e:
        print(f"Erreur lors de la finalisation de la préparation: {str(e)}")
        return JsonResponse({
            'success': False, 
            'error': f'Erreur lors de la finalisation: {str(e)}'
        }, status=500)
    

@csrf_exempt
@login_required
def export_commandes_envoi_excel(request, envoi_id):
    """Exporter les commandes préparées d'un envoi en fichier Excel"""
    from django.http import HttpResponse
    from django.utils import timezone

    try:

        # Récupérer l'envoi
        envoi = Envoi.objects.select_related('region').get(id=envoi_id)

        # Importer openpyxl après avoir vérifié l'envoi
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        except ImportError as import_error:
            print(f"🚫 ERREUR - Import openpyxl échoué: {import_error}")
            return JsonResponse({
                'success': False,
                'error': f'Erreur d\'import openpyxl: {import_error}'
            }, status=500)
        
        # Récupérer les commandes associées à cet envoi. Si aucune n'est encore liée,
        # fallback: prendre les commandes "Préparée" de la région de l'envoi.
        commandes = (
            Commande.objects.filter(envoi=envoi)
            .select_related('client', 'ville', 'ville__region')
            .prefetch_related(
                'paniers__article',
                'paniers__variante',
                'paniers__variante__couleur',
                'paniers__variante__pointure',
            )
            .distinct()
        )

        if not commandes.exists():
            print("🔎 Aucun lien direct commande→envoi. Fallback sur TOUTES les commandes de la région.")
            commandes = (
                Commande.objects.filter(
                    ville__region=envoi.region,
                )
                .select_related('client', 'ville', 'ville__region')
                .prefetch_related(
                    'paniers__article',
                    'paniers__variante',
                    'paniers__variante__couleur',
                    'paniers__variante__pointure',
                )
                .distinct()
            )
        # Test: Créer un workbook simple d'abord
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Envoi_{envoi.numero_envoi}"[:31]  # Limiter à 31 caractères
        except Exception as wb_error:
            print(f"🚫 ERREUR - Création workbook échouée: {wb_error}")
            raise wb_error
        
        # Ajout structuré et stylé des données dans le fichier Excel
        try:
            # En-tête global
            title_cell = ws['A1']
            title_cell.value = f"EXPORT COMMANDES - ENVOI {envoi.numero_envoi}"
            title_cell.font = Font(size=14, bold=True)
            ws['A2'] = f"Région: {envoi.region.nom_region}"
            ws['A3'] = f"Date: {timezone.now().strftime('%d/%m/%Y %H:%M')}"

            # En-têtes des colonnes (structuré)
            # Demande: supprimer "ID YZ" et "N° Externe", renommer "Articles" en "Panier"
            headers = ['N° Commande', 'Client', 'Téléphone', 'Adresse', 'Ville', 'Région', 'Total (DH)', 'Panier']
            header_fill = PatternFill("solid", fgColor="1E3A8A")  # var(--preparation-border-accent) approchée
            header_font = Font(color="FFFFFF", bold=True)
            center = Alignment(horizontal='center', vertical='center')
            wrap = Alignment(wrap_text=True, vertical='top')

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
            

            # Largeurs de colonnes lisibles
            col_widths = {
                1: 16,  # N° Commande
                2: 28,  # Client
                3: 16,  # Téléphone
                4: 40,  # Adresse
                5: 18,  # Ville
                6: 18,  # Région
                7: 14,  # Total (DH)
                8: 80,  # Panier
            }
            for idx, width in col_widths.items():
                ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
            
            # Données simplifiées des commandes
            row = 6
            for commande in commandes:
                try:
                    # Données de base
                    client_nom = (
                        f"{(commande.client.prenom or '').strip()} {(commande.client.nom or '').strip()}".strip()
                        if commande.client else "N/A"
                    )
                    adresse_livraison = commande.adresse or (commande.client.adresse if getattr(commande, 'client', None) else '') or ""
                    ville_nom = commande.ville.nom if commande.ville else "N/A"
                    region_nom = commande.ville.region.nom_region if getattr(commande, 'ville', None) and commande.ville.region else (
                        envoi.region.nom_region if envoi and envoi.region else "N/A"
                    )

                    # Détails du panier (items)
                    items_texts = []
                    for p in commande.paniers.all():
                        try:
                            article_nom = p.article.nom if p.article else "Article N/A"
                            variante_text = ""
                            if p.variante:
                                couleur = getattr(p.variante.couleur, 'nom', None)
                                pointure = getattr(p.variante.pointure, 'pointure', None)
                                details = [v for v in [couleur, pointure] if v]
                                if details:
                                    variante_text = f" ({' - '.join(details)})"
                            items_texts.append(f"{p.quantite}x {article_nom}{variante_text}")
                        except Exception as p_err:
                            print(f"⚠️ ERREUR panier commande {commande.id}: {p_err}")
                            continue
                    # Articles sur la même ligne, séparateur lisible
                    panier_str = " | ".join(items_texts) if items_texts else ""

                    # Ecrire la ligne
                    thin = Side(border_style="thin", color="E5E7EB")
                    border = Border(left=thin, right=thin, top=thin, bottom=thin)

                    ws.cell(row=row, column=1, value=str(getattr(commande, 'id', '')))  # N° Commande (ID interne)
                    ws.cell(row=row, column=2, value=client_nom)
                    ws.cell(row=row, column=3, value=getattr(commande.client, 'numero_tel', '') if commande.client else '')
                    ws.cell(row=row, column=4, value=adresse_livraison)
                    ws.cell(row=row, column=5, value=ville_nom)
                    ws.cell(row=row, column=6, value=region_nom)
                    # Total commande (montant total)
                    total_cell = ws.cell(row=row, column=7, value=float(getattr(commande, 'total_cmd', 0) or 0))
                    total_cell.number_format = '#,##0.00'
                    panier_cell = ws.cell(row=row, column=8, value=panier_str)
                    # Pas de retour à la ligne: une seule ligne pour tout le panier
                    panier_cell.alignment = Alignment(wrap_text=False, vertical='top')

                    # Appliquer bordures et alignements par défaut
                    for col in range(1, 9):
                        c = ws.cell(row=row, column=col)
                        c.border = border
                        if col in (1, 3, 5, 6, 7):
                            c.alignment = center
                        else:
                            if col in (4, 8):
                                c.alignment = wrap
                            else:
                                c.alignment = Alignment(vertical='center')

                    row += 1

                except Exception as row_error:
                    print(f"🚫 ERREUR - Ligne commande {commande.id}: {row_error}")
                    continue
            
            # Auto-filter et freeze header
            ws.auto_filter.ref = f"A5:H{max(5, row-1)}"
            ws.freeze_panes = 'A6'
            
        except Exception as data_error:
            print(f"🚫 ERREUR - Ajout des données: {data_error}")
            raise data_error
        
        # Préparer la réponse HTTP
        try:
           
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="Envoi_{envoi.numero_envoi}_Commandes.xlsx"'
            # Sauvegarder le workbook dans la réponse
            wb.save(response)
            
            return response
            
        except Exception as save_error:
            print(f"🚫 ERREUR - Sauvegarde échouée: {save_error}")
            raise save_error
        
    except Envoi.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Envoi non trouvé'}, status=404)
    except Exception as e:
        import traceback
        print(f"Erreur dans export_commandes_envoi_excel: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la génération du fichier Excel: {str(e)}'
        }, status=500)


@superviseur_preparation_required
def api_commande_info(request, commande_id):
    """API pour récupérer les informations détaillées d'une commande pour la modale"""
    try:
        # Récupérer la commande
        commande = get_object_or_404(Commande, id=commande_id)
        
        # Récupérer les informations détaillées de manière sécurisée
        try:
            etat_actuel = commande.etat_actuel
        except:
            etat_actuel = None
        try:
            etat_precedent = commande.etat_precedent
        except:
            etat_precedent = None
        try:
            etat_confirmation = commande.etat_confirmation
        except:
            etat_confirmation = None
            
        try:
            operations = commande.operations.all().order_by('-date_operation')
        except:
            operations = []
        
        # Récupérer les informations détaillées
        context = {
            'commande': commande,
            'etat_actuel': etat_actuel,
            'etat_precedent': etat_precedent,
            'etat_confirmation': etat_confirmation,
            'operations': operations,
        }
        
        # Rendre le template HTML pour les informations détaillées
        html_content = render_to_string('Superpreparation/partials/_commande_info_modal.html', context)
        
        return JsonResponse({
            'success': True,
            'html_content': html_content
        })
        
    except Exception as e:
        import traceback
        print(f"Erreur dans api_commande_info: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors du chargement des informations: {str(e)}'
        }, status=500)

