from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Sum, F, Case, When, Value, CharField, Avg
from django.utils import timezone
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from decimal import Decimal
from datetime import datetime, timedelta
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import csv
import io

from parametre.models import Ville, Operateur, Region
from commande.models import Commande, EtatCommande, EnumEtatCmd, Panier, Envoi
from article.models import Article
from client.models import Client
from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur

# Fonction get_prix_upsell remplacée par get_prix_upsell_avec_compteur depuis commande_filters

def commande_a_impacte_stock(commande):
    """
    Vérifie si une commande a déjà impacté le stock (donc a été confirmée).
    Le stock est généralement décrémenté lors de la confirmation de la commande.
    """
    etats_impact_stock = ['Confirmée', 'Préparée', 'En cours de livraison', 'Livrée']
    
    # Vérifier si la commande a un de ces états dans son historique
    for etat in commande.etats.all():
        if etat.enum_etat.libelle in etats_impact_stock:
            return True
    
    return False

def creer_mouvement_stock(article, quantite, type_mouvement, commande, operateur, commentaire):
    """
    Crée un mouvement de stock pour tracer les changements
    """
    try:
        from article.models import MouvementStock
        
        # Calculer la nouvelle quantité disponible
        if type_mouvement in ['retour_client', 'ajustement_pos']:
            # Mouvements qui augmentent le stock
            nouvelle_qte = article.qte_disponible + abs(quantite)
            mouvement_quantite = abs(quantite)
        else:
            # Mouvements qui diminuent le stock
            nouvelle_qte = article.qte_disponible - abs(quantite)
            mouvement_quantite = -abs(quantite)
        
        # Mettre à jour le stock de l'article
        article.qte_disponible = max(0, nouvelle_qte)
        article.save(update_fields=['qte_disponible'])
        
        # Créer le mouvement de stock
        mouvement = MouvementStock.objects.create(
            article=article,
            type_mouvement=type_mouvement,
            quantite=mouvement_quantite,
            qte_apres_mouvement=article.qte_disponible,
            commentaire=commentaire,
            commande_associee=commande,
            operateur=operateur
        )
        
        print(f"📦 Mouvement de stock créé: {article.nom} - {type_mouvement} - {mouvement_quantite} (stock: {article.qte_disponible})")
        return mouvement
        
    except ImportError as e:
        print(f"❌ Erreur d'import MouvementStock: {str(e)}")
        return None
    except Exception as e:
        print(f"❌ Erreur lors de la création du mouvement de stock: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

# Create your views here.

@login_required
def dashboard(request):
    """Page d'accueil de l'interface opérateur logistique avec statistiques réelles"""
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil d'opérateur logistique non trouvé.")
        return redirect('login')
    
    # Dates pour les statistiques
    today = timezone.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    start_of_month = today.replace(day=1)
    
    # === STATISTIQUES GLOBALES (TOUS LES OPÉRATEURS) ===
    
    # 1. Commandes en préparation (En préparation actif)
    en_preparation = Commande.objects.filter(
        etats__enum_etat__libelle='En préparation',
        etats__date_fin__isnull=True
    ).distinct().count()
    
    # 2. Commandes prêtes à expédier (Préparée actif)
    prets_expedition = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée',
        etats__date_fin__isnull=True
    ).distinct().count()
    
    # 3. Commandes expédiées (En cours de livraison actif)
    expedies = Commande.objects.filter(
        etats__enum_etat__libelle='En cours de livraison',
        etats__date_fin__isnull=True
    ).distinct().count()
    
    # 4. Stock manquant (récupérer depuis la base de données)
    from article.models import Article
    
    # Compter les articles en rupture de stock (qte_disponible = 0) et actifs
    articles_rupture = Article.objects.filter(
        qte_disponible=0,
        actif=True
    ).count()
    
    # Compter les articles avec stock faible (moins de 5 unités)
    articles_stock_faible = Article.objects.filter(
        qte_disponible__lt=5,
        qte_disponible__gt=0,
        actif=True
    ).count()
    
    # Total des problèmes de stock
    stock_manquant = articles_rupture + articles_stock_faible
    
    # === ALERTES DE STOCK DÉTAILLÉES ===
    alertes_stock = []
    
    # Alertes de rupture de stock
    if articles_rupture > 0:
        alertes_stock.append({
            'type': 'danger',
            'message': f'{articles_rupture} article(s) en rupture de stock',
            'count': articles_rupture,
            'icon': 'fas fa-exclamation-triangle'
        })
    
    # Alertes de stock faible
    if articles_stock_faible > 0:
        alertes_stock.append({
            'type': 'warning', 
            'message': f'{articles_stock_faible} article(s) avec stock faible (< 5 unités)',
            'count': articles_stock_faible,
            'icon': 'fas fa-exclamation-circle'
        })
    
    # Articles à réapprovisionner bientôt (stock < 10)
    articles_a_reapprovisionner = Article.objects.filter(
        actif=True,
        qte_disponible__lt=10,
        qte_disponible__gt=0
    ).order_by('qte_disponible')[:3]
    
    if articles_a_reapprovisionner.exists():
        noms_articles = [art.nom for art in articles_a_reapprovisionner]
        alertes_stock.append({
            'type': 'info',
            'message': f'À réapprovisionner bientôt : {", ".join(noms_articles)}',
            'count': len(noms_articles),
            'icon': 'fas fa-info-circle'
        })
    
    # === DÉTAILS SUPPLÉMENTAIRES DU STOCK ===
    
    # Articles les plus en demande (avec le plus de commandes)
    articles_populaires = Article.objects.filter(
        actif=True,
        paniers__isnull=False
    ).annotate(
        nb_commandes=Count('paniers__commande', distinct=True)
    ).order_by('-nb_commandes')[:5]
    
    # Stock total et valeur
    stock_total = Article.objects.filter(actif=True).aggregate(
        total_articles=Count('id'),
        total_stock=Sum('qte_disponible'),
        valeur_totale=Sum(F('qte_disponible') * F('prix_unitaire'))
    )
    
    # Articles par catégorie avec problèmes de stock
    categories_stock = Article.objects.filter(
        actif=True
    ).values('categorie').annotate(
        total_articles=Count('id'),
        articles_rupture=Count('id', filter=Q(qte_disponible=0)),
        articles_faible=Count('id', filter=Q(qte_disponible__lt=5, qte_disponible__gt=0)),
        stock_moyen=Avg('qte_disponible')
    ).order_by('-articles_rupture', '-articles_faible')
    
    # Si aucune alerte, message positif
    if not alertes_stock:
        alertes_stock.append({
            'type': 'success',
            'message': 'Tous les articles sont bien approvisionnés',
            'count': 0,
            'icon': 'fas fa-check-circle'
        })
    
    # === STATISTIQUES SPÉCIFIQUES À CET OPÉRATEUR ===
    
    # Commandes affectées à cet opérateur
    mes_commandes = Commande.objects.filter(
        etats__operateur=operateur,
        etats__enum_etat__libelle='En cours de livraison',
        etats__date_fin__isnull=True
    ).distinct()
    
    # Mes commandes aujourd'hui
    mes_commandes_aujourd_hui = mes_commandes.filter(
        etats__date_debut__date=today
    ).count()
    
    # Mes commandes cette semaine
    mes_commandes_semaine = mes_commandes.filter(
        etats__date_debut__date__gte=start_of_week
    ).count()
    
    # Mes commandes ce mois
    mes_commandes_mois = mes_commandes.filter(
        etats__date_debut__date__gte=start_of_month
    ).count()
    
    # Montant total de mes commandes
    mes_commandes_montant = mes_commandes.aggregate(total=Sum('total_cmd'))['total'] or 0
    
    # === COMMANDES URGENTES (exemple: commandes de plus de 3 jours) ===
    date_limite_urgence = today - timedelta(days=3)
    commandes_urgentes = mes_commandes.filter(
        etats__date_debut__date__lte=date_limite_urgence
    )
    
    # === STATISTIQUES PAR VILLE POUR CET OPÉRATEUR ===
    stats_par_ville = mes_commandes.values(
        'ville__nom', 'ville__region__nom_region'
    ).annotate(
        nb_commandes=Count('id'),
        total_montant=Sum('total_cmd')
    ).order_by('-nb_commandes')[:5]  # Top 5 des villes
    
    # === ÉVOLUTION DES LIVRAISONS (7 derniers jours) ===
    livraisons_semaine = []
    for i in range(7):
        date_jour = today - timedelta(days=6-i)
        nb_livraisons = Commande.objects.filter(
            etats__operateur=operateur,
            etats__enum_etat__libelle='Livrée',
            etats__date_debut__date=date_jour
        ).distinct().count()
        livraisons_semaine.append({
            'date': date_jour,
            'nb_livraisons': nb_livraisons
        })
    
    # === CALCULS DE POURCENTAGES ===
    # Pourcentage d'évolution cette semaine vs semaine dernière
    semaine_derniere_debut = start_of_week - timedelta(days=7)
    semaine_derniere_fin = start_of_week - timedelta(days=1)
    
    commandes_semaine_derniere = Commande.objects.filter(
        etats__operateur=operateur,
        etats__enum_etat__libelle='En cours de livraison',
        etats__date_debut__date__range=[semaine_derniere_debut, semaine_derniere_fin]
    ).distinct().count()
    
    if commandes_semaine_derniere > 0:
        evolution_semaine = ((mes_commandes_semaine - commandes_semaine_derniere) / commandes_semaine_derniere) * 100
    else:
        evolution_semaine = 100 if mes_commandes_semaine > 0 else 0
    
    context = {
        'operateur': operateur,
        # Statistiques globales
        'en_preparation': en_preparation,
        'prets_expedition': prets_expedition,
        'expedies': expedies,
        'stock_manquant': stock_manquant,
        # Mes statistiques
        'mes_commandes_total': mes_commandes.count(),
        'mes_commandes_aujourd_hui': mes_commandes_aujourd_hui,
        'mes_commandes_semaine': mes_commandes_semaine,
        'mes_commandes_mois': mes_commandes_mois,
        'mes_commandes_montant': mes_commandes_montant,
        # Commandes urgentes
        'commandes_urgentes': commandes_urgentes,
        'nb_commandes_urgentes': commandes_urgentes.count(),
        # Statistiques par ville
        'stats_par_ville': stats_par_ville,
        # Alertes de stock détaillées
        'alertes_stock': alertes_stock,
        'articles_rupture': articles_rupture,
        'articles_stock_faible': articles_stock_faible,
        # Détails du stock
        'articles_populaires': articles_populaires,
        'stock_total': stock_total,
        'categories_stock': categories_stock,
        'articles_a_reapprovisionner': articles_a_reapprovisionner,
        # Évolution
        'evolution_semaine': round(evolution_semaine, 1),
        'livraisons_semaine': livraisons_semaine,
        # Méta
        'page_title': 'Tableau de Bord',
        'page_subtitle': f'Interface Opérateur Logistique - {operateur.nom_complet}',
    }
    
    return render(request, 'composant_generale/operatLogistic/home.html', context)

@login_required
def liste_commandes(request):
    """Liste des commandes affectées à cet opérateur logistique"""
    try:
        # Récupérer l'opérateur logistique connecté
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil d'opérateur logistique non trouvé.")
        return redirect('login')
    
    # Récupérer les commandes affectées à cet opérateur logistique
    # Les commandes sont dans l'état "En cours de livraison" et affectées à cet opérateur
    commandes_affectees = Commande.objects.filter(
        etats__enum_etat__libelle='En cours de livraison',
        etats__operateur=operateur,
        etats__date_fin__isnull=True  # État actif
    ).select_related('client', 'ville', 'ville__region').prefetch_related('etats').distinct().order_by('-etats__date_debut')
    
    # Si aucune commande trouvée avec la méthode stricte, essayer une approche plus large
    if not commandes_affectees.exists():
        # Chercher toutes les commandes qui ont été affectées à cet opérateur pour la livraison
        # et qui n'ont pas encore d'état "Livrée" ou "Retournée"
        commandes_affectees = Commande.objects.filter(
            etats__operateur=operateur,
            etats__enum_etat__libelle__in=['En cours de livraison', 'Préparée']
        ).exclude(
            # Exclure les commandes qui ont déjà un état ultérieur actif
            Q(etats__enum_etat__libelle__in=['Livrée', 'Retournée', 'Annulée'], etats__date_fin__isnull=True)
        ).select_related('client', 'ville', 'ville__region').prefetch_related('etats').distinct().order_by('-etats__date_debut')
    
    # Recherche
    search_query = request.GET.get('search', '')
    if search_query:
        commandes_affectees = commandes_affectees.filter(
            Q(id_yz__icontains=search_query) |
            Q(num_cmd__icontains=search_query) |
            Q(client__nom__icontains=search_query) |
            Q(client__prenom__icontains=search_query) |
            Q(client__numero_tel__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(commandes_affectees, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistiques
    today = timezone.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    start_of_month = today.replace(day=1)
    
    # Compter les commandes par période (utiliser la même logique flexible)
    base_filter = Q(etats__operateur=operateur) & (
        Q(etats__enum_etat__libelle='En cours de livraison', etats__date_fin__isnull=True) |
        Q(etats__enum_etat__libelle='Préparée')
    )
    
    affectees_aujourd_hui = Commande.objects.filter(
        base_filter & Q(etats__date_debut__date=today)
    ).distinct().count()
    
    affectees_semaine = Commande.objects.filter(
        base_filter & Q(etats__date_debut__date__gte=start_of_week)
    ).distinct().count()
    
    affectees_mois = Commande.objects.filter(
        base_filter & Q(etats__date_debut__date__gte=start_of_month)
    ).distinct().count()
    
    total_commandes = commandes_affectees.count()
    
    # Montant total
    total_montant = commandes_affectees.aggregate(total=Sum('total_cmd'))['total'] or 0
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'operateur': operateur,
        'total_commandes': total_commandes,
        'total_montant': total_montant,
        'affectees_aujourd_hui': affectees_aujourd_hui,
        'affectees_semaine': affectees_semaine,
        'affectees_mois': affectees_mois,
        'page_title': 'Mes Commandes',
        'page_subtitle': f'Commandes affectées à {operateur.nom_complet}',
    }
    
    return render(request, 'operatLogistic/liste_commande.html', context)

@login_required
def detail_commande(request, commande_id):
    """Affiche les détails d'une commande spécifique pour l'opérateur logistique."""
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil d'opérateur logistique non trouvé.")
        return redirect('login')

    try:
        commande = get_object_or_404(Commande, id=commande_id)

        # Vérifier que la commande est bien affectée à cet opérateur
        is_affectee = commande.etats.filter(
            Q(enum_etat__libelle='En cours de livraison') | Q(enum_etat__libelle='Préparée'),
            operateur=operateur
        ).exists()

        if not is_affectee:
            messages.error(request, "Vous n'avez pas l'autorisation de voir les détails de cette commande.")
            return redirect('operatLogistic:liste_commandes')
            
    except Commande.DoesNotExist:
        raise Http404("La commande n'existe pas.")

    # Déterminer si la commande a déjà été confirmée (impact sur le stock)
    commande_confirmee = commande_a_impacte_stock(commande)

    context = {
        'commande': commande,
        'commande_confirmee': commande_confirmee,
    }
    
    return render(request, 'operatLogistic/detail_commande.html', context)

@login_required
def repartition(request):
    """Page de répartition automatique des commandes par ville et région"""
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil d'opérateur logistique non trouvé.")
        return redirect('login')
    
    # Récupérer les commandes actuellement en cours de livraison (déjà réparties)
    from commande.models import EnumEtatCmd
    from django.db.models import Count
    
    commandes_reparties = Commande.objects.filter(
        etats__enum_etat__libelle='En cours de livraison',
        etats__date_fin__isnull=True  # État actif
    ).select_related('client', 'ville', 'ville__region').prefetch_related('etats__operateur').distinct()
    
    # Statistiques par ville et région des commandes réparties
    stats_par_ville = commandes_reparties.values(
        'ville__nom', 'ville__region__nom_region'
    ).annotate(
        nb_commandes=Count('id'),
        total_montant=Sum('total_cmd')
    ).order_by('ville__region__nom_region', 'ville__nom')
    
    stats_par_region = commandes_reparties.values(
        'ville__region__nom_region'
    ).annotate(
        nb_commandes=Count('id'),
        total_montant=Sum('total_cmd')
    ).order_by('ville__region__nom_region')
    
    # Calculer le montant total général
    total_montant_general = commandes_reparties.aggregate(total=Sum('total_cmd'))['total'] or 0
    
    context = {
        'operateur': operateur,
        'commandes_reparties': commandes_reparties,
        'stats_par_ville': stats_par_ville,
        'stats_par_region': stats_par_region,
        'total_commandes_reparties': commandes_reparties.count(),
        'total_montant_general': total_montant_general,
        'page_title': 'Répartition des Commandes',
        'page_subtitle': 'Répartition automatique par ville et région',
    }
    
    return render(request, 'operatLogistic/repartition.html', context)

@login_required
def details_region(request, nom_region):
    """Page de détails d'une région avec toutes les commandes par ville et opérateur"""
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil d'opérateur logistique non trouvé.")
        return redirect('login')
    
    # Statistiques détaillées par ville
    commandes_region = Commande.objects.filter(
        etats__enum_etat__libelle='En cours de livraison',
        etats__date_fin__isnull=True,  # État actif
        ville__region__nom_region=nom_region
    ).select_related('client', 'ville', 'ville__region').prefetch_related('etats__operateur').distinct()
    stats_par_ville = commandes_region.values(
        'ville__nom'
    ).annotate(
        nb_commandes=Count('id'),
        total_montant=Sum('total_cmd')
    ).order_by('ville__nom')
    
    # Calculer les totaux
    total_commandes = commandes_region.count()
    total_montant = commandes_region.aggregate(total=Sum('total_cmd'))['total'] or 0
    
    context = {
        'operateur': operateur,
        'nom_region': nom_region,
        'commandes_region': commandes_region,
        'stats_par_ville': stats_par_ville,
        'total_commandes': total_commandes,
        'total_montant': total_montant,
        'page_title': f'Détails - {nom_region}',
        'page_subtitle': 'Répartition détaillée par ville',
    }
    
    return render(request, 'operatLogistic/details_region.html', context)

@login_required
def parametre(request):
    """Page des paramètres opérateur logistique"""
    return render(request, 'operatLogistic/parametre.html')

@login_required
def creer_operateur_logistique(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        nom = request.POST.get('nom')
        prenom = request.POST.get('prenom')
        mail = request.POST.get('mail')
        telephone = request.POST.get('telephone')
        adresse = request.POST.get('adresse')

        # Validation de base
        if not all([username, password, nom, prenom, mail]):
            messages.error(request, "Tous les champs obligatoires doivent être remplis.")
            return render(request, 'composant_generale/creer_operateur.html', {'form_data': request.POST})
        
        if User.objects.filter(username=username).exists():
            messages.error(request, "Ce nom d'utilisateur existe déjà.")
            return render(request, 'composant_generale/creer_operateur.html', {'form_data': request.POST})

        if User.objects.filter(email=mail).exists():
            messages.error(request, "Cet email est déjà utilisé.")
            return render(request, 'composant_generale/creer_operateur.html', {'form_data': request.POST})

        try:
            # Créer l'utilisateur Django
            user = User.objects.create_user(
                username=username,
                email=mail,
                password=password,
                first_name=prenom,
                last_name=nom
            )
            user.save()

            # Créer le profil Operateur
            operateur = Operateur.objects.create(
                user=user,
                nom=nom,
                prenom=prenom,
                mail=mail,
                type_operateur='LOGISTIQUE', # Type spécifique pour logistique
                telephone=telephone,
                adresse=adresse
            )
            operateur.save()

            # Ajouter l'utilisateur au groupe 'operateur_logistique'
            group, created = Group.objects.get_or_create(name='operateur_logistique')
            user.groups.add(group)

            messages.success(request, f"L'opérateur logistique {prenom} {nom} a été créé avec succès.")
            return redirect('app_admin:liste_operateurs') # Rediriger vers la liste des opérateurs

        except Exception as e:
            messages.error(request, f"Une erreur est survenue lors de la création de l'opérateur : {e}")
            if 'user' in locals() and user.pk:
                user.delete()
            return render(request, 'composant_generale/creer_operateur.html', {'form_data': request.POST})

    return render(request, 'composant_generale/creer_operateur.html')

@login_required
def profile_logistique(request):
    """Page de profil pour l'opérateur logistique"""
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil d'opérateur non trouvé.")
        return redirect('login') # Rediriger vers la page de connexion ou une page d'erreur
    return render(request, 'operatLogistic/profile.html', {'operateur': operateur})

@login_required
def modifier_profile_logistique(request):
    """Page de modification de profil pour l'opérateur logistique"""
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil d'opérateur non trouvé.")
        return redirect('login')

    user = request.user

    if request.method == 'POST':
        # Récupérer les données du formulaire
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        telephone = request.POST.get('telephone')
        adresse = request.POST.get('adresse')

        # Validation de base
        if not all([first_name, last_name, email]):
            messages.error(request, "Le prénom, le nom et l'email sont obligatoires.")
            return render(request, 'operatLogistic/modifier_profile.html', {'operateur': operateur, 'user': user})

        # Vérifier si l'email est déjà utilisé par un autre utilisateur (sauf l'utilisateur actuel)
        if User.objects.filter(email=email).exclude(pk=user.pk).exists():
            messages.error(request, "Cet email est déjà utilisé par un autre compte.")
            return render(request, 'operatLogistic/modifier_profile.html', {'operateur': operateur, 'user': user})
        
        try:
            # Mettre à jour l'objet User
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.save()

            # Mettre à jour l'objet Operateur
            operateur.nom = last_name # Mettre à jour le nom de famille de l'opérateur
            operateur.prenom = first_name # Mettre à jour le prénom de l'opérateur
            operateur.mail = email
            operateur.telephone = telephone
            operateur.adresse = adresse
            # Ne pas modifier type_operateur ou actif
            operateur.save()

            messages.success(request, "Votre profil a été mis à jour avec succès.")
            return redirect('operatLogistic:profile')

        except Exception as e:
            messages.error(request, f"Une erreur est survenue lors de la mise à jour : {e}")

    return render(request, 'operatLogistic/modifier_profile.html', {'operateur': operateur, 'user': user})

@login_required
def changer_mot_de_passe_logistique(request):
    """Page de changement de mot de passe pour l'opérateur logistique - Désactivée"""
    return redirect('operatLogistic:profile')

@login_required
def marquer_livree(request, commande_id):
    """Marquer une commande comme livrée"""
    if request.method == 'POST':
        try:
            # Vérifier que l'utilisateur est un opérateur logistique
            operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
            
            # Récupérer la commande
            commande = Commande.objects.get(id=commande_id)
            
            # Vérifier que la commande est affectée à cet opérateur
            etat_actuel = commande.etat_actuel
            if not (etat_actuel and etat_actuel.operateur == operateur and etat_actuel.enum_etat.libelle == 'En cours de livraison'):
                return JsonResponse({'success': False, 'message': 'Cette commande ne vous est pas affectée ou n\'est pas en cours de livraison.'})
            
            # Marquer comme livrée
            gerer_changement_etat_automatique(
                commande, 
                'Livrée', 
                operateur=operateur,
                commentaire="Commande livrée par l'opérateur logistique"
            )
            
            return JsonResponse({'success': True, 'message': f'Commande {commande.id_yz} marquée comme livrée avec succès.'})
            
        except Operateur.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Opérateur non trouvé.'})
        except Commande.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Commande non trouvée.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée.'})

@login_required
def signaler_probleme(request, commande_id):
    """Signaler un problème avec une commande"""
    if request.method == 'POST':
        try:
            # Vérifier que l'utilisateur est un opérateur logistique
            operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
            
            # Récupérer les données
            data = json.loads(request.body)
            description = data.get('description', '')
            
            if not description:
                return JsonResponse({'success': False, 'message': 'Description du problème requise.'})
            
            # Récupérer la commande
            commande = Commande.objects.get(id=commande_id)
            
            # Vérifier que la commande est affectée à cet opérateur
            etat_actuel = commande.etat_actuel
            if not (etat_actuel and etat_actuel.operateur == operateur):
                return JsonResponse({'success': False, 'message': 'Cette commande ne vous est pas affectée.'})
            
            # Créer une opération pour signaler le problème
            Operation.objects.create(
                commande=commande,
                type_operation='COMMENTAIRE',
                operateur=operateur,
                commentaire=f"PROBLÈME SIGNALÉ: {description}"
            )
            
            return JsonResponse({'success': True, 'message': f'Problème signalé pour la commande {commande.id_yz}.'})
            
        except Operateur.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Opérateur non trouvé.'})
        except Commande.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Commande non trouvée.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée.'})

# === FONCTIONS D'EXPORTATION ===

def _get_commandes_preparees():
    """Fonction utilitaire pour récupérer les commandes préparées (en cours de livraison)"""
    return Commande.objects.filter(
        etats__enum_etat__libelle='En cours de livraison',
        etats__date_fin__isnull=True  # État actif
    ).select_related('client', 'ville', 'ville__region').prefetch_related('etats__operateur').distinct()

def _prepare_commande_data(commandes):
    """Fonction utilitaire pour préparer les données de commande pour l'exportation"""
    data = []
    for commande in commandes:
        # Récupérer l'opérateur affecté pour la livraison
        etat_livraison = commande.etats.filter(
            enum_etat__libelle='En cours de livraison',
            date_fin__isnull=True
        ).first()
        
        operateur_affecte = etat_livraison.operateur if etat_livraison else None
        
        # Vérifier si les champs existent avant de les utiliser
        ville_nom = commande.ville.nom if commande.ville else commande.ville_init or 'Non définie'
        region_nom = commande.ville.region.nom_region if commande.ville and commande.ville.region else 'Non définie'
        client_email = getattr(commande.client, 'email', '') or ''
        
        # Construire le panier (articles commandés)
        panier_details = []
        paniers = commande.paniers.all().select_related('article')
        
        for panier in paniers:
            if panier.article:
                # Format: "Nom Article [Référence] (Qté: X) - Prix unitaire DH"
                article_info = f"{panier.article.nom or 'Article sans nom'}"
                
                if panier.article.reference:
                    article_info += f" [{panier.article.reference}]"
                
                # Ajouter couleur et pointure si disponibles
                details_supplementaires = []
                if hasattr(panier.article, 'couleur') and panier.article.couleur:
                    details_supplementaires.append(f"Couleur: {panier.article.couleur}")
                if hasattr(panier.article, 'pointure') and panier.article.pointure:
                    details_supplementaires.append(f"Pointure: {panier.article.pointure}")
                
                if details_supplementaires:
                    article_info += f" ({', '.join(details_supplementaires)})"
                
                article_info += f" - Qté: {panier.quantite}"
                article_info += f" - {panier.article.prix_unitaire:.2f} DH/unité"
                article_info += f" - Sous-total: {panier.sous_total:.2f} DH"
                
                panier_details.append(article_info)
        
        # Joindre tous les articles avec un séparateur
        panier_complet = " | ".join(panier_details) if panier_details else "Aucun article"
        
        data.append({
            'ID YZ': commande.id_yz or '',
            'Numéro Commande': commande.num_cmd or '',
            'Client': f"{commande.client.prenom} {commande.client.nom}",
            'Téléphone': commande.client.numero_tel or '',
            'Email': client_email,
            'Ville': ville_nom,
            'Région': region_nom,
            'Adresse': commande.adresse or '',
            'Panier (Articles)': panier_complet,
            'Nombre d\'Articles': paniers.count(),
            'Montant Total': f"{commande.total_cmd:.2f} DH",
            'Date Commande': commande.date_cmd.strftime('%d/%m/%Y') if commande.date_cmd else '',
            'Date Création': commande.date_creation.strftime('%d/%m/%Y %H:%M') if commande.date_creation else '',
            'Opérateur Affecté': operateur_affecte.nom_complet if operateur_affecte else 'Non affecté',
            'Date Affectation': etat_livraison.date_debut.strftime('%d/%m/%Y %H:%M') if etat_livraison and etat_livraison.date_debut else '',
            'Statut': 'En cours de livraison'
        })
    return data

@login_required
def export_all_regions_excel(request):
    """Exporter toutes les commandes de toutes les régions en Excel"""
    try:
        # Vérifier que l'utilisateur est un opérateur logistique
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        messages.error(request, "Accès non autorisé.")
        return redirect('operatLogistic:repartition')
    
    if not EXCEL_AVAILABLE:
        messages.error(request, "L'exportation Excel n'est pas disponible. Veuillez installer openpyxl.")
        return redirect('operatLogistic:repartition')
    
    # Récupérer toutes les commandes préparées
    commandes = _get_commandes_preparees()
    
    if not commandes.exists():
        messages.warning(request, "Aucune commande disponible pour l'exportation.")
        return redirect('operatLogistic:repartition')
    
    # Préparer les données
    data = _prepare_commande_data(commandes)
    
    # Créer le fichier Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Toutes les Régions"
    
    # En-têtes
    headers = list(data[0].keys()) if data else []
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajouter les en-têtes
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Ajouter les données
    for row, commande_data in enumerate(data, 2):
        for col, value in enumerate(commande_data.values(), 1):
            worksheet.cell(row=row, column=col, value=value)
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 20
    
    # Préparer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"commandes_toutes_regions_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Sauvegarder le fichier
    workbook.save(response)
    
    return response

@login_required
def export_all_regions_csv(request):
    """Exporter toutes les commandes de toutes les régions en CSV"""
    try:
        # Vérifier que l'utilisateur est un opérateur logistique
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        messages.error(request, "Accès non autorisé.")
        return redirect('operatLogistic:repartition')
    
    # Récupérer toutes les commandes préparées
    commandes = _get_commandes_preparees()
    
    if not commandes.exists():
        messages.warning(request, "Aucune commande disponible pour l'exportation.")
        return redirect('operatLogistic:repartition')
    
    # Préparer les données
    data = _prepare_commande_data(commandes)
    
    # Créer la réponse CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"commandes_toutes_regions_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Ajouter BOM pour Excel
    response.write('\ufeff')
    
    # Créer le writer CSV
    writer = csv.writer(response, delimiter=';')  # Utiliser point-virgule pour Excel
    
    # Écrire les en-têtes
    if data:
        writer.writerow(data[0].keys())
        
        # Écrire les données
        for commande_data in data:
            writer.writerow(commande_data.values())
    
    return response

@login_required
def export_region_excel(request, nom_region):
    """Exporter les commandes d'une région spécifique en Excel"""
    try:
        # Vérifier que l'utilisateur est un opérateur logistique
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        messages.error(request, "Accès non autorisé.")
        return redirect('operatLogistic:repartition')
    
    if not EXCEL_AVAILABLE:
        messages.error(request, "L'exportation Excel n'est pas disponible. Veuillez installer openpyxl.")
        return redirect('operatLogistic:repartition')
    
    from urllib.parse import unquote
    nom_region = unquote(nom_region)
    
    # Récupérer les commandes de cette région
    commandes = _get_commandes_preparees().filter(ville__region__nom_region=nom_region)
    
    if not commandes.exists():
        messages.warning(request, f"Aucune commande disponible pour la région {nom_region}.")
        return redirect('operatLogistic:repartition')
    
    # Préparer les données
    data = _prepare_commande_data(commandes)
    
    # Créer le fichier Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = f"Région {nom_region}"
    
    # En-têtes
    headers = list(data[0].keys()) if data else []
    
    # Style pour les en-têtes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajouter les en-têtes
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Ajouter les données
    for row, commande_data in enumerate(data, 2):
        for col, value in enumerate(commande_data.values(), 1):
            worksheet.cell(row=row, column=col, value=value)
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 20
    
    # Préparer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"commandes_{nom_region.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Sauvegarder le fichier
    workbook.save(response)
    
    return response

@login_required
def export_region_csv(request, nom_region):
    """Exporter les commandes d'une région spécifique en CSV"""
    try:
        # Vérifier que l'utilisateur est un opérateur logistique
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        messages.error(request, "Accès non autorisé.")
        return redirect('operatLogistic:repartition')
    
    from urllib.parse import unquote
    nom_region = unquote(nom_region)
    
    # Récupérer les commandes de cette région
    commandes = _get_commandes_preparees().filter(ville__region__nom_region=nom_region)
    
    if not commandes.exists():
        messages.warning(request, f"Aucune commande disponible pour la région {nom_region}.")
        return redirect('operatLogistic:repartition')
    
    # Préparer les données
    data = _prepare_commande_data(commandes)
    
    # Créer la réponse CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"commandes_{nom_region.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Ajouter BOM pour Excel
    response.write('\ufeff')
    
    # Créer le writer CSV
    writer = csv.writer(response, delimiter=';')  # Utiliser point-virgule pour Excel
    
    # Écrire les en-têtes
    if data:
        writer.writerow(data[0].keys())
        
        # Écrire les données
        for commande_data in data:
            writer.writerow(commande_data.values())
    
    return response

# === FONCTIONS UTILITAIRES POUR L'EXPORTATION ===

def _format_commande_data(commande):
    """Fonction utilitaire pour formater les données d'une commande pour l'export"""
    return {
        'id_yz': commande.id_yz,
        'num_cmd': commande.num_cmd or '',
        'date_cmd': commande.date_cmd.strftime('%d/%m/%Y') if commande.date_cmd else '',
        'date_creation': commande.date_creation.strftime('%d/%m/%Y %H:%M') if commande.date_creation else '',
        'client_nom': f"{commande.client.prenom} {commande.client.nom}",
        'client_tel': commande.client.numero_tel or '',
        'client_email': commande.client.email or '',
        'ville': commande.ville.nom if commande.ville else '',
        'region': commande.ville.region.nom_region if commande.ville and commande.ville.region else '',
        'frais_livraison': float(commande.ville.frais_livraison) if commande.ville and commande.ville.frais_livraison else 0,
        'adresse': commande.adresse or '',
        'total_cmd': float(commande.total_cmd) if commande.total_cmd else 0,
        'etat_actuel': commande.etat_actuel.enum_etat.libelle if commande.etat_actuel else '',
        'operateur': commande.etat_actuel.operateur.nom_complet if commande.etat_actuel and commande.etat_actuel.operateur else '',
        'date_affectation': commande.etat_actuel.date_debut.strftime('%d/%m/%Y %H:%M') if commande.etat_actuel and commande.etat_actuel.date_debut else '',
        'nb_articles': commande.paniers.count(),
        'articles': [
            {
                'nom': panier.article.nom,
                'reference': panier.article.reference or '',
                'quantite': panier.quantite,
                'prix_unitaire': float(panier.prix_unitaire) if panier.prix_unitaire else 0,
                'sous_total': float(panier.total_ligne) if panier.total_ligne else 0,
            }
            for panier in commande.paniers.all()
        ]
    }

def _create_excel_workbook(commandes_data, title):
    """Fonction utilitaire pour créer un classeur Excel"""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl n'est pas installé")
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = title
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # En-têtes
    headers = [
        'ID YZ', 'N° Externe', 'Date Commande', 'Date Création',
        'Client', 'Téléphone', 'Email', 'Ville', 'Région',
        'Frais Livraison', 'Adresse', 'Total Commande', 'État',
        'Opérateur', 'Date Affectation', 'Nb Articles'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    for row, commande in enumerate(commandes_data, 2):
        worksheet.cell(row=row, column=1, value=commande['id_yz']).border = border
        worksheet.cell(row=row, column=2, value=commande['num_cmd']).border = border
        worksheet.cell(row=row, column=3, value=commande['date_cmd']).border = border
        worksheet.cell(row=row, column=4, value=commande['date_creation']).border = border
        worksheet.cell(row=row, column=5, value=commande['client_nom']).border = border
        worksheet.cell(row=row, column=6, value=commande['client_tel']).border = border
        worksheet.cell(row=row, column=7, value=commande['client_email']).border = border
        worksheet.cell(row=row, column=8, value=commande['ville']).border = border
        worksheet.cell(row=row, column=9, value=commande['region']).border = border
        worksheet.cell(row=row, column=10, value=commande['frais_livraison']).border = border
        worksheet.cell(row=row, column=11, value=commande['adresse']).border = border
        worksheet.cell(row=row, column=12, value=commande['total_cmd']).border = border
        worksheet.cell(row=row, column=13, value=commande['etat_actuel']).border = border
        worksheet.cell(row=row, column=14, value=commande['operateur']).border = border
        worksheet.cell(row=row, column=15, value=commande['date_affectation']).border = border
        worksheet.cell(row=row, column=16, value=commande['nb_articles']).border = border
    
    # Ajuster la largeur des colonnes
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return workbook

# === NOUVELLES VUES POUR LA GESTION DES ARTICLES ===

@login_required
def api_articles_disponibles(request):
    """API pour récupérer la liste des articles disponibles"""
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Profil d\'opérateur logistique non trouvé.'}, status=403)
    
    try:
        from django.utils import timezone
        
        # Récupérer le compteur upsell de la commande si fourni
        commande_id = request.GET.get('commande_id')
        compteur_upsell = 0
        
        if commande_id:
            try:
                commande = Commande.objects.get(id=commande_id)
                compteur_upsell = commande.compteur
            except Commande.DoesNotExist:
                pass  # Garder compteur à 0
        
        # Récupérer tous les articles actifs avec leurs promotions
        articles = Article.objects.filter(actif=True).prefetch_related('promotions').order_by('nom')
        
        articles_data = []
        for article in articles:
            try:
                # Utiliser des valeurs par défaut sécurisées
                has_promo_active = getattr(article, 'has_promo_active', False)
                prix_actuel = float(article.prix_actuel) if article.prix_actuel is not None else float(article.prix_unitaire)
                prix_unitaire = float(article.prix_unitaire) if article.prix_unitaire is not None else 0.0
                prix_upsell_1 = float(article.prix_upsell_1) if article.prix_upsell_1 is not None else 0.0
                prix_upsell_2 = float(article.prix_upsell_2) if article.prix_upsell_2 is not None else 0.0
                prix_upsell_3 = float(article.prix_upsell_3) if article.prix_upsell_3 is not None else 0.0
                prix_upsell_4 = float(article.prix_upsell_4) if article.prix_upsell_4 is not None else 0.0
                
                # Déterminer la phase de l'article
                phase = getattr(article, 'phase', 'EN_COURS')
                
                # Déterminer le prix selon le compteur upsell
                prix_avec_compteur = get_prix_upsell_avec_compteur(article, compteur_upsell)
                
                article_data = {
                    'id': article.id,
                    'nom': article.nom or '',
                    'reference': article.reference or '',
                    'prix_unitaire': prix_unitaire,
                    'prix_actuel': prix_avec_compteur,
                    'prix_upsell_1': prix_upsell_1,
                    'prix_upsell_2': prix_upsell_2,
                    'prix_upsell_3': prix_upsell_3,
                    'prix_upsell_4': prix_upsell_4,
                    'pointure': article.pointure or '',
                    'couleur': article.couleur or '',
                    'categorie': article.categorie or '',
                    'qte_disponible': article.qte_disponible or 0,
                    'has_promo_active': has_promo_active,
                    'phase': phase,
                    'isUpsell': getattr(article, 'isUpsell', False)
                }
                
                articles_data.append(article_data)
                
            except Exception as article_error:
                print(f"Erreur avec l'article {article.id}: {str(article_error)}")
                continue  # Ignorer cet article et passer au suivant
        
        return JsonResponse({
            'success': True,
            'articles': articles_data,
            'compteur_upsell': compteur_upsell
        })
        
    except Exception as e:
        print(f"Erreur dans api_articles_disponibles: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Erreur interne: {str(e)}'}, status=500)

@login_required
def ajouter_article_commande(request, commande_id):
    """Ajouter un article à la commande"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Profil d\'opérateur logistique non trouvé.'}, status=403)
    
    try:
        from django.db import transaction
        
        with transaction.atomic():
            commande = Commande.objects.select_for_update().get(id=commande_id)
            
            # Vérifier que la commande est affectée à cet opérateur
            etat_livraison = commande.etats.filter(
                operateur=operateur,
                enum_etat__libelle='En cours de livraison',
                date_fin__isnull=True
            ).first()
            
            if not etat_livraison:
                return JsonResponse({'error': 'Cette commande n\'est pas affectée à votre livraison.'}, status=403)
            
            # Récupérer les paramètres
            article_id = request.POST.get('article_id')
            quantite = int(request.POST.get('quantite', 1))
            
            if not article_id:
                return JsonResponse({'error': 'Article manquant'}, status=400)
            
            if quantite <= 0:
                return JsonResponse({'error': 'Quantité invalide'}, status=400)
            
            # Récupérer l'article
            try:
                article = Article.objects.get(id=article_id)
            except Article.DoesNotExist:
                return JsonResponse({'error': 'Article non trouvé'}, status=404)
            
            # Vérifier si la commande a déjà impacté le stock
            stock_deja_impacte = commande_a_impacte_stock(commande)
            
            if stock_deja_impacte:
                # Vérifier le stock disponible avant d'ajouter l'article
                if article.qte_disponible < quantite:
                    return JsonResponse({
                        'error': f'Stock insuffisant pour {article.nom}. Stock disponible: {article.qte_disponible}, demandé: {quantite}'
                    }, status=400)
                
                # Décrémenter le stock et créer un mouvement
                creer_mouvement_stock(
                    article=article,
                    quantite=quantite,
                    type_mouvement='sortie',
                    commande=commande,
                    operateur=operateur,
                    commentaire=f'Ajout d\'article lors de la modification de la commande {commande.id_yz} par {operateur.nom_complet}'
                )
                
                print(f"📦 Stock décrémenté: {article.nom} - {quantite} unités (commande confirmée)")
            
            # Vérifier si l'article existe déjà dans le panier
            panier_existant = Panier.objects.filter(commande=commande, article=article).first()
            
            if panier_existant:
                # Si l'article existe déjà, augmenter la quantité
                if stock_deja_impacte:
                    # Le stock a déjà été décrémenté ci-dessus pour la nouvelle quantité
                    pass
                
                panier_existant.quantite += quantite
                
                # Recalculer le sous-total avec la nouvelle quantité
                prix_actuel = get_prix_upsell_avec_compteur(article, commande.compteur)
                panier_existant.sous_total = prix_actuel * panier_existant.quantite
                panier_existant.save()
                
                print(f"📝 Article existant mis à jour: {article.nom} - nouvelle quantité: {panier_existant.quantite}")
            else:
                # Créer un nouveau panier
                prix_actuel = get_prix_upsell_avec_compteur(article, commande.compteur)
                sous_total = prix_actuel * quantite
                
                panier = Panier.objects.create(
                    commande=commande,
                    article=article,
                    quantite=quantite,
                    sous_total=sous_total
                )
                
                print(f"➕ Nouvel article ajouté: {article.nom} - quantité: {quantite}")
            
            # Recalculer le compteur upsell selon la logique d'operatConfirme
            paniers_upsell = commande.paniers.filter(article__isUpsell=True)
            total_quantite_upsell = sum(p.quantite for p in paniers_upsell)
            
            if total_quantite_upsell >= 2:
                commande.compteur = total_quantite_upsell - 1
            else:
                commande.compteur = 0
            
            # Recalculer le total de la commande
            commande.total_cmd = sum(p.sous_total for p in commande.paniers.all())
            commande.save()
            
            print(f"🎯 Compteur upsell mis à jour: {commande.compteur} (total upsell: {total_quantite_upsell})")
            
            return JsonResponse({
                'success': True,
                'message': f'Article {article.nom} ajouté avec succès'
            })
            
    except ValueError as e:
        return JsonResponse({'error': f'Quantité invalide: {str(e)}'}, status=400)
    except Exception as e:
        print(f"Erreur dans ajouter_article_commande: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Erreur interne: {str(e)}'}, status=500)

@login_required
def modifier_quantite_article(request, commande_id):
    """Modifier la quantité d'un article dans la commande"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Profil d\'opérateur logistique non trouvé.'}, status=403)
    
    try:
        from django.db import transaction
        
        with transaction.atomic():
            commande = Commande.objects.select_for_update().get(id=commande_id)
            
            # Vérifier que la commande est affectée à cet opérateur
            etat_livraison = commande.etats.filter(
                operateur=operateur,
                enum_etat__libelle='En cours de livraison',
                date_fin__isnull=True
            ).first()
            
            if not etat_livraison:
                return JsonResponse({'error': 'Cette commande n\'est pas affectée à votre livraison.'}, status=403)
            
            # Récupérer les paramètres
            panier_id = request.POST.get('panier_id')
            nouvelle_quantite = int(request.POST.get('quantite', 1))
            
            if not panier_id:
                return JsonResponse({'error': 'Panier manquant'}, status=400)
            
            if nouvelle_quantite <= 0:
                return JsonResponse({'error': 'Quantité invalide'}, status=400)
            
            # Récupérer le panier
            panier = Panier.objects.get(id=panier_id, commande=commande)
            ancienne_quantite = panier.quantite
            article = panier.article
            
            # Vérifier si la commande a déjà impacté le stock
            stock_deja_impacte = commande_a_impacte_stock(commande)
            
            if stock_deja_impacte:
                # Calculer la différence de quantité
                difference = nouvelle_quantite - ancienne_quantite
                
                if difference > 0:
                    # Augmentation de quantité - vérifier le stock disponible
                    if article.qte_disponible < difference:
                        return JsonResponse({
                            'error': f'Stock insuffisant pour {article.nom}. Stock disponible: {article.qte_disponible}, augmentation demandée: {difference}'
                        }, status=400)
                    
                    # Décrémenter le stock pour la différence
                    creer_mouvement_stock(
                        article=article,
                        quantite=difference,
                        type_mouvement='sortie',
                        commande=commande,
                        operateur=operateur,
                        commentaire=f'Augmentation de quantité pour la commande {commande.id_yz} par {operateur.nom_complet} (de {ancienne_quantite} à {nouvelle_quantite})'
                    )
                    
                    print(f"📦 Stock décrémenté: {article.nom} - {difference} unités (augmentation)")
                
                elif difference < 0:
                    # Diminution de quantité - remettre en stock
                    quantite_retour = abs(difference)
                    
                    creer_mouvement_stock(
                        article=article,
                        quantite=quantite_retour,
                        type_mouvement='retour_client',
                        commande=commande,
                        operateur=operateur,
                        commentaire=f'Diminution de quantité pour la commande {commande.id_yz} par {operateur.nom_complet} (de {ancienne_quantite} à {nouvelle_quantite})'
                    )
                    
                    print(f"📦 Stock incrémenté: {article.nom} - {quantite_retour} unités (diminution)")
                
                # Si difference == 0, pas de changement de stock
            
            # Mettre à jour la quantité du panier
            panier.quantite = nouvelle_quantite
            
            # Recalculer le sous-total
            prix_actuel = get_prix_upsell_avec_compteur(article, commande.compteur)
            panier.sous_total = prix_actuel * nouvelle_quantite
            panier.save()
            
            print(f"📝 Quantité modifiée: {article.nom} - {ancienne_quantite} → {nouvelle_quantite}")
            
            # Recalculer le compteur upsell selon la logique d'operatConfirme
            paniers_upsell = commande.paniers.filter(article__isUpsell=True)
            total_quantite_upsell = sum(p.quantite for p in paniers_upsell)
            
            if total_quantite_upsell >= 2:
                commande.compteur = total_quantite_upsell - 1
            else:
                commande.compteur = 0
            
            # Recalculer le total de la commande
            commande.total_cmd = sum(p.sous_total for p in commande.paniers.all())
            commande.save()
            
            print(f"🎯 Compteur upsell mis à jour: {commande.compteur} (total upsell: {total_quantite_upsell})")
            
            return JsonResponse({
                'success': True,
                'message': f'Quantité de {article.nom} modifiée avec succès'
            })
            
    except Panier.DoesNotExist:
        return JsonResponse({'error': 'Panier non trouvé'}, status=404)
    except ValueError:
        return JsonResponse({'error': 'Quantité invalide'}, status=400)
    except Exception as e:
        print(f"Erreur dans modifier_quantite_article: {str(e)}")
        return JsonResponse({'error': f'Erreur interne: {str(e)}'}, status=500)

@login_required
def supprimer_article_commande(request, commande_id):
    """Supprimer un article de la commande"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Profil d\'opérateur logistique non trouvé.'}, status=403)
    
    try:
        from django.db import transaction
        
        with transaction.atomic():
            commande = Commande.objects.select_for_update().get(id=commande_id)
            
            # Vérifier que la commande est affectée à cet opérateur
            etat_livraison = commande.etats.filter(
                operateur=operateur,
                enum_etat__libelle='En cours de livraison',
                date_fin__isnull=True
            ).first()
            
            if not etat_livraison:
                return JsonResponse({'error': 'Cette commande n\'est pas affectée à votre livraison.'}, status=403)
            
            # Récupérer les paramètres
            panier_id = request.POST.get('panier_id')
            
            if not panier_id:
                return JsonResponse({'error': 'Panier manquant'}, status=400)
            
            # Récupérer le panier
            panier = Panier.objects.get(id=panier_id, commande=commande)
            article = panier.article
            quantite_supprimee = panier.quantite
            
            # Vérifier si la commande a déjà impacté le stock
            stock_deja_impacte = commande_a_impacte_stock(commande)
            
            if stock_deja_impacte:
                # Remettre les articles en stock
                creer_mouvement_stock(
                    article=article,
                    quantite=quantite_supprimee,
                    type_mouvement='retour_client',
                    commande=commande,
                    operateur=operateur,
                    commentaire=f'Suppression d\'article lors de la modification de la commande {commande.id_yz} par {operateur.nom_complet}'
                )
                
                print(f"📦 Stock incrémenté: {article.nom} - {quantite_supprimee} unités (suppression)")
            
            # Supprimer le panier
            nom_article = article.nom
            panier.delete()
            
            print(f"🗑️ Article supprimé: {nom_article} - {quantite_supprimee} unités")
            
            # Recalculer le compteur upsell selon la logique d'operatConfirme
            paniers_upsell = commande.paniers.filter(article__isUpsell=True)
            total_quantite_upsell = sum(p.quantite for p in paniers_upsell)
            
            if total_quantite_upsell >= 2:
                commande.compteur = total_quantite_upsell - 1
            else:
                commande.compteur = 0
            
            # Recalculer le total de la commande
            commande.total_cmd = sum(p.sous_total for p in commande.paniers.all())
            commande.save()
            
            print(f"🎯 Compteur upsell mis à jour: {commande.compteur} (total upsell: {total_quantite_upsell})")
            
            return JsonResponse({
                'success': True,
                'message': f'Article {nom_article} supprimé avec succès'
            })
            
    except Panier.DoesNotExist:
        return JsonResponse({'error': 'Panier non trouvé'}, status=404)
    except Exception as e:
        print(f"Erreur dans supprimer_article_commande: {str(e)}")
        return JsonResponse({'error': f'Erreur interne: {str(e)}'}, status=500)

@login_required
def rafraichir_articles_commande(request, commande_id):
    """Rafraîchir la section des articles de la commande"""
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Profil d\'opérateur logistique non trouvé.'}, status=403)
    
    try:
        from commande.models import Commande
        
        commande = Commande.objects.select_related('client', 'ville', 'ville__region').prefetch_related('paniers__article').get(id=commande_id)
        
        # Vérifier que la commande est affectée à cet opérateur
        etat_livraison = commande.etats.filter(
            operateur=operateur,
            enum_etat__libelle='En cours de livraison',
            date_fin__isnull=True
        ).first()
        
        if not etat_livraison:
            return JsonResponse({'error': 'Cette commande ne vous est pas affectée.'}, status=403)
        
        # Recalculer le compteur upsell avec la logique d'operatConfirme
        from django.db.models import Sum
        total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
            total=Sum('quantite')
        )['total'] or 0
        
        # Logique d'operatConfirme : 0-1 unités → compteur = 0 | 2+ unités → compteur = total-1
        if total_quantite_upsell >= 2:
            commande.compteur = total_quantite_upsell - 1
        else:
            commande.compteur = 0
        
        commande.save()
        
        # Recalculer TOUS les articles de la commande avec le nouveau compteur
        if hasattr(commande, 'recalculer_totaux_upsell'):
            commande.recalculer_totaux_upsell()
        else:
            # Fallback si la méthode n'existe pas
            for p in commande.paniers.all():
                prix_recalcule = get_prix_upsell_avec_compteur(p.article, commande.compteur)
                p.sous_total = prix_recalcule * p.quantite
                p.save()
        
        # Recalculer le total de la commande
        total_commande = commande.paniers.aggregate(
            total=Sum('sous_total')
        )['total'] or 0
        commande.total_cmd = total_commande
        commande.save()
        
        # Générer le HTML avec le template
        from django.template.loader import render_to_string
        
        html = render_to_string('operatLogistic/partials/_articles_section.html', {
            'commande': commande
        }, request=request)
        
        return JsonResponse({
            'success': True,
            'html': html,
            'count': commande.paniers.count(),
            'total': float(commande.total_cmd),
            'compteur': commande.compteur
        })
        
    except Commande.DoesNotExist:
        return JsonResponse({'error': 'Commande non trouvée'}, status=404)
    except Exception as e:
        print(f"Erreur dans rafraichir_articles_commande: {str(e)}")
        return JsonResponse({'error': f'Erreur interne: {str(e)}'}, status=500)

@login_required
def changer_etat_sav(request, commande_id):
    """Changer l'état d'une commande via le SAV (Service Après-Vente)"""
    if request.method != 'POST':
        return redirect('operatLogistic:detail_commande', commande_id=commande_id)
    
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil d'opérateur logistique non trouvé.")
        return redirect('operatLogistic:detail_commande', commande_id=commande_id)
    
    try:
        from django.db import transaction
        
        with transaction.atomic():
            commande = Commande.objects.select_for_update().get(id=commande_id)
            
            # Vérifier que la commande est affectée à cet opérateur
            etat_livraison = commande.etats.filter(
                operateur=operateur,
                enum_etat__libelle='En cours de livraison',
                date_fin__isnull=True
            ).first()
            
            if not etat_livraison:
                messages.error(request, "Cette commande ne vous est pas affectée.")
                return redirect('operatLogistic:detail_commande', commande_id=commande_id)
            
            from commande.models import EnumEtatCmd, EtatCommande
            from django.utils import timezone
            from datetime import datetime
            
            nouvel_etat_libelle = request.POST.get('nouvel_etat')
            commentaire = request.POST.get('commentaire', '').strip()
            
            if not nouvel_etat_libelle or not commentaire:
                messages.error(request, "L'état et le commentaire sont obligatoires.")
                return redirect('operatLogistic:detail_commande', commande_id=commande_id)
            
            # Récupérer ou créer l'état cible
            nouvel_etat, created = EnumEtatCmd.objects.get_or_create(
                libelle=nouvel_etat_libelle,
                defaults={'ordre': 50, 'couleur': '#6366F1'}
            )
            
            # Terminer l'état actuel
            etat_livraison.terminer_etat(operateur)
            
            # Traitement spécifique selon le type d'action SAV
            if nouvel_etat_libelle == 'Reportée':
                # Gestion du report avec nouvelle date
                date_report_str = request.POST.get('date_report')
                if date_report_str:
                    try:
                        date_report = datetime.strptime(date_report_str, '%Y-%m-%d').date()
                        
                        # Créer ou mettre à jour l'envoi
                        envoi, created = Envoi.objects.get_or_create(
                            commande=commande,
                            defaults={
                                'operateur': operateur,
                                'date_livraison_prevue': date_report,
                                'status': 'reporte'
                            }
                        )
                        
                        if not created:
                            envoi.date_report = date_report
                            envoi.motif_report = commentaire
                            envoi.status = 'reporte'
                            envoi.save()
                        
                        commentaire += f" - Nouvelle date de livraison : {date_report.strftime('%d/%m/%Y')}"
                        
                    except ValueError:
                        messages.error(request, "Format de date invalide.")
                        return redirect('operatLogistic:detail_commande', commande_id=commande_id)
                else:
                    messages.error(request, "Date de report obligatoire.")
                    return redirect('operatLogistic:detail_commande', commande_id=commande_id)
            
            elif nouvel_etat_libelle == 'Livrée':
                # Marquer comme livrée
                envoi, created = Envoi.objects.get_or_create(
                    commande=commande,
                    defaults={
                        'operateur': operateur,
                        'date_livraison_prevue': timezone.now().date(),
                        'status': 'livre'
                    }
                )
                
                if not created:
                    envoi.status = 'livre'
                    envoi.date_livraison_effective = timezone.now()
                    envoi.save()
            
            elif nouvel_etat_libelle == 'Retournée':
                # Gestion de l'annulation avec gestion du stock
                type_annulation = request.POST.get('type_annulation')
                
                if not type_annulation:
                    messages.error(request, "Type d'annulation obligatoire.")
                    return redirect('operatLogistic:detail_commande', commande_id=commande_id)
                
                if type_annulation == 'bonne':
                    # Réincrémenter le stock pour tous les articles
                    for panier in commande.paniers.all():
                        panier.article.qte_disponible += panier.quantite
                        panier.article.save()
                        
                        # Créer un mouvement de stock pour tracer le retour
                        creer_mouvement_stock(
                            article=panier.article,
                            quantite=panier.quantite,  # Quantité positive pour le retour
                            type_mouvement='Retour Client',
                            commande=commande,
                            operateur=operateur,
                            commentaire=f"Retour suite à annulation de la commande {commande.id_yz} - Produit en bon état"
                        )
                    
                    commentaire += " - Stock réincrémenté (produits en bon état)."
                else:
                    commentaire += " - Stock non réincrémenté (produits endommagés)."
                
                # Remettre le compteur à zéro lors de l'annulation
                commande.compteur = 0
                commande.save()
            
            # Créer le nouvel état
            EtatCommande.objects.create(
                commande=commande,
                enum_etat=nouvel_etat,
                operateur=operateur,
                commentaire=commentaire
            )
            
            messages.success(request, f"État de la commande changé vers '{nouvel_etat_libelle}' avec succès.")
            
    except Commande.DoesNotExist:
        messages.error(request, "Commande non trouvée.")
    except Exception as e:
        print(f"Erreur dans changer_etat_sav: {str(e)}")
        messages.error(request, f"Erreur lors du changement d'état: {str(e)}")
    
    return redirect('operatLogistic:detail_commande', commande_id=commande_id)

@login_required
def diagnostiquer_compteur(request, commande_id):
    """
    Fonction pour diagnostiquer et corriger le compteur d'une commande
    """
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Profil d\'opérateur logistique non trouvé.'}, status=403)
    
    try:
        commande = Commande.objects.get(id=commande_id)
        
        # Vérifier que la commande est affectée à cet opérateur
        etat_livraison = commande.etats.filter(
            operateur=operateur,
            enum_etat__libelle='En cours de livraison',
            date_fin__isnull=True
        ).first()
        
        if not etat_livraison:
            return JsonResponse({'error': 'Cette commande ne vous est pas affectée.'}, status=403)
        
        # Diagnostiquer la situation actuelle
        articles_upsell = commande.paniers.filter(article__isUpsell=True)
        compteur_actuel = commande.compteur
        
        # Calculer la quantité totale d'articles upsell
        from django.db.models import Sum
        total_quantite_upsell = articles_upsell.aggregate(
            total=Sum('quantite')
        )['total'] or 0
        
        print(f"🔍 DIAGNOSTIC Commande {commande.id_yz}:")
        print(f"📊 Compteur actuel: {compteur_actuel}")
        print(f"📦 Articles upsell trouvés: {articles_upsell.count()}")
        print(f"🔢 Quantité totale d'articles upsell: {total_quantite_upsell}")
        
        if articles_upsell.exists():
            print("📋 Articles upsell dans la commande:")
            for panier in articles_upsell:
                print(f"  - {panier.article.nom} (Qté: {panier.quantite}, ID: {panier.article.id}, isUpsell: {panier.article.isUpsell})")
        
        # Déterminer le compteur correct selon la logique d'operatConfirme :
        # 0-1 unités upsell → compteur = 0
        # 2+ unités upsell → compteur = total_quantite_upsell - 1
        if total_quantite_upsell >= 2:
            compteur_correct = total_quantite_upsell - 1
        else:
            compteur_correct = 0
        
        print(f"✅ Compteur correct: {compteur_correct}")
        print("📖 Logique: 0-1 unités upsell → compteur=0 | 2+ unités upsell → compteur=total_quantité-1")
        
        # Corriger si nécessaire
        if compteur_actuel != compteur_correct:
            print(f"🔧 CORRECTION: {compteur_actuel} -> {compteur_correct}")
            commande.compteur = compteur_correct
            commande.save()
            
            # Recalculer tous les totaux
            if hasattr(commande, 'recalculer_totaux_upsell'):
                commande.recalculer_totaux_upsell()
            else:
                # Fallback si la méthode n'existe pas
                for p in commande.paniers.all():
                    prix_recalcule = get_prix_upsell_avec_compteur(p.article, commande.compteur)
                    p.sous_total = prix_recalcule * p.quantite
                    p.save()
                
                # Recalculer le total de la commande
                from django.db.models import Sum
                total_commande = commande.paniers.aggregate(
                    total=Sum('sous_total')
                )['total'] or 0
                commande.total_cmd = total_commande
                commande.save()
            
            # Retourner les nouvelles données
            return JsonResponse({
                'success': True,
                'message': f'Compteur corrigé de {compteur_actuel} vers {compteur_correct}',
                'ancien_compteur': compteur_actuel,
                'nouveau_compteur': compteur_correct,
                'total_commande': float(commande.total_cmd),
                'articles_upsell': articles_upsell.count(),
                'quantite_totale_upsell': total_quantite_upsell
            })
        else:
            return JsonResponse({
                'success': True,
                'message': 'Compteur déjà correct',
                'compteur': compteur_actuel,
                'articles_upsell': articles_upsell.count(),
                'quantite_totale_upsell': total_quantite_upsell
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def historique_stock_commande(request, commande_id):
    """Afficher l'historique des mouvements de stock pour une commande"""
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Profil d\'opérateur logistique non trouvé.'}, status=403)
    
    try:
        commande = Commande.objects.get(id=commande_id)
        
        # Vérifier que la commande est affectée à cet opérateur
        etat_livraison = commande.etats.filter(
            operateur=operateur,
            enum_etat__libelle='En cours de livraison',
            date_fin__isnull=True
        ).first()
        
        if not etat_livraison:
            return JsonResponse({'error': 'Cette commande n\'est pas affectée à votre livraison.'}, status=403)
        
        # Récupérer tous les mouvements de stock liés à cette commande
        from article.models import MouvementStock
        
        mouvements = MouvementStock.objects.filter(
            commande_associee=commande
        ).select_related('article', 'operateur').order_by('-date_mouvement')
        
        # Préparer les données pour le JSON
        mouvements_data = []
        for mouvement in mouvements:
            mouvements_data.append({
                'id': mouvement.id,
                'article_nom': mouvement.article.nom,
                'article_reference': mouvement.article.reference or '',
                'type_mouvement': mouvement.get_type_mouvement_display(),
                'quantite': mouvement.quantite,
                'qte_apres_mouvement': mouvement.qte_apres_mouvement,
                'date_mouvement': mouvement.date_mouvement.strftime('%d/%m/%Y %H:%M'),
                'commentaire': mouvement.commentaire or '',
                'operateur': mouvement.operateur.nom_complet if mouvement.operateur else 'Système'
            })
        
        return JsonResponse({
            'success': True,
            'commande_id': commande.id_yz,
            'mouvements': mouvements_data,
            'count': len(mouvements_data)
        })
        
    except Commande.DoesNotExist:
        return JsonResponse({'error': 'Commande non trouvée'}, status=404)
    except Exception as e:
        print(f"Erreur dans historique_stock_commande: {str(e)}")
        return JsonResponse({'error': f'Erreur interne: {str(e)}'}, status=500)
